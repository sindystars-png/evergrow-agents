#!/usr/bin/env python3
"""
Generate IRS-style Form 1040 Tax Return PDF Package from populated Excel workbook.
Reads from 1040_Tax_Return_2025.xlsx and produces a complete tax return PDF.
"""

import sys
import os
from openpyxl import load_workbook
from reportlab.lib.pagesizes import letter
from reportlab.lib.units import inch
from reportlab.lib.colors import black, white, HexColor
from reportlab.pdfgen import canvas
from reportlab.lib.enums import TA_LEFT, TA_RIGHT, TA_CENTER

# ── Constants ────────────────────────────────────────────────────────────────
PAGE_W, PAGE_H = letter  # 612 x 792
MARGIN_L = 0.5 * inch
MARGIN_R = 0.5 * inch
MARGIN_T = 0.5 * inch
MARGIN_B = 0.5 * inch
CONTENT_W = PAGE_W - MARGIN_L - MARGIN_R

IRS_DARK = HexColor('#1a1a1a')
IRS_GRAY = HexColor('#f0f0f0')
IRS_MED_GRAY = HexColor('#d0d0d0')
IRS_LINE_GRAY = HexColor('#999999')
IRS_BLUE = HexColor('#003366')
IRS_LIGHT_BLUE = HexColor('#e8f0fe')
SECTION_BG = HexColor('#1a3a5c')
AMOUNT_BG = HexColor('#f5f5f5')

FONT_TITLE = ('Helvetica-Bold', 14)
FONT_FORM_NAME = ('Helvetica-Bold', 11)
FONT_SECTION = ('Helvetica-Bold', 9)
FONT_LINE = ('Helvetica', 8)
FONT_LINE_BOLD = ('Helvetica-Bold', 8)
FONT_AMOUNT = ('Courier-Bold', 9)
FONT_SMALL = ('Helvetica', 6.5)
FONT_HEADER = ('Helvetica-Bold', 8)


def fmt_curr(val):
    if val is None or val == '' or val == 0:
        return '—'
    try:
        v = float(val)
        if v < 0:
            return f'({abs(v):,.0f})'
        return f'{v:,.0f}'
    except (ValueError, TypeError):
        return str(val)


def fmt_curr_dollar(val):
    if val is None or val == '' or val == 0:
        return '—'
    try:
        v = float(val)
        if v < 0:
            return f'(${abs(v):,.0f})'
        return f'${v:,.0f}'
    except (ValueError, TypeError):
        return str(val)


def safe_str(val):
    if val is None:
        return ''
    return str(val).strip()


def safe_num(val):
    if val is None or val == '':
        return 0
    try:
        return float(val)
    except (ValueError, TypeError):
        return 0


# ═══════════════════════════════════════════════════════════════════════════════
# DATA READER — Reads populated Excel workbook
# ═══════════════════════════════════════════════════════════════════════════════
class TaxData:
    def __init__(self, xlsx_path):
        self.wb = load_workbook(xlsx_path, data_only=True)

    def cell(self, sheet, ref):
        try:
            ws = self.wb[sheet]
            return ws[ref].value
        except Exception:
            return None

    def val(self, sheet, ref):
        return safe_num(self.cell(sheet, ref))

    def txt(self, sheet, ref):
        return safe_str(self.cell(sheet, ref))

    def filing_status_text(self):
        fs = self.val('Intake', 'C6')
        return {1: 'Single', 2: 'Married filing jointly', 3: 'Married filing separately',
                4: 'Head of household', 5: 'Qualifying surviving spouse'}.get(int(fs) if fs else 0, '')

    def taxpayer_name(self):
        first = self.txt('Intake', 'C14')
        mi = self.txt('Intake', 'C15')
        last = self.txt('Intake', 'C16')
        parts = [p for p in [first, mi, last] if p]
        return ' '.join(parts)

    def spouse_name(self):
        first = self.txt('Intake', 'E14')
        last = self.txt('Intake', 'E16')
        parts = [p for p in [first, last] if p]
        return ' '.join(parts)

    def ssn(self):
        return self.txt('Intake', 'C17')

    def spouse_ssn(self):
        return self.txt('Intake', 'E17')

    def address_line1(self):
        street = self.txt('Intake', 'C28')
        apt = self.txt('Intake', 'C29')
        if apt:
            return f'{street}, {apt}'
        return street

    def address_line2(self):
        city = self.txt('Intake', 'C30')
        state = self.txt('Intake', 'C31')
        zipcode = self.txt('Intake', 'C32')
        parts = []
        if city: parts.append(city)
        if state: parts.append(state)
        csz = ', '.join(parts)
        if zipcode: csz += f' {zipcode}'
        return csz


# ═══════════════════════════════════════════════════════════════════════════════
# IRS FORM RENDERER — Core drawing engine
# ═══════════════════════════════════════════════════════════════════════════════
class IRSFormRenderer:
    def __init__(self, c, data):
        self.c = c
        self.data = data
        self.y = PAGE_H - MARGIN_T

    def new_page(self):
        self.c.showPage()
        self.y = PAGE_H - MARGIN_T

    def _set_font(self, font_tuple):
        self.c.setFont(font_tuple[0], font_tuple[1])

    # ── Form Title Bar ────────────────────────────────────────────────────────
    def form_header(self, form_name, form_title, tax_year='2025', subtitle=None):
        y = self.y
        # Dark banner
        self.c.setFillColor(SECTION_BG)
        self.c.rect(MARGIN_L, y - 28, CONTENT_W, 28, fill=True, stroke=False)

        # Form name on left
        self.c.setFillColor(white)
        self._set_font(('Helvetica-Bold', 13))
        self.c.drawString(MARGIN_L + 8, y - 20, form_name)

        # Title centered
        self._set_font(('Helvetica-Bold', 10))
        title_w = self.c.stringWidth(form_title, 'Helvetica-Bold', 10)
        self.c.drawString(MARGIN_L + (CONTENT_W - title_w) / 2, y - 12, form_title)

        # Subtitle if present
        if subtitle:
            self._set_font(('Helvetica', 8))
            sub_w = self.c.stringWidth(subtitle, 'Helvetica', 8)
            self.c.drawString(MARGIN_L + (CONTENT_W - sub_w) / 2, y - 24, subtitle)

        # Tax year on right
        self._set_font(('Helvetica-Bold', 11))
        yr_w = self.c.stringWidth(tax_year, 'Helvetica-Bold', 11)
        self.c.drawString(MARGIN_L + CONTENT_W - yr_w - 8, y - 20, tax_year)

        self.c.setFillColor(black)
        self.y = y - 34

    # ── Section Header ────────────────────────────────────────────────────────
    def section_header(self, text):
        y = self.y
        self.c.setFillColor(HexColor('#2c5282'))
        self.c.rect(MARGIN_L, y - 16, CONTENT_W, 16, fill=True, stroke=False)
        self.c.setFillColor(white)
        self._set_font(FONT_SECTION)
        self.c.drawString(MARGIN_L + 6, y - 12, text)
        self.c.setFillColor(black)
        self.y = y - 19

    # ── Info Row (label + value, no line number) ──────────────────────────────
    def info_row(self, label, value, height=14):
        y = self.y
        # Light background
        self.c.setFillColor(IRS_GRAY)
        self.c.rect(MARGIN_L, y - height, CONTENT_W * 0.5, height, fill=True, stroke=False)
        self.c.setFillColor(black)
        self.c.setStrokeColor(IRS_MED_GRAY)
        self.c.rect(MARGIN_L, y - height, CONTENT_W, height, fill=False, stroke=True)

        self._set_font(FONT_LINE)
        self.c.drawString(MARGIN_L + 4, y - height + 3.5, label)

        self._set_font(FONT_AMOUNT)
        self.c.drawString(MARGIN_L + CONTENT_W * 0.52, y - height + 3.5, safe_str(value))

        self.c.setStrokeColor(black)
        self.y = y - height

    # ── Form Line (line_num | description ... | amount box) ───────────────────
    def form_line(self, line_num, description, amount, bold=False, is_total=False, double_line=False, show_dollar=True):
        y = self.y
        h = 16
        line_col_w = 28
        desc_w = CONTENT_W - line_col_w - 100
        amt_w = 100

        # Background for totals
        if is_total:
            self.c.setFillColor(HexColor('#e8f4e8'))
            self.c.rect(MARGIN_L, y - h, CONTENT_W, h, fill=True, stroke=False)
        elif double_line:
            self.c.setFillColor(HexColor('#fff3e0'))
            self.c.rect(MARGIN_L, y - h, CONTENT_W, h, fill=True, stroke=False)

        # Border
        self.c.setStrokeColor(IRS_MED_GRAY)
        self.c.setLineWidth(0.5)
        self.c.rect(MARGIN_L, y - h, CONTENT_W, h, fill=False, stroke=True)

        # Line number column
        self.c.setFillColor(HexColor('#e8edf3'))
        self.c.rect(MARGIN_L, y - h, line_col_w, h, fill=True, stroke=True)
        self.c.setFillColor(black)
        self._set_font(FONT_LINE_BOLD)
        num_w = self.c.stringWidth(str(line_num), 'Helvetica-Bold', 8)
        self.c.drawString(MARGIN_L + (line_col_w - num_w) / 2, y - h + 4.5, str(line_num))

        # Description
        font = FONT_LINE_BOLD if bold or is_total else FONT_LINE
        self._set_font(font)
        # Truncate if too long
        desc_text = description
        while self.c.stringWidth(desc_text, font[0], font[1]) > desc_w - 10 and len(desc_text) > 5:
            desc_text = desc_text[:-2]
        self.c.drawString(MARGIN_L + line_col_w + 4, y - h + 4.5, desc_text)

        # Dotted leader
        if not is_total:
            desc_end = MARGIN_L + line_col_w + 4 + self.c.stringWidth(desc_text, font[0], font[1]) + 4
            amt_start = MARGIN_L + CONTENT_W - amt_w - 2
            if amt_start - desc_end > 10:
                self.c.setStrokeColor(IRS_LINE_GRAY)
                self.c.setDash(1, 3)
                self.c.line(desc_end, y - h + 6, amt_start, y - h + 6)
                self.c.setDash()
                self.c.setStrokeColor(black)

        # Amount box
        amt_x = MARGIN_L + CONTENT_W - amt_w
        self.c.setFillColor(AMOUNT_BG if not is_total else HexColor('#d4edda'))
        self.c.rect(amt_x, y - h, amt_w, h, fill=True, stroke=True)
        self.c.setFillColor(black)

        self._set_font(FONT_AMOUNT)
        amt_text = fmt_curr_dollar(amount) if show_dollar else fmt_curr(amount)
        tw = self.c.stringWidth(amt_text, 'Courier-Bold', 9)
        self.c.drawString(amt_x + amt_w - tw - 6, y - h + 4.5, amt_text)

        # Double underline for key totals
        if double_line:
            self.c.setLineWidth(1.5)
            self.c.line(amt_x + 2, y - h + 1, amt_x + amt_w - 2, y - h + 1)
            self.c.setLineWidth(0.5)

        self.c.setStrokeColor(black)
        self.c.setFillColor(black)
        self.y = y - h

    # ── Spacer ────────────────────────────────────────────────────────────────
    def spacer(self, h=6):
        self.y -= h

    # ── Table header row ──────────────────────────────────────────────────────
    def table_header(self, columns, col_widths):
        y = self.y
        h = 16
        x = MARGIN_L
        self.c.setFillColor(HexColor('#2c5282'))
        self.c.rect(MARGIN_L, y - h, CONTENT_W, h, fill=True, stroke=False)
        self.c.setFillColor(white)
        self._set_font(('Helvetica-Bold', 7))
        for i, col in enumerate(columns):
            w = col_widths[i]
            tw = self.c.stringWidth(col, 'Helvetica-Bold', 7)
            self.c.drawString(x + (w - tw) / 2, y - h + 4.5, col)
            x += w
        self.c.setFillColor(black)
        self.y = y - h

    # ── Table data row ────────────────────────────────────────────────────────
    def table_row(self, values, col_widths, aligns=None, is_total=False):
        y = self.y
        h = 14
        x = MARGIN_L

        if is_total:
            self.c.setFillColor(HexColor('#e8f4e8'))
            self.c.rect(MARGIN_L, y - h, CONTENT_W, h, fill=True, stroke=False)

        self.c.setStrokeColor(IRS_MED_GRAY)
        self.c.rect(MARGIN_L, y - h, CONTENT_W, h, fill=False, stroke=True)
        self.c.setFillColor(black)

        font = FONT_LINE_BOLD if is_total else FONT_LINE
        self._set_font(font)

        for i, val in enumerate(values):
            w = col_widths[i]
            txt = safe_str(val) if not isinstance(val, (int, float)) else fmt_curr(val)
            align = (aligns[i] if aligns else 'left')
            if align == 'right':
                tw = self.c.stringWidth(txt, font[0], font[1])
                self.c.drawString(x + w - tw - 4, y - h + 3.5, txt)
            elif align == 'center':
                tw = self.c.stringWidth(txt, font[0], font[1])
                self.c.drawString(x + (w - tw) / 2, y - h + 3.5, txt)
            else:
                self.c.drawString(x + 4, y - h + 3.5, txt)
            # Column separator
            self.c.setStrokeColor(IRS_MED_GRAY)
            self.c.line(x + w, y - h, x + w, y)
            x += w

        self.c.setStrokeColor(black)
        self.y = y - h

    # ── Footer ────────────────────────────────────────────────────────────────
    def page_footer(self, form_name, page_num=None):
        self._set_font(FONT_SMALL)
        self.c.setFillColor(IRS_LINE_GRAY)
        self.c.drawString(MARGIN_L, MARGIN_B - 8, f'{form_name}  |  Tax Year 2025')
        if page_num:
            pg_text = f'Page {page_num}'
            tw = self.c.stringWidth(pg_text, 'Helvetica', 6.5)
            self.c.drawString(PAGE_W - MARGIN_R - tw, MARGIN_B - 8, pg_text)
        self.c.setFillColor(black)

    # ── Check if space remains ────────────────────────────────────────────────
    def space_left(self):
        return self.y - MARGIN_B


# ═══════════════════════════════════════════════════════════════════════════════
# FORM BUILDERS
# ═══════════════════════════════════════════════════════════════════════════════

def build_form_1040_page1(r, d):
    r.form_header('Form 1040', 'U.S. Individual Income Tax Return', '2025',
                  'Department of the Treasury — Internal Revenue Service')
    r.spacer(4)

    # Taxpayer info section
    r.section_header('TAXPAYER INFORMATION')
    r.info_row('Name', d.taxpayer_name())
    if d.spouse_name():
        r.info_row('Spouse Name', d.spouse_name())
    r.info_row('Social Security Number', d.ssn())
    r.info_row('Address', d.address_line1())
    r.info_row('City, State, ZIP', d.address_line2())
    r.info_row('Filing Status', d.filing_status_text())
    r.spacer(6)

    # Income section
    r.section_header('INCOME')
    lines = [
        ('1a', 'Wages, salaries, tips (W-2, Box 1)', 'Form 1040', 'D5'),
        ('1z', 'Add lines 1a through 1c', 'Form 1040', 'D8'),
        ('2a', 'Tax-exempt interest', 'Form 1040', 'D9'),
        ('2b', 'Taxable interest', 'Form 1040', 'D10'),
        ('3a', 'Qualified dividends', 'Form 1040', 'D11'),
        ('3b', 'Ordinary dividends', 'Form 1040', 'D12'),
        ('4a', 'IRA distributions', 'Form 1040', 'D13'),
        ('4b', 'Taxable amount', 'Form 1040', 'D14'),
        ('5a', 'Pensions and annuities', 'Form 1040', 'D15'),
        ('5b', 'Taxable amount', 'Form 1040', 'D16'),
        ('6a', 'Social Security benefits', 'Form 1040', 'D17'),
        ('6b', 'Taxable amount', 'Form 1040', 'D18'),
        ('7', 'Capital gain or (loss)', 'Form 1040', 'D19'),
        ('8', 'Other income from Schedule 1, line 10', 'Form 1040', 'D20'),
    ]
    for ln, desc, sheet, ref in lines:
        r.form_line(ln, desc, d.val(sheet, ref))

    r.form_line('9', 'Total income', d.val('Form 1040', 'D21'), bold=True, is_total=True)
    r.spacer(4)

    # AGI
    r.section_header('ADJUSTED GROSS INCOME')
    r.form_line('10', 'Adjustments to income from Schedule 1, line 26', d.val('Form 1040', 'D24'))
    r.form_line('11', 'Adjusted gross income (AGI)', d.val('Form 1040', 'D25'), bold=True, is_total=True)

    r.page_footer('Form 1040', 1)


def build_form_1040_page2(r, d):
    r.form_header('Form 1040', 'U.S. Individual Income Tax Return (continued)', '2025')
    r.spacer(4)

    r.section_header('DEDUCTIONS & TAXABLE INCOME')
    r.form_line('12', 'Standard deduction or itemized deductions', d.val('Form 1040', 'D28'))
    r.form_line('13', 'Qualified business income deduction', d.val('Form 1040', 'D29'))
    r.form_line('14', 'Total deductions', d.val('Form 1040', 'D30'))
    r.form_line('15', 'Taxable income', d.val('Form 1040', 'D31'), bold=True, is_total=True)
    r.spacer(4)

    r.section_header('TAX COMPUTATION')
    r.form_line('16', 'Tax', d.val('Form 1040', 'D34'))
    r.form_line('17', 'Amount from Schedule 2, line 21', d.val('Form 1040', 'D35'))
    r.form_line('18', 'Add lines 16 and 17', d.val('Form 1040', 'D36'))
    r.form_line('19', 'Nonrefundable credits (Schedule 3, line 8)', d.val('Form 1040', 'D37'))
    r.form_line('20', 'Subtract line 19 from line 18', d.val('Form 1040', 'D38'))
    r.form_line('21', 'Other taxes from Schedule 2', d.val('Form 1040', 'D39'))
    r.form_line('22', 'Total tax', d.val('Form 1040', 'D40'), bold=True, is_total=True)
    r.spacer(4)

    r.section_header('PAYMENTS')
    r.form_line('23', 'Federal income tax withheld', d.val('Form 1040', 'D43'))
    r.form_line('24', 'Estimated tax payments and amount applied from prior year', d.val('Form 1040', 'D44'))
    r.form_line('25', 'Earned income credit', d.val('Form 1040', 'D45'))
    r.form_line('26', 'Additional child tax credit', d.val('Form 1040', 'D46'))
    r.form_line('27', 'American opportunity credit', d.val('Form 1040', 'D47'))
    r.form_line('33', 'Total payments', d.val('Form 1040', 'D53'), bold=True, is_total=True)
    r.spacer(4)

    r.section_header('REFUND OR AMOUNT OWED')
    refund = d.val('Form 1040', 'D56')
    owed = d.val('Form 1040', 'D58')
    r.form_line('34', 'Overpayment', refund, double_line=(refund > 0))
    r.form_line('35a', 'Refunded to you', refund)
    r.form_line('37', 'Amount you owe', owed, bold=True, double_line=(owed > 0))

    r.page_footer('Form 1040', 2)


def build_schedule_1(r, d):
    r.form_header('Schedule 1', 'Additional Income and Adjustments to Income', '2025',
                  '(Form 1040)')
    r.spacer(4)

    r.section_header('PART I — ADDITIONAL INCOME')
    s1_lines = [
        ('1', 'Taxable refunds, credits, or offsets of state and local income taxes'),
        ('2a', 'Alimony received'),
        ('3', 'Business income or (loss) (Schedule C)'),
        ('4', 'Other gains or (losses) (Form 4797)'),
        ('5', 'Rental real estate, royalties, partnerships, S corporations (Schedule E)'),
        ('6', 'Farm income or (loss) (Schedule F)'),
        ('7', 'Unemployment compensation'),
        ('8b', 'Gambling income'),
        ('8z', 'Other income'),
    ]
    # Read from Schedule 1 tab
    row = 4  # Starting row in Schedule 1 sheet
    for ln, desc in s1_lines:
        r.form_line(ln, desc, d.val('Schedule 1', f'D{row}'))
        row += 1

    r.form_line('10', 'Total additional income', d.val('Schedule 1', 'D17'), bold=True, is_total=True)
    r.spacer(6)

    r.section_header('PART II — ADJUSTMENTS TO INCOME')
    adj_lines = [
        ('11', 'Educator expenses'),
        ('12', 'Business expenses of reservists, etc.'),
        ('13', 'HSA deduction'),
        ('14', 'Moving expenses (military only)'),
        ('15', 'Deductible part of self-employment tax'),
        ('16', 'Self-employed SEP, SIMPLE, qualified plans'),
        ('17', 'Self-employed health insurance deduction'),
        ('18', 'Penalty on early withdrawal of savings'),
        ('19', 'IRA deduction'),
        ('20', 'Student loan interest deduction'),
        ('22', 'Other adjustments'),
    ]
    row = 33
    for ln, desc in adj_lines:
        r.form_line(ln, desc, d.val('Schedule 1', f'D{row}'))
        row += 1

    r.form_line('26', 'Total adjustments to income', d.val('Schedule 1', 'D45'), bold=True, is_total=True)
    r.page_footer('Schedule 1 (Form 1040)')


def build_schedule_2(r, d):
    r.form_header('Schedule 2', 'Additional Taxes', '2025', '(Form 1040)')
    r.spacer(4)

    r.section_header('PART I — TAX')
    r.form_line('1', 'AMT (Form 6251)', d.val('Schedule 2', 'D4'))
    r.form_line('2', 'Excess premium tax credit repayment', d.val('Schedule 2', 'D5'))
    r.form_line('3', 'Add lines 1 and 2', d.val('Schedule 2', 'D6'), is_total=True)
    r.spacer(6)

    r.section_header('PART II — OTHER TAXES')
    r.form_line('6', 'Self-employment tax (Schedule SE)', d.val('Schedule 2', 'D9'))
    r.form_line('8', 'Additional Medicare Tax (Form 8959)', d.val('Schedule 2', 'D10'))
    r.form_line('9', 'Net investment income tax (Form 8960)', d.val('Schedule 2', 'D11'))
    r.form_line('10', 'Uncollected SS/Medicare on tips', d.val('Schedule 2', 'D12'))
    r.form_line('11', 'Additional tax on IRAs/retirement', d.val('Schedule 2', 'D13'))
    r.form_line('12', 'Additional tax on HSA distributions', d.val('Schedule 2', 'D14'))
    r.form_line('21', 'Total other taxes', d.val('Schedule 2', 'D16'), bold=True, is_total=True)

    r.page_footer('Schedule 2 (Form 1040)')


def build_schedule_3(r, d):
    r.form_header('Schedule 3', 'Additional Credits and Payments', '2025', '(Form 1040)')
    r.spacer(4)

    r.section_header('PART I — NONREFUNDABLE CREDITS')
    r.form_line('1', 'Foreign tax credit (Form 1116)', d.val('Schedule 3', 'D4'))
    r.form_line('2', 'Child and dependent care credit (Form 2441)', d.val('Schedule 3', 'D5'))
    r.form_line('3', 'Education credits (Form 8863)', d.val('Schedule 3', 'D6'))
    r.form_line('4', 'Retirement savings credit (Form 8880)', d.val('Schedule 3', 'D7'))
    r.form_line('5a', 'Residential energy credit', d.val('Schedule 3', 'D8'))
    r.form_line('5b', 'EV credit', d.val('Schedule 3', 'D9'))
    r.form_line('7', 'Other nonrefundable credits', d.val('Schedule 3', 'D10'))
    r.form_line('8', 'Total nonrefundable credits', d.val('Schedule 3', 'D11'), bold=True, is_total=True)
    r.spacer(6)

    r.section_header('PART II — OTHER PAYMENTS AND REFUNDABLE CREDITS')
    r.form_line('9', 'Net premium tax credit (Form 8962)', d.val('Schedule 3', 'D18'))
    r.form_line('10', 'Amount paid with extension request', d.val('Schedule 3', 'D19'))
    r.form_line('11', 'Excess social security tax withheld', d.val('Schedule 3', 'D20'))
    r.form_line('13', 'Other refundable credits', d.val('Schedule 3', 'D21'))
    r.form_line('15', 'Total other payments and refundable credits', d.val('Schedule 3', 'D22'), bold=True, is_total=True)

    r.page_footer('Schedule 3 (Form 1040)')


def build_schedule_a(r, d):
    r.form_header('Schedule A', 'Itemized Deductions', '2025', '(Form 1040)')
    r.spacer(4)

    r.section_header('MEDICAL AND DENTAL EXPENSES')
    r.form_line('1', 'Medical and dental expenses', d.val('Schedule A', 'D4'))
    r.form_line('2', 'Enter AGI from Form 1040, line 11', d.val('Schedule A', 'D5'))
    r.form_line('3', 'Multiply line 2 by 7.5% (0.075)', d.val('Schedule A', 'D6'))
    r.form_line('4', 'Subtract line 3 from line 1 (if more than zero)', d.val('Schedule A', 'D7'), is_total=True)
    r.spacer(4)

    r.section_header('TAXES YOU PAID')
    r.form_line('5a', 'State and local income taxes or general sales taxes', d.val('Schedule A', 'D10'))
    r.form_line('5b', 'State and local real estate taxes', d.val('Schedule A', 'D11'))
    r.form_line('5c', 'State and local personal property taxes', d.val('Schedule A', 'D12'))
    r.form_line('5d', 'Total SALT (limited to $10,000)', d.val('Schedule A', 'D13'), is_total=True)
    r.spacer(4)

    r.section_header('INTEREST YOU PAID')
    r.form_line('8a', 'Home mortgage interest (Form 1098)', d.val('Schedule A', 'D16'))
    r.form_line('8d', 'Mortgage insurance premiums', d.val('Schedule A', 'D17'))
    r.form_line('9', 'Investment interest expense', d.val('Schedule A', 'D18'))
    r.form_line('10', 'Total interest you paid', d.val('Schedule A', 'D19'), is_total=True)
    r.spacer(4)

    r.section_header('GIFTS TO CHARITY')
    r.form_line('11', 'Gifts by cash or check', d.val('Schedule A', 'D22'))
    r.form_line('12', 'Other than cash or check', d.val('Schedule A', 'D23'))
    r.form_line('13', 'Carryover from prior year', d.val('Schedule A', 'D24'))
    r.form_line('14', 'Total gifts to charity', d.val('Schedule A', 'D25'), is_total=True)
    r.spacer(4)

    r.section_header('OTHER ITEMIZED DEDUCTIONS')
    r.form_line('15', 'Casualty and theft loss(es)', d.val('Schedule A', 'D28'))
    r.form_line('16', 'Gambling losses', d.val('Schedule A', 'D29'))
    r.form_line('17', 'Other itemized deductions', d.val('Schedule A', 'D30'))
    r.spacer(4)

    r.form_line('18', 'TOTAL ITEMIZED DEDUCTIONS', d.val('Schedule A', 'D32'), bold=True, is_total=True, double_line=True)

    r.page_footer('Schedule A (Form 1040)')


def build_schedule_b(r, d):
    r.form_header('Schedule B', 'Interest and Ordinary Dividends', '2025', '(Form 1040)')
    r.spacer(4)

    r.section_header('PART I — INTEREST')
    # Read interest entries from Intake
    wb = d.wb
    ws = wb['Intake']
    has_int = False
    for row in range(61, 69):  # Approximate interest rows
        payer = ws.cell(row=row, column=2).value
        amt = ws.cell(row=row, column=4).value
        if payer and amt:
            has_int = True
            r.form_line('', safe_str(payer), safe_num(amt))
    if not has_int:
        r.form_line('', '(No interest income reported)', 0)

    r.form_line('4', 'Total interest', d.val('Schedule B', 'D6'), bold=True, is_total=True)
    r.spacer(6)

    r.section_header('PART II — ORDINARY DIVIDENDS')
    has_div = False
    for row in range(76, 84):
        payer = ws.cell(row=row, column=2).value
        amt = ws.cell(row=row, column=3).value
        if payer and amt:
            has_div = True
            r.form_line('', safe_str(payer), safe_num(amt))
    if not has_div:
        r.form_line('', '(No dividend income reported)', 0)

    r.form_line('6', 'Total ordinary dividends', d.val('Schedule B', 'D10'), bold=True, is_total=True)
    r.spacer(6)

    r.section_header('PART III — FOREIGN ACCOUNTS AND TRUSTS')
    r.form_line('7a', 'Foreign accounts or foreign trusts?', d.txt('Schedule B', 'D13'), show_dollar=False)
    r.form_line('8', 'FBAR (FinCEN Form 114) required?', d.txt('Schedule B', 'D15'), show_dollar=False)

    r.page_footer('Schedule B (Form 1040)')


def build_schedule_c(r, d):
    r.form_header('Schedule C', 'Profit or Loss From Business', '2025',
                  '(Form 1040) — Sole Proprietorship')
    r.spacer(4)

    r.section_header('BUSINESS INFORMATION')
    r.info_row('Business Name', d.txt('Schedule C', 'D5'))
    r.info_row('Business EIN', d.txt('Schedule C', 'D6'))
    r.info_row('Principal Business Code', d.txt('Schedule C', 'D4'))
    r.info_row('Accounting Method', d.txt('Schedule C', 'D8'))
    r.spacer(4)

    r.section_header('INCOME')
    r.form_line('1', 'Gross receipts or sales', d.val('Schedule C', 'D11'))
    r.form_line('2', 'Returns and allowances', d.val('Schedule C', 'D12'))
    r.form_line('4', 'Cost of goods sold', d.val('Schedule C', 'D13'))
    r.form_line('5', 'Gross profit', d.val('Schedule C', 'D14'), is_total=True)
    r.spacer(4)

    r.section_header('EXPENSES')
    expense_lines = [
        ('8', 'Advertising'), ('9', 'Car and truck expenses'),
        ('10', 'Commissions and fees'), ('11', 'Contract labor'),
        ('13', 'Depreciation'), ('14', 'Employee benefit programs'),
        ('15', 'Insurance'), ('16a', 'Mortgage interest'),
        ('16b', 'Other interest'), ('17', 'Legal and professional services'),
        ('18', 'Office expense'), ('20a', 'Rent — vehicles/machinery'),
        ('20b', 'Rent — other'), ('21', 'Repairs and maintenance'),
        ('22', 'Supplies'), ('23', 'Taxes and licenses'),
        ('24a', 'Travel'), ('24b', 'Meals (50%)'),
        ('25', 'Utilities'), ('26', 'Wages'),
        ('27a', 'Other expenses'),
    ]
    row = 17
    for ln, desc in expense_lines:
        val = d.val('Schedule C', f'D{row}')
        if val != 0:
            r.form_line(ln, desc, val)
        row += 1
    r.form_line('28', 'Total expenses', d.val('Schedule C', f'D{row}'), is_total=True)
    r.spacer(4)

    r.form_line('30', 'Business use of home', d.val('Schedule C', f'D{row+2}'))
    r.form_line('31', 'Net profit or (loss)', d.val('Schedule C', 'D45'), bold=True, is_total=True, double_line=True)

    r.page_footer('Schedule C (Form 1040)')


def build_schedule_d(r, d):
    r.form_header('Schedule D', 'Capital Gains and Losses', '2025', '(Form 1040)')
    r.spacer(4)

    r.section_header('PART I — SHORT-TERM CAPITAL GAINS AND LOSSES')
    r.form_line('7', 'Net short-term gain (loss) from Form 8949', d.val('Schedule D', 'D4'))
    r.form_line('8', 'Short-term gain from installment sales and K-1s', d.val('Schedule D', 'D5'))
    r.form_line('9', 'Short-term capital loss carryover', d.val('Schedule D', 'D6'))
    r.form_line('10', 'Net short-term capital gain or (loss)', d.val('Schedule D', 'D7'), is_total=True)
    r.spacer(6)

    r.section_header('PART II — LONG-TERM CAPITAL GAINS AND LOSSES')
    r.form_line('15', 'Net long-term gain (loss) from Form 8949', d.val('Schedule D', 'D10'))
    r.form_line('16', 'Long-term gain from K-1s and 1099-DIV', d.val('Schedule D', 'D11'))
    r.form_line('17', 'Long-term capital loss carryover', d.val('Schedule D', 'D12'))
    r.form_line('18', 'Net long-term capital gain or (loss)', d.val('Schedule D', 'D13'), is_total=True)
    r.spacer(6)

    r.section_header('PART III — SUMMARY')
    r.form_line('19', 'Combine short-term and long-term', d.val('Schedule D', 'D16'))
    r.form_line('21', 'Net capital gain or (loss)', d.val('Schedule D', 'D50'), bold=True, is_total=True, double_line=True)

    r.page_footer('Schedule D (Form 1040)')


def build_form_8949(r, d):
    r.form_header('Form 8949', 'Sales and Other Dispositions of Capital Assets', '2025')
    r.spacer(4)

    wb = d.wb
    ws = wb['Form 8949']

    # Part I — Short-term
    r.section_header('PART I — SHORT-TERM (Assets held one year or less)')
    cols = ['Description', 'Date Acquired', 'Date Sold', 'Proceeds', 'Cost Basis', 'Gain/(Loss)']
    widths = [160, 75, 75, 80, 80, 80]
    aligns = ['left', 'center', 'center', 'right', 'right', 'right']
    r.table_header(cols, widths)

    has_st = False
    for row in range(6, 21):
        desc = ws.cell(row=row, column=2).value
        if desc:
            has_st = True
            vals = [
                safe_str(desc),
                safe_str(ws.cell(row=row, column=3).value),
                safe_str(ws.cell(row=row, column=4).value),
                safe_num(ws.cell(row=row, column=5).value),
                safe_num(ws.cell(row=row, column=6).value),
                safe_num(ws.cell(row=row, column=7).value),
            ]
            r.table_row(vals, widths, aligns)
    if not has_st:
        r.table_row(['(No short-term transactions)', '', '', '', '', ''], widths, aligns)

    r.table_row(['TOTAL SHORT-TERM', '', '', fmt_curr(d.val('Form 8949', 'E29')),
                 fmt_curr(d.val('Form 8949', 'F29')), fmt_curr(d.val('Form 8949', 'D30'))],
                widths, aligns, is_total=True)
    r.spacer(8)

    # Part II — Long-term
    r.section_header('PART II — LONG-TERM (Assets held more than one year)')
    r.table_header(cols, widths)

    has_lt = False
    for row in range(35, 50):
        desc = ws.cell(row=row, column=2).value
        if desc:
            has_lt = True
            vals = [
                safe_str(desc),
                safe_str(ws.cell(row=row, column=3).value),
                safe_str(ws.cell(row=row, column=4).value),
                safe_num(ws.cell(row=row, column=5).value),
                safe_num(ws.cell(row=row, column=6).value),
                safe_num(ws.cell(row=row, column=7).value),
            ]
            r.table_row(vals, widths, aligns)
    if not has_lt:
        r.table_row(['(No long-term transactions)', '', '', '', '', ''], widths, aligns)

    r.table_row(['TOTAL LONG-TERM', '', '', fmt_curr(d.val('Form 8949', 'E59')),
                 fmt_curr(d.val('Form 8949', 'F59')), fmt_curr(d.val('Form 8949', 'D60'))],
                widths, aligns, is_total=True)

    r.page_footer('Form 8949')


def build_schedule_e(r, d):
    r.form_header('Schedule E', 'Supplemental Income and Loss', '2025',
                  '(Form 1040) — Rental Real Estate, Royalties, Partnerships, S Corporations')
    r.spacer(4)

    r.section_header('PART I — INCOME OR LOSS FROM RENTAL REAL ESTATE AND ROYALTIES')
    cols = ['', 'Property 1', 'Property 2', 'Property 3']
    widths = [200, 100, 100, 100]  # roughly fills 550
    # Simplified: show select lines
    items = [
        ('3', 'Rents received', 4),
        ('12', 'Mortgage interest', 4+8),
        ('16', 'Taxes', 4+12),
        ('18', 'Depreciation', 4+14),
        ('20', 'Total expenses', None),
        ('21', 'Net income/(loss)', None),
    ]

    for ln, desc, base_row in items:
        if base_row:
            vals = [f'{ln}  {desc}',
                    fmt_curr(d.val('Schedule E', f'D{base_row}')),
                    fmt_curr(d.val('Schedule E', f'E{base_row}')),
                    fmt_curr(d.val('Schedule E', f'F{base_row}'))]
        else:
            row_map = {'20': None, '21': None}  # compute from sheet
            vals = [f'{ln}  {desc}', '', '', '']
        aligns_e = ['left', 'right', 'right', 'right']
        is_tot = ln in ('20', '21')
        r.table_row(vals, widths, aligns_e, is_total=is_tot)
    r.spacer(6)

    r.section_header('PART II — INCOME OR LOSS FROM PARTNERSHIPS AND S CORPORATIONS')
    r.form_line('32', 'Total K-1 ordinary income', d.val('Schedule E', 'D35'))
    r.form_line('33', 'Total K-1 rental income', d.val('Schedule E', 'D36'))
    r.spacer(4)
    r.form_line('41', 'Total Schedule E income (loss)', d.val('Schedule E', 'D50'), bold=True, is_total=True, double_line=True)

    r.page_footer('Schedule E (Form 1040)')


def build_schedule_se(r, d):
    r.form_header('Schedule SE', 'Self-Employment Tax', '2025', '(Form 1040)')
    r.spacer(4)

    r.section_header('SELF-EMPLOYMENT TAX CALCULATION')
    se_lines = [
        ('1', 'Net farm profit or (loss)', 'D4'),
        ('2', 'Net profit or (loss) from Schedule C', 'D5'),
        ('3', 'Combine lines 1 and 2', 'D6'),
        ('4', 'Multiply line 3 by 92.35%', 'D7'),
        ('5', 'Social Security wage base ($176,100)', 'D8'),
        ('6', 'Total Social Security wages from W-2', 'D9'),
        ('7', 'Remaining wage room', 'D10'),
        ('8', 'Social Security portion (12.4%)', 'D11'),
        ('9', 'Medicare portion (2.9%)', 'D12'),
    ]
    for ln, desc, ref in se_lines:
        r.form_line(ln, desc, d.val('Schedule SE', ref))

    r.form_line('10', 'Total self-employment tax', d.val('Schedule SE', 'D13'), bold=True, is_total=True, double_line=True)
    r.spacer(4)
    r.form_line('11', 'Deductible part of SE tax (50%)', d.val('Schedule SE', 'D20'), is_total=True)

    r.page_footer('Schedule SE (Form 1040)')


def build_form_8938(r, d):
    r.form_header('Form 8938', 'Statement of Specified Foreign Financial Assets', '2025',
                  'FATCA — Foreign Account Tax Compliance Act')
    r.spacer(4)

    r.section_header('FILING THRESHOLD DETERMINATION')
    r.info_row('Filing Status', d.filing_status_text())
    r.info_row('Maximum aggregate value during year', fmt_curr_dollar(d.val('Form 8938', 'D6')))
    r.info_row('Year-end aggregate value', fmt_curr_dollar(d.val('Form 8938', 'D7')))
    r.info_row('Filing threshold (year-end)', fmt_curr_dollar(d.val('Form 8938', 'D8')))
    r.info_row('Filing threshold (any time)', fmt_curr_dollar(d.val('Form 8938', 'D9')))
    r.info_row('FORM 8938 REQUIRED?', d.txt('Form 8938', 'D10'))
    r.spacer(6)

    r.section_header('FOREIGN DEPOSIT AND CUSTODIAL ACCOUNTS')
    cols = ['Account Type', 'Institution', 'Country', 'Max Value', 'Year-End Value']
    widths = [100, 130, 80, 100, 100]  # ~510
    aligns = ['left', 'left', 'left', 'right', 'right']
    r.table_header(cols, widths)

    wb = d.wb
    ws = wb['Form 8938']
    has_accts = False
    for row in range(14, 24):
        acct_type = ws.cell(row=row, column=2).value
        if acct_type:
            has_accts = True
            vals = [
                safe_str(acct_type),
                safe_str(ws.cell(row=row, column=3).value),
                safe_str(ws.cell(row=row, column=4).value),
                safe_num(ws.cell(row=row, column=5).value),
                safe_num(ws.cell(row=row, column=6).value),
            ]
            r.table_row(vals, widths, aligns)
    if not has_accts:
        r.table_row(['(No foreign accounts reported)', '', '', '', ''], widths, aligns)

    r.page_footer('Form 8938')


def build_form_8621(r, d):
    r.form_header('Form 8621', 'Return by a Shareholder of a PFIC', '2025',
                  'Passive Foreign Investment Company')
    r.spacer(4)

    r.section_header('PFIC HOLDINGS SUMMARY')
    cols = ['PFIC Name', 'Country', 'EIN/Ref ID', 'Fair Market Value']
    widths = [160, 100, 120, 120]
    aligns = ['left', 'left', 'left', 'right']
    r.table_header(cols, widths)

    wb = d.wb
    ws = wb['Form 8621']
    has_pfic = False
    for row in range(5, 10):
        name = ws.cell(row=row, column=2).value
        if name:
            has_pfic = True
            vals = [
                safe_str(name),
                safe_str(ws.cell(row=row, column=3).value),
                safe_str(ws.cell(row=row, column=4).value),
                safe_num(ws.cell(row=row, column=5).value),
            ]
            r.table_row(vals, widths, aligns)
    if not has_pfic:
        r.table_row(['(No PFICs reported)', '', '', ''], widths, aligns)

    r.spacer(8)
    r.section_header('ELECTION AND INCOME DETAILS')
    cols2 = ['PFIC Name', 'Election Type', 'Ordinary Income', 'Net Capital Gain']
    r.table_header(cols2, widths)
    for row in range(14, 19):
        name = ws.cell(row=row, column=2).value
        if name:
            vals = [
                safe_str(name),
                safe_str(ws.cell(row=row, column=3).value),
                safe_num(ws.cell(row=row, column=4).value),
                safe_num(ws.cell(row=row, column=5).value),
            ]
            r.table_row(vals, widths, aligns)

    r.page_footer('Form 8621')


def build_form_3520(r, d):
    r.form_header('Form 3520', 'Annual Return to Report Transactions', '2025',
                  'With Foreign Trusts and Receipt of Certain Foreign Gifts')
    r.spacer(4)

    r.section_header('PART IV — U.S. PERSONS RECEIVING FOREIGN GIFTS')
    r.form_line('', 'Total gifts received from foreign individuals', d.val('Form 3520', 'D5'))
    r.form_line('', 'Total gifts received from foreign corporations', d.val('Form 3520', 'D6'))
    r.info_row('Form 3520 required for gifts?', d.txt('Form 3520', 'D7'))
    r.spacer(6)

    r.section_header('PART III — U.S. PERSONS WITH FOREIGN TRUSTS')
    r.info_row('Name of foreign trust', d.txt('Form 3520', 'D10'))
    r.info_row('Country of trust', d.txt('Form 3520', 'D11'))
    r.info_row('EIN or Reference ID', d.txt('Form 3520', 'D12'))
    r.form_line('', 'Distributions received from trust', d.val('Form 3520', 'D13'))
    r.form_line('', 'Transfers to trust during year', d.val('Form 3520', 'D14'))
    r.info_row('Form 3520 required for trust?', d.txt('Form 3520', 'D16'))

    r.page_footer('Form 3520')


# ═══════════════════════════════════════════════════════════════════════════════
# COVER PAGE
# ═══════════════════════════════════════════════════════════════════════════════
def build_cover_page(r, d):
    c = r.c
    y = PAGE_H - 1.5 * inch

    # Title block
    c.setFillColor(SECTION_BG)
    c.rect(MARGIN_L, y - 80, CONTENT_W, 80, fill=True, stroke=False)
    c.setFillColor(white)
    c.setFont('Helvetica-Bold', 22)
    c.drawCentredString(PAGE_W / 2, y - 30, 'Individual Income Tax Return')
    c.setFont('Helvetica', 14)
    c.drawCentredString(PAGE_W / 2, y - 52, 'Form 1040 — Tax Year 2025')
    c.setFont('Helvetica', 10)
    c.drawCentredString(PAGE_W / 2, y - 70, 'Prepared for filing with the Internal Revenue Service')
    c.setFillColor(black)
    y -= 110

    # Taxpayer info box
    c.setStrokeColor(IRS_MED_GRAY)
    c.setLineWidth(1)
    c.rect(MARGIN_L + 40, y - 120, CONTENT_W - 80, 120, fill=False, stroke=True)

    c.setFont('Helvetica-Bold', 12)
    c.drawString(MARGIN_L + 60, y - 20, 'Taxpayer:')
    c.setFont('Helvetica', 12)
    c.drawString(MARGIN_L + 180, y - 20, d.taxpayer_name() or '(Not provided)')

    if d.spouse_name():
        c.setFont('Helvetica-Bold', 12)
        c.drawString(MARGIN_L + 60, y - 40, 'Spouse:')
        c.setFont('Helvetica', 12)
        c.drawString(MARGIN_L + 180, y - 40, d.spouse_name())

    c.setFont('Helvetica-Bold', 12)
    c.drawString(MARGIN_L + 60, y - 65, 'SSN:')
    c.setFont('Helvetica', 12)
    c.drawString(MARGIN_L + 180, y - 65, d.ssn() or 'XXX-XX-XXXX')

    c.setFont('Helvetica-Bold', 12)
    c.drawString(MARGIN_L + 60, y - 85, 'Filing Status:')
    c.setFont('Helvetica', 12)
    c.drawString(MARGIN_L + 180, y - 85, d.filing_status_text() or '(Not specified)')

    c.setFont('Helvetica-Bold', 12)
    c.drawString(MARGIN_L + 60, y - 105, 'Address:')
    c.setFont('Helvetica', 12)
    addr = d.address_line1()
    addr2 = d.address_line2()
    c.drawString(MARGIN_L + 180, y - 105, f'{addr}, {addr2}' if addr else '(Not provided)')

    y -= 155

    # Summary box
    c.setFillColor(HexColor('#f8f9fa'))
    c.rect(MARGIN_L + 40, y - 200, CONTENT_W - 80, 200, fill=True, stroke=True)
    c.setFillColor(black)

    c.setFont('Helvetica-Bold', 13)
    c.drawCentredString(PAGE_W / 2, y - 20, 'TAX RETURN SUMMARY')

    summary_items = [
        ('Total Income (Line 9)', d.val('Form 1040', 'D21')),
        ('Adjusted Gross Income (Line 11)', d.val('Form 1040', 'D25')),
        ('Taxable Income (Line 15)', d.val('Form 1040', 'D31')),
        ('Total Tax (Line 22)', d.val('Form 1040', 'D40')),
        ('Total Payments (Line 33)', d.val('Form 1040', 'D53')),
    ]

    sy = y - 48
    for label, val in summary_items:
        c.setFont('Helvetica', 10)
        c.drawString(MARGIN_L + 70, sy, label)
        c.setFont('Courier-Bold', 11)
        c.drawRightString(MARGIN_L + CONTENT_W - 70, sy, fmt_curr_dollar(val))
        sy -= 22

    # Refund / Owed
    refund = d.val('Form 1040', 'D56')
    owed = d.val('Form 1040', 'D58')

    sy -= 8
    c.setLineWidth(1.5)
    c.line(MARGIN_L + 70, sy + 14, MARGIN_L + CONTENT_W - 70, sy + 14)
    sy -= 4

    if refund > 0:
        c.setFillColor(HexColor('#006600'))
        c.setFont('Helvetica-Bold', 13)
        c.drawString(MARGIN_L + 70, sy, 'REFUND DUE:')
        c.setFont('Courier-Bold', 14)
        c.drawRightString(MARGIN_L + CONTENT_W - 70, sy, fmt_curr_dollar(refund))
    elif owed > 0:
        c.setFillColor(HexColor('#cc0000'))
        c.setFont('Helvetica-Bold', 13)
        c.drawString(MARGIN_L + 70, sy, 'AMOUNT OWED:')
        c.setFont('Courier-Bold', 14)
        c.drawRightString(MARGIN_L + CONTENT_W - 70, sy, fmt_curr_dollar(owed))
    else:
        c.setFont('Helvetica-Bold', 12)
        c.drawString(MARGIN_L + 70, sy, 'No balance due / No refund')

    c.setFillColor(black)
    c.setLineWidth(0.5)
    y -= 230

    # Forms included
    c.setFont('Helvetica-Bold', 11)
    c.drawString(MARGIN_L + 40, y, 'Forms and Schedules Included:')
    y -= 18
    c.setFont('Helvetica', 9)
    forms_list = [
        'Form 1040 — U.S. Individual Income Tax Return',
        'Schedule 1 — Additional Income and Adjustments',
        'Schedule 2 — Additional Taxes',
        'Schedule 3 — Additional Credits and Payments',
        'Schedule A — Itemized Deductions',
        'Schedule B — Interest and Ordinary Dividends',
        'Schedule C — Profit or Loss From Business',
        'Schedule D — Capital Gains and Losses',
        'Form 8949 — Sales and Dispositions of Capital Assets',
        'Schedule E — Supplemental Income and Loss',
        'Schedule SE — Self-Employment Tax',
        'Form 8938 — FATCA (Foreign Financial Assets)',
        'Form 8621 — PFIC Information Return',
        'Form 3520 — Foreign Trusts and Gifts',
    ]
    for form in forms_list:
        c.drawString(MARGIN_L + 60, y, f'  {form}')
        y -= 13

    # Footer
    c.setFont('Helvetica', 7)
    c.setFillColor(IRS_LINE_GRAY)
    c.drawCentredString(PAGE_W / 2, MARGIN_B, 'This document is for informational purposes. Verify all information before filing.')
    c.setFillColor(black)


# ═══════════════════════════════════════════════════════════════════════════════
# MAIN — Assemble PDF Package
# ═══════════════════════════════════════════════════════════════════════════════
def generate_pdf(xlsx_path, output_path):
    print(f"Reading data from: {xlsx_path}")
    data = TaxData(xlsx_path)

    print(f"Generating PDF: {output_path}")
    c_pdf = canvas.Canvas(output_path, pagesize=letter)
    c_pdf.setTitle('Form 1040 Individual Income Tax Return — Tax Year 2025')
    c_pdf.setAuthor('Tax Return Preparation System')

    builders = [
        ('Cover Page', build_cover_page),
        ('Form 1040 Page 1', build_form_1040_page1),
        ('Form 1040 Page 2', build_form_1040_page2),
        ('Schedule 1', build_schedule_1),
        ('Schedule 2', build_schedule_2),
        ('Schedule 3', build_schedule_3),
        ('Schedule A', build_schedule_a),
        ('Schedule B', build_schedule_b),
        ('Schedule C', build_schedule_c),
        ('Schedule D', build_schedule_d),
        ('Form 8949', build_form_8949),
        ('Schedule E', build_schedule_e),
        ('Schedule SE', build_schedule_se),
        ('Form 8938', build_form_8938),
        ('Form 8621', build_form_8621),
        ('Form 3520', build_form_3520),
    ]

    for i, (name, builder) in enumerate(builders):
        if i > 0:
            c_pdf.showPage()
        renderer = IRSFormRenderer(c_pdf, data)
        print(f"  Building {name}...")
        builder(renderer, data)

    c_pdf.save()
    print(f"\nPDF generated successfully: {output_path}")
    print(f"Total pages: {len(builders)}")


if __name__ == '__main__':
    xlsx = sys.argv[1] if len(sys.argv) > 1 else 'mnt/Tax templates/1040_Tax_Return_2025.xlsx'
    pdf_out = sys.argv[2] if len(sys.argv) > 2 else 'mnt/Tax templates/1040_Tax_Return_2025.pdf'
    generate_pdf(xlsx, pdf_out)
