#!/usr/bin/env python3
"""Build comprehensive 2025 Form 1040 Tax Return Workbook."""

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side, numbers
from openpyxl.utils import get_column_letter
from copy import copy

wb = Workbook()

# ── Style constants ──────────────────────────────────────────────────────────
BLUE_INPUT = Font(name='Arial', size=10, color='0000FF')
BLACK = Font(name='Arial', size=10, color='000000')
BLACK_BOLD = Font(name='Arial', size=10, color='000000', bold=True)
GREEN_LINK = Font(name='Arial', size=10, color='008000')
HEADER_FONT = Font(name='Arial', size=11, color='FFFFFF', bold=True)
SECTION_FONT = Font(name='Arial', size=10, color='000000', bold=True)
TITLE_FONT = Font(name='Arial', size=14, color='000000', bold=True)
SUBTITLE_FONT = Font(name='Arial', size=11, color='000000', bold=True)
SMALL_FONT = Font(name='Arial', size=9, color='666666', italic=True)

YELLOW_FILL = PatternFill('solid', fgColor='FFFF00')
LIGHT_BLUE = PatternFill('solid', fgColor='DCE6F1')
LIGHT_GRAY = PatternFill('solid', fgColor='F2F2F2')
DARK_BLUE = PatternFill('solid', fgColor='1F4E79')
MEDIUM_BLUE = PatternFill('solid', fgColor='2E75B6')
LIGHT_GREEN = PatternFill('solid', fgColor='E2EFDA')
LIGHT_YELLOW = PatternFill('solid', fgColor='FFF2CC')
LIGHT_RED = PatternFill('solid', fgColor='FCE4EC')

CENTER = Alignment(horizontal='center', vertical='center', wrap_text=True)
LEFT = Alignment(horizontal='left', vertical='center', wrap_text=True)
RIGHT = Alignment(horizontal='right', vertical='center')
WRAP = Alignment(wrap_text=True, vertical='top')

THIN_BORDER = Border(
    left=Side(style='thin'), right=Side(style='thin'),
    top=Side(style='thin'), bottom=Side(style='thin')
)
BOTTOM_BORDER = Border(bottom=Side(style='thin'))
THICK_BOTTOM = Border(bottom=Side(style='medium'))

CURR_FMT = '$#,##0;($#,##0);"-"'
CURR_FMT2 = '$#,##0.00;($#,##0.00);"-"'
PCT_FMT = '0.0%'
NUM_FMT = '#,##0;(#,##0);"-"'
DATE_FMT = 'MM/DD/YYYY'

def style_cell(ws, row, col, value=None, font=None, fill=None, align=None, border=None, fmt=None, merge_end_col=None):
    cell = ws.cell(row=row, column=col, value=value)
    if font: cell.font = font
    if fill: cell.fill = fill
    if align: cell.alignment = align
    if border: cell.border = border
    if fmt: cell.number_format = fmt
    if merge_end_col:
        ws.merge_cells(start_row=row, start_column=col, end_row=row, end_column=merge_end_col)
    return cell

def section_header(ws, row, col_start, col_end, text):
    for c in range(col_start, col_end + 1):
        cell = ws.cell(row=row, column=c)
        cell.fill = DARK_BLUE
        cell.font = HEADER_FONT
        cell.alignment = CENTER
        cell.border = THIN_BORDER
    ws.cell(row=row, column=col_start, value=text)
    if col_end > col_start:
        ws.merge_cells(start_row=row, start_column=col_start, end_row=row, end_column=col_end)

def sub_header(ws, row, col_start, col_end, text):
    for c in range(col_start, col_end + 1):
        cell = ws.cell(row=row, column=c)
        cell.fill = MEDIUM_BLUE
        cell.font = Font(name='Arial', size=10, color='FFFFFF', bold=True)
        cell.alignment = CENTER
        cell.border = THIN_BORDER
    ws.cell(row=row, column=col_start, value=text)
    if col_end > col_start:
        ws.merge_cells(start_row=row, start_column=col_start, end_row=row, end_column=col_end)

def input_row(ws, row, col_label, label, col_val, font=BLUE_INPUT, fill=LIGHT_YELLOW, fmt=None, note=None, note_col=None):
    style_cell(ws, row, col_label, label, BLACK, LIGHT_GRAY, LEFT, THIN_BORDER)
    c = style_cell(ws, row, col_val, None, font, fill, RIGHT, THIN_BORDER, fmt)
    if note and note_col:
        style_cell(ws, row, note_col, note, SMALL_FONT, None, LEFT)
    return c

def formula_row(ws, row, col_label, label, col_val, formula, fmt=CURR_FMT, fill=None):
    style_cell(ws, row, col_label, label, BLACK, LIGHT_GRAY, LEFT, THIN_BORDER)
    style_cell(ws, row, col_val, formula, GREEN_LINK if "!" in str(formula) else BLACK, fill or LIGHT_BLUE, RIGHT, THIN_BORDER, fmt)


# ═══════════════════════════════════════════════════════════════════════════════
# TAB 1: INTAKE
# ═══════════════════════════════════════════════════════════════════════════════
ws = wb.active
ws.title = 'Intake'
ws.sheet_properties.tabColor = '1F4E79'
ws.column_dimensions['A'].width = 3
ws.column_dimensions['B'].width = 35
ws.column_dimensions['C'].width = 30
ws.column_dimensions['D'].width = 15
ws.column_dimensions['E'].width = 30
ws.column_dimensions['F'].width = 30
ws.column_dimensions['G'].width = 20
ws.column_dimensions['H'].width = 20
ws.column_dimensions['I'].width = 25

r = 1
style_cell(ws, r, 2, 'Form 1040 — Individual Income Tax Return', TITLE_FONT, None, LEFT, None, None, 6)
r += 1
style_cell(ws, r, 2, 'Tax Year 2025 — Comprehensive Intake Worksheet', SUBTITLE_FONT, None, LEFT, None, None, 6)
r += 1
style_cell(ws, r, 2, 'Blue cells = Client input  |  Yellow highlight = Required fields', SMALL_FONT)
r += 2

# ── Section 1: Filing Information ─────────────────────────────────────────────
section_header(ws, r, 2, 6, 'SECTION 1: FILING INFORMATION'); r += 1
input_row(ws, r, 2, 'Filing Status', 3, fmt=None, note='1=Single, 2=MFJ, 3=MFS, 4=HOH, 5=QSS', note_col=4); r += 1
input_row(ws, r, 2, 'Digital Assets Question (Y/N)', 3, note='Did you receive/sell/exchange digital assets?', note_col=4); r += 1
input_row(ws, r, 2, 'Standard Deduction or Itemize?', 3, note='S=Standard, I=Itemize', note_col=4); r += 1
input_row(ws, r, 2, 'Are you a dependent of another taxpayer?', 3, note='Y/N', note_col=4); r += 1
input_row(ws, r, 2, 'Presidential Election Campaign Fund?', 3, note='Y/N', note_col=4); r += 1
r += 1

# ── Section 2: Taxpayer Information ───────────────────────────────────────────
section_header(ws, r, 2, 6, 'SECTION 2: TAXPAYER INFORMATION'); r += 1
sub_header(ws, r, 2, 3, 'Primary Taxpayer'); sub_header(ws, r, 4, 6, 'Spouse (if MFJ/MFS)'); r += 1
for field in ['First Name', 'Middle Initial', 'Last Name', 'Social Security Number',
              'Date of Birth', 'Occupation', 'Phone Number', 'Email Address',
              'Legally Blind? (Y/N)', 'Age 65 or Older? (Y/N)',
              'US Citizen? (Y/N)', 'Resident Alien? (Y/N)']:
    style_cell(ws, r, 2, field, BLACK, LIGHT_GRAY, LEFT, THIN_BORDER)
    style_cell(ws, r, 3, None, BLUE_INPUT, LIGHT_YELLOW, LEFT, THIN_BORDER,
               DATE_FMT if 'Date' in field else None)
    style_cell(ws, r, 4, field, BLACK, LIGHT_GRAY, LEFT, THIN_BORDER)
    style_cell(ws, r, 5, None, BLUE_INPUT, LIGHT_YELLOW, LEFT, THIN_BORDER,
               DATE_FMT if 'Date' in field else None)
    r += 1
r += 1

# ── Section 3: Address ────────────────────────────────────────────────────────
section_header(ws, r, 2, 6, 'SECTION 3: MAILING ADDRESS'); r += 1
for field in ['Street Address', 'Apartment Number', 'City', 'State', 'ZIP Code',
              'Foreign Country Name', 'Foreign Province/State', 'Foreign Postal Code']:
    input_row(ws, r, 2, field, 3); r += 1
r += 1

# ── Section 4: Dependents ────────────────────────────────────────────────────
section_header(ws, r, 2, 9, 'SECTION 4: DEPENDENTS'); r += 1
dep_headers = ['First Name', 'Last Name', 'SSN', 'Relationship', 'Date of Birth',
               'Months Lived w/ You', 'Child Tax Credit (Y/N)', 'Other Dependent Credit (Y/N)']
for i, h in enumerate(dep_headers, 2):
    style_cell(ws, r, i, h, Font(name='Arial', size=9, color='FFFFFF', bold=True),
               MEDIUM_BLUE, CENTER, THIN_BORDER)
r += 1
DEP_START = r
for _ in range(10):
    for c in range(2, 10):
        style_cell(ws, r, c, None, BLUE_INPUT, LIGHT_YELLOW, LEFT, THIN_BORDER,
                   DATE_FMT if c == 6 else None)
    r += 1
DEP_END = r - 1
r += 1

# ── Section 5: Income — Wages ────────────────────────────────────────────────
section_header(ws, r, 2, 9, 'SECTION 5: INCOME — WAGES (W-2)'); r += 1
w2_headers = ['Employer Name', 'EIN', 'Wages (Box 1)', 'Federal Tax Withheld (Box 2)',
              'SS Wages (Box 3)', 'SS Tax (Box 4)', 'Medicare Wages (Box 5)', 'Medicare Tax (Box 6)']
for i, h in enumerate(w2_headers, 2):
    style_cell(ws, r, i, h, Font(name='Arial', size=9, color='FFFFFF', bold=True),
               MEDIUM_BLUE, CENTER, THIN_BORDER)
r += 1
W2_START = r
for _ in range(5):
    for c in range(2, 10):
        fmt = CURR_FMT if c >= 4 else None
        style_cell(ws, r, c, None, BLUE_INPUT, LIGHT_YELLOW, RIGHT if c >= 4 else LEFT, THIN_BORDER, fmt)
    r += 1
W2_END = r - 1
style_cell(ws, r, 2, 'Total Wages', BLACK_BOLD, LIGHT_GREEN, LEFT, THICK_BOTTOM)
style_cell(ws, r, 4, f'=SUM(D{W2_START}:D{W2_END})', BLACK_BOLD, LIGHT_GREEN, RIGHT, THICK_BOTTOM, CURR_FMT)
style_cell(ws, r, 5, f'=SUM(E{W2_START}:E{W2_END})', BLACK_BOLD, LIGHT_GREEN, RIGHT, THICK_BOTTOM, CURR_FMT)
TOTAL_WAGES_ROW = r
r += 2

# ── Section 6: Income — Interest (1099-INT) ──────────────────────────────────
section_header(ws, r, 2, 7, 'SECTION 6: INTEREST INCOME (1099-INT)'); r += 1
int_headers = ['Payer Name', 'EIN', 'Interest (Box 1)', 'Tax-Exempt Interest', 'Federal Tax Withheld', 'Foreign Tax Paid']
for i, h in enumerate(int_headers, 2):
    style_cell(ws, r, i, h, Font(name='Arial', size=9, color='FFFFFF', bold=True),
               MEDIUM_BLUE, CENTER, THIN_BORDER)
r += 1
INT_START = r
for _ in range(8):
    for c in range(2, 8):
        fmt = CURR_FMT if c >= 4 else None
        style_cell(ws, r, c, None, BLUE_INPUT, LIGHT_YELLOW, RIGHT if c >= 4 else LEFT, THIN_BORDER, fmt)
    r += 1
INT_END = r - 1
style_cell(ws, r, 2, 'Total Taxable Interest', BLACK_BOLD, LIGHT_GREEN, LEFT, THICK_BOTTOM)
style_cell(ws, r, 4, f'=SUM(D{INT_START}:D{INT_END})', BLACK_BOLD, LIGHT_GREEN, RIGHT, THICK_BOTTOM, CURR_FMT)
TOTAL_INT_ROW = r
r += 1
style_cell(ws, r, 2, 'Total Tax-Exempt Interest', BLACK_BOLD, LIGHT_GREEN, LEFT, THICK_BOTTOM)
style_cell(ws, r, 5, f'=SUM(E{INT_START}:E{INT_END})', BLACK_BOLD, LIGHT_GREEN, RIGHT, THICK_BOTTOM, CURR_FMT)
r += 2

# ── Section 7: Dividends (1099-DIV) ──────────────────────────────────────────
section_header(ws, r, 2, 8, 'SECTION 7: DIVIDEND INCOME (1099-DIV)'); r += 1
div_headers = ['Payer Name', 'Total Ordinary Div', 'Qualified Dividends', 'Capital Gain Dist',
               'Nondividend Dist', 'Federal Tax Withheld', 'Foreign Tax Paid']
for i, h in enumerate(div_headers, 2):
    style_cell(ws, r, i, h, Font(name='Arial', size=9, color='FFFFFF', bold=True),
               MEDIUM_BLUE, CENTER, THIN_BORDER)
r += 1
DIV_START = r
for _ in range(8):
    for c in range(2, 9):
        fmt = CURR_FMT if c >= 3 else None
        style_cell(ws, r, c, None, BLUE_INPUT, LIGHT_YELLOW, RIGHT if c >= 3 else LEFT, THIN_BORDER, fmt)
    r += 1
DIV_END = r - 1
style_cell(ws, r, 2, 'Total Ordinary Dividends', BLACK_BOLD, LIGHT_GREEN, LEFT, THICK_BOTTOM)
style_cell(ws, r, 3, f'=SUM(C{DIV_START}:C{DIV_END})', BLACK_BOLD, LIGHT_GREEN, RIGHT, THICK_BOTTOM, CURR_FMT)
TOTAL_ORD_DIV_ROW = r
r += 1
style_cell(ws, r, 2, 'Total Qualified Dividends', BLACK_BOLD, LIGHT_GREEN, LEFT, THICK_BOTTOM)
style_cell(ws, r, 4, f'=SUM(D{DIV_START}:D{DIV_END})', BLACK_BOLD, LIGHT_GREEN, RIGHT, THICK_BOTTOM, CURR_FMT)
TOTAL_QUAL_DIV_ROW = r
r += 2

# ── Section 8: Capital Gains / 1099-B ────────────────────────────────────────
section_header(ws, r, 2, 9, 'SECTION 8: CAPITAL GAINS & LOSSES (1099-B / Form 8949)'); r += 1
cg_headers = ['Description of Property', 'Date Acquired', 'Date Sold', 'Proceeds',
              'Cost Basis', 'Adjustment Code', 'Adjustment Amt', 'Short/Long (S/L)']
for i, h in enumerate(cg_headers, 2):
    style_cell(ws, r, i, h, Font(name='Arial', size=9, color='FFFFFF', bold=True),
               MEDIUM_BLUE, CENTER, THIN_BORDER)
r += 1
CG_START = r
for _ in range(20):
    for c in range(2, 10):
        fmt = DATE_FMT if c in [3, 4] else (CURR_FMT if c in [5, 6, 8] else None)
        style_cell(ws, r, c, None, BLUE_INPUT, LIGHT_YELLOW, RIGHT if c >= 5 else LEFT, THIN_BORDER, fmt)
    r += 1
CG_END = r - 1
style_cell(ws, r, 2, 'Total Proceeds', BLACK_BOLD, LIGHT_GREEN, LEFT, THICK_BOTTOM)
style_cell(ws, r, 5, f'=SUM(E{CG_START}:E{CG_END})', BLACK_BOLD, LIGHT_GREEN, RIGHT, THICK_BOTTOM, CURR_FMT)
r += 1
style_cell(ws, r, 2, 'Total Cost Basis', BLACK_BOLD, LIGHT_GREEN, LEFT, THICK_BOTTOM)
style_cell(ws, r, 6, f'=SUM(F{CG_START}:F{CG_END})', BLACK_BOLD, LIGHT_GREEN, RIGHT, THICK_BOTTOM, CURR_FMT)
r += 2

# ── Section 9: Business Income (Schedule C) ──────────────────────────────────
section_header(ws, r, 2, 6, 'SECTION 9: BUSINESS INCOME (Schedule C)'); r += 1
SCHC_START = r
for field, note in [
    ('Business Name', ''), ('Business EIN', ''), ('Principal Business Code', 'NAICS code'),
    ('Business Address', ''), ('Accounting Method', 'Cash/Accrual/Other'),
    ('Gross Receipts/Sales', ''), ('Returns and Allowances', ''), ('Cost of Goods Sold', ''),
    ('Advertising', ''), ('Car/Truck Expenses', ''), ('Commissions/Fees', ''),
    ('Contract Labor', ''), ('Depreciation (Form 4562)', ''), ('Employee Benefit Programs', ''),
    ('Insurance (non-health)', ''), ('Interest (Mortgage)', ''), ('Interest (Other)', ''),
    ('Legal/Professional Services', ''), ('Office Expense', ''), ('Rent/Lease (Vehicles)', ''),
    ('Rent/Lease (Other)', ''), ('Repairs/Maintenance', ''), ('Supplies', ''),
    ('Taxes/Licenses', ''), ('Travel', ''), ('Meals (50%)', ''),
    ('Utilities', ''), ('Wages Paid', ''), ('Other Expenses (describe)', ''),
    ('Other Expenses Amount', ''), ('Business Use of Home', 'Form 8829 amount')
]:
    fmt = CURR_FMT if any(k in field for k in ['Receipts', 'Returns', 'Cost', 'Advertising',
        'Car', 'Commissions', 'Contract', 'Depreciation', 'Employee', 'Insurance', 'Interest',
        'Legal', 'Office', 'Rent', 'Repairs', 'Supplies', 'Taxes', 'Travel', 'Meals',
        'Utilities', 'Wages', 'Amount', 'Home']) else None
    input_row(ws, r, 2, field, 3, fmt=fmt, note=note, note_col=4)
    r += 1
r += 1

# ── Section 10: Rental Income (Schedule E) ───────────────────────────────────
section_header(ws, r, 2, 8, 'SECTION 10: RENTAL REAL ESTATE (Schedule E)'); r += 1
sub_header(ws, r, 2, 2, 'Item'); sub_header(ws, r, 3, 3, 'Property 1')
sub_header(ws, r, 4, 4, 'Property 2'); sub_header(ws, r, 5, 5, 'Property 3'); r += 1
SCHE_START = r
for field in ['Property Address', 'Property Type', 'Days Rented', 'Days Personal Use',
              'Rents Received', 'Advertising', 'Auto/Travel', 'Cleaning/Maintenance',
              'Commissions', 'Insurance', 'Legal/Professional', 'Management Fees',
              'Mortgage Interest', 'Other Interest', 'Repairs', 'Supplies', 'Taxes',
              'Utilities', 'Depreciation', 'Other Expenses']:
    style_cell(ws, r, 2, field, BLACK, LIGHT_GRAY, LEFT, THIN_BORDER)
    for c in range(3, 6):
        fmt = CURR_FMT if field not in ['Property Address', 'Property Type', 'Days Rented', 'Days Personal Use'] else (NUM_FMT if 'Days' in field else None)
        style_cell(ws, r, c, None, BLUE_INPUT, LIGHT_YELLOW, RIGHT if fmt else LEFT, THIN_BORDER, fmt)
    r += 1
r += 1

# ── Section 11: Other Income ─────────────────────────────────────────────────
section_header(ws, r, 2, 6, 'SECTION 11: OTHER INCOME'); r += 1
OTHER_INC_START = r
for field in ['State/Local Tax Refund (if itemized prior year)', 'Alimony Received (pre-2019 divorce)',
              'IRA Distributions — Gross', 'IRA Distributions — Taxable',
              'Pensions/Annuities — Gross', 'Pensions/Annuities — Taxable',
              'Social Security Benefits — Gross', 'Social Security Benefits — Taxable',
              'Unemployment Compensation', 'Gambling Winnings',
              'Other Income (describe)', 'Other Income Amount',
              'Farm Income/Loss (Schedule F)', 'Rental Royalties/Partnerships/S-Corps (Sch E pg 2)',
              'Tip Income Not on W-2']:
    input_row(ws, r, 2, field, 3, fmt=CURR_FMT)
    r += 1
r += 1

# ── Section 12: Adjustments to Income ────────────────────────────────────────
section_header(ws, r, 2, 6, 'SECTION 12: ADJUSTMENTS TO INCOME (Schedule 1, Part II)'); r += 1
ADJ_START = r
for field in ['Educator Expenses (max $300)', 'Business Expenses (reservists, etc.)',
              'HSA Deduction', 'Moving Expenses (military only)',
              'Deductible Self-Employment Tax', 'Self-Employed SEP/SIMPLE/Qualified Plans',
              'Self-Employed Health Insurance', 'Penalty on Early Withdrawal of Savings',
              'IRA Deduction', 'Student Loan Interest (max $2,500)',
              'Tuition and Fees', 'Charitable Contributions (non-itemizer)',
              'Other Adjustments (describe)', 'Other Adjustments Amount']:
    input_row(ws, r, 2, field, 3, fmt=CURR_FMT)
    r += 1
r += 1

# ── Section 13: Itemized Deductions (Schedule A) ─────────────────────────────
section_header(ws, r, 2, 6, 'SECTION 13: ITEMIZED DEDUCTIONS (Schedule A)'); r += 1
SCHA_START = r
for field in ['Medical/Dental Expenses (total)', 'State/Local Income Taxes Paid',
              'State/Local Sales Taxes Paid', 'Real Estate Taxes Paid',
              'Personal Property Taxes', 'Total SALT (auto-capped at $10,000)',
              'Home Mortgage Interest (Form 1098)', 'Mortgage Insurance Premiums',
              'Investment Interest Expense',
              'Gifts to Charity — Cash', 'Gifts to Charity — Noncash',
              'Carryover from Prior Year',
              'Casualty/Theft Losses (federally declared disaster)',
              'Gambling Losses (limited to winnings)', 'Other Itemized Deductions']:
    input_row(ws, r, 2, field, 3, fmt=CURR_FMT)
    r += 1
r += 1

# ── Section 14: Tax Credits ──────────────────────────────────────────────────
section_header(ws, r, 2, 6, 'SECTION 14: TAX CREDITS'); r += 1
CREDIT_START = r
for field in ['Foreign Tax Credit (Form 1116)', 'Child/Dependent Care Credit (Form 2441)',
              'Education Credits (Form 8863 — AOTC)', 'Education Credits (Form 8863 — LLC)',
              'Retirement Savings Credit (Form 8880)', 'Child Tax Credit (auto-calculated)',
              'Residential Energy Credit (Form 5695)', 'EV Credit (Form 8936)',
              'Other Nonrefundable Credits', 'Additional Child Tax Credit (Form 8812)',
              'American Opportunity Credit (refundable)', 'Net Premium Tax Credit (Form 8962)',
              'Earned Income Credit', 'Other Refundable Credits']:
    input_row(ws, r, 2, field, 3, fmt=CURR_FMT)
    r += 1
r += 1

# ── Section 15: Payments ─────────────────────────────────────────────────────
section_header(ws, r, 2, 6, 'SECTION 15: TAX PAYMENTS & WITHHOLDING'); r += 1
PAY_START = r
for field in ['Federal Tax Withheld (total from W-2s, 1099s, etc.)',
              'Estimated Tax Payments (Q1)', 'Estimated Tax Payments (Q2)',
              'Estimated Tax Payments (Q3)', 'Estimated Tax Payments (Q4)',
              'Amount Applied from Prior Year Return',
              'Extension Payment (Form 4868)', 'Excess Social Security Tax Withheld',
              'Other Payments/Credits']:
    input_row(ws, r, 2, field, 3, fmt=CURR_FMT)
    r += 1
r += 1

# ── Section 16: Foreign Assets ────────────────────────────────────────────────
section_header(ws, r, 2, 9, 'SECTION 16: FOREIGN FINANCIAL ASSETS (FBAR / FATCA / Form 8938)'); r += 1
style_cell(ws, r, 2, 'Did you have a financial interest in or signature authority over foreign accounts? (Y/N)', BLACK, LIGHT_GRAY, LEFT, THIN_BORDER)
style_cell(ws, r, 3, None, BLUE_INPUT, LIGHT_YELLOW, LEFT, THIN_BORDER)
ws.merge_cells(start_row=r, start_column=2, end_row=r, end_column=2)
r += 1
input_row(ws, r, 2, 'Maximum aggregate value of all foreign accounts during year', 3, fmt=CURR_FMT); r += 1
input_row(ws, r, 2, 'Were any accounts in a foreign country? (Y/N)', 3); r += 1
r += 1

sub_header(ws, r, 2, 9, 'Foreign Account Details (for FBAR and Form 8938)'); r += 1
fa_headers = ['Account Type', 'Financial Institution', 'Country', 'Account Number',
              'Max Value During Year', 'Year-End Value', 'Income Earned', 'Joint Account? (Y/N)']
for i, h in enumerate(fa_headers, 2):
    style_cell(ws, r, i, h, Font(name='Arial', size=9, color='FFFFFF', bold=True),
               MEDIUM_BLUE, CENTER, THIN_BORDER)
r += 1
FA_START = r
for _ in range(10):
    for c in range(2, 10):
        fmt = CURR_FMT if c in [6, 7, 8] else None
        style_cell(ws, r, c, None, BLUE_INPUT, LIGHT_YELLOW, RIGHT if fmt else LEFT, THIN_BORDER, fmt)
    r += 1
FA_END = r - 1
style_cell(ws, r, 2, 'Total Max Value', BLACK_BOLD, LIGHT_GREEN, LEFT, THICK_BOTTOM)
style_cell(ws, r, 6, f'=SUM(F{FA_START}:F{FA_END})', BLACK_BOLD, LIGHT_GREEN, RIGHT, THICK_BOTTOM, CURR_FMT)
style_cell(ws, r, 7, f'=SUM(G{FA_START}:G{FA_END})', BLACK_BOLD, LIGHT_GREEN, RIGHT, THICK_BOTTOM, CURR_FMT)
style_cell(ws, r, 8, f'=SUM(H{FA_START}:H{FA_END})', BLACK_BOLD, LIGHT_GREEN, RIGHT, THICK_BOTTOM, CURR_FMT)
r += 2

# ── Section 17: Foreign Trusts / Gifts (Form 3520) ───────────────────────────
section_header(ws, r, 2, 8, 'SECTION 17: FOREIGN TRUSTS & GIFTS (Form 3520)'); r += 1
input_row(ws, r, 2, 'Did you receive gifts from foreign persons > $100,000? (Y/N)', 3); r += 1
input_row(ws, r, 2, 'Total gifts from foreign individuals', 3, fmt=CURR_FMT); r += 1
input_row(ws, r, 2, 'Total gifts from foreign corporations (> $19,570)', 3, fmt=CURR_FMT); r += 1
input_row(ws, r, 2, 'Foreign trust — Name of trust', 3); r += 1
input_row(ws, r, 2, 'Foreign trust — Country', 3); r += 1
input_row(ws, r, 2, 'Foreign trust — EIN/Reference ID', 3); r += 1
input_row(ws, r, 2, 'Foreign trust — Distributions received', 3, fmt=CURR_FMT); r += 1
input_row(ws, r, 2, 'Foreign trust — Transfers to trust', 3, fmt=CURR_FMT); r += 1
r += 1

# ── Section 18: PFIC (Form 8621) ─────────────────────────────────────────────
section_header(ws, r, 2, 9, 'SECTION 18: PASSIVE FOREIGN INVESTMENT COMPANIES (Form 8621)'); r += 1
pfic_headers = ['PFIC Name', 'Country', 'EIN/Ref ID', 'Date Acquired', 'Shares Held',
                'Fair Market Value (YE)', 'QEF/MTM Election', 'Distributions Received']
for i, h in enumerate(pfic_headers, 2):
    style_cell(ws, r, i, h, Font(name='Arial', size=9, color='FFFFFF', bold=True),
               MEDIUM_BLUE, CENTER, THIN_BORDER)
r += 1
PFIC_START = r
for _ in range(5):
    for c in range(2, 10):
        fmt = CURR_FMT if c in [7, 9] else (DATE_FMT if c == 5 else None)
        style_cell(ws, r, c, None, BLUE_INPUT, LIGHT_YELLOW, RIGHT if fmt else LEFT, THIN_BORDER, fmt)
    r += 1
PFIC_END = r - 1
r += 1

# ── Section 19: K-1 Income ───────────────────────────────────────────────────
section_header(ws, r, 2, 8, 'SECTION 19: PARTNERSHIP / S-CORP / TRUST K-1 INCOME'); r += 1
k1_headers = ['Entity Name', 'EIN', 'Type (P/S/T)', 'Ordinary Income', 'Rental Income',
              'Interest/Dividends', 'Capital Gains']
for i, h in enumerate(k1_headers, 2):
    style_cell(ws, r, i, h, Font(name='Arial', size=9, color='FFFFFF', bold=True),
               MEDIUM_BLUE, CENTER, THIN_BORDER)
r += 1
K1_START = r
for _ in range(5):
    for c in range(2, 9):
        fmt = CURR_FMT if c >= 5 else None
        style_cell(ws, r, c, None, BLUE_INPUT, LIGHT_YELLOW, RIGHT if fmt else LEFT, THIN_BORDER, fmt)
    r += 1
K1_END = r - 1
r += 1

# ── Section 20: Retirement / HSA ─────────────────────────────────────────────
section_header(ws, r, 2, 6, 'SECTION 20: RETIREMENT CONTRIBUTIONS & HSA'); r += 1
for field in ['Traditional IRA Contributions', 'Roth IRA Contributions',
              '401(k) Employee Deferrals', '401(k) Employer Match',
              'SEP IRA Contributions', 'SIMPLE IRA Contributions',
              'HSA Contributions (Self)', 'HSA Contributions (Employer)',
              'HSA Coverage Type (Self/Family)',
              'HSA Distributions — Total', 'HSA Distributions — Qualified Medical']:
    input_row(ws, r, 2, field, 3, fmt=CURR_FMT if 'Type' not in field else None)
    r += 1
r += 1

# ── Section 21: Additional Info ──────────────────────────────────────────────
section_header(ws, r, 2, 6, 'SECTION 21: ADDITIONAL INFORMATION'); r += 1
for field in ['Bank Routing Number (for direct deposit)', 'Bank Account Number',
              'Account Type (Checking/Savings)', 'Third Party Designee Name',
              'Third Party Designee Phone', 'Third Party Designee PIN',
              'Preparer Name', 'Preparer PTIN', 'Preparer Firm Name',
              'Preparer Firm EIN', 'Preparer Firm Address']:
    input_row(ws, r, 2, field, 3)
    r += 1

INTAKE_LAST_ROW = r
print(f"Intake tab complete. Last row: {INTAKE_LAST_ROW}")


# ═══════════════════════════════════════════════════════════════════════════════
# TAB 2: TAX TABLES
# ═══════════════════════════════════════════════════════════════════════════════
ws_tt = wb.create_sheet('Tax Tables')
ws_tt.sheet_properties.tabColor = '7030A0'
ws_tt.column_dimensions['A'].width = 3
ws_tt.column_dimensions['B'].width = 25
ws_tt.column_dimensions['C'].width = 18
ws_tt.column_dimensions['D'].width = 18
ws_tt.column_dimensions['E'].width = 18
ws_tt.column_dimensions['F'].width = 18
ws_tt.column_dimensions['G'].width = 18

r = 1
style_cell(ws_tt, r, 2, '2025 Federal Tax Tables & Key Thresholds', TITLE_FONT); r += 2

# ── Tax Brackets ──────────────────────────────────────────────────────────────
section_header(ws_tt, r, 2, 7, 'ORDINARY INCOME TAX BRACKETS — TAX YEAR 2025'); r += 1
for i, h in enumerate(['Rate', 'Single', 'MFJ', 'MFS', 'HOH'], 2):
    style_cell(ws_tt, r, i, h, HEADER_FONT, MEDIUM_BLUE, CENTER, THIN_BORDER)
r += 1
brackets = [
    ('10%', 0, 11925, 0, 23850, 0, 11925, 0, 17000),
    ('12%', 11925, 48475, 23850, 96950, 11925, 48475, 17000, 64850),
    ('22%', 48475, 103350, 96950, 206700, 48475, 103350, 64850, 103350),
    ('24%', 103350, 197300, 206700, 394600, 103350, 197300, 103350, 197300),
    ('32%', 197300, 250525, 394600, 501050, 197300, 250525, 197300, 250500),
    ('35%', 250525, 626350, 501050, 751600, 250525, 626350, 250500, 626350),
    ('37%', 626350, None, 751600, None, 626350, None, 626350, None),
]
BRACKET_START = r
for rate, s1, s2, m1, m2, ms1, ms2, h1, h2 in brackets:
    style_cell(ws_tt, r, 2, rate, BLACK_BOLD, LIGHT_GRAY, CENTER, THIN_BORDER, PCT_FMT if False else None)
    s_text = f'${s1:,} – ${s2:,}' if s2 else f'Over ${s1:,}'
    m_text = f'${m1:,} – ${m2:,}' if m2 else f'Over ${m1:,}'
    ms_text = f'${ms1:,} – ${ms2:,}' if ms2 else f'Over ${ms1:,}'
    h_text = f'${h1:,} – ${h2:,}' if h2 else f'Over ${h1:,}'
    style_cell(ws_tt, r, 3, s_text, BLACK, None, CENTER, THIN_BORDER)
    style_cell(ws_tt, r, 4, m_text, BLACK, None, CENTER, THIN_BORDER)
    style_cell(ws_tt, r, 5, ms_text, BLACK, None, CENTER, THIN_BORDER)
    style_cell(ws_tt, r, 6, h_text, BLACK, None, CENTER, THIN_BORDER)
    r += 1
r += 1

# Bracket thresholds for formulas (numeric)
section_header(ws_tt, r, 2, 7, 'BRACKET THRESHOLDS (Numeric — Used by Formulas)'); r += 1
for i, h in enumerate(['Rate', 'Single Lower', 'Single Upper', 'MFJ Lower', 'MFJ Upper', 'HOH Lower'], 2):
    style_cell(ws_tt, r, i, h, HEADER_FONT, MEDIUM_BLUE, CENTER, THIN_BORDER)
r += 1
TT_RATES_START = r
rates_num = [
    (0.10, 0, 11925, 0, 23850, 0),
    (0.12, 11925, 48475, 23850, 96950, 17000),
    (0.22, 48475, 103350, 96950, 206700, 64850),
    (0.24, 103350, 197300, 206700, 394600, 103350),
    (0.32, 197300, 250525, 394600, 501050, 197300),
    (0.35, 250525, 626350, 501050, 751600, 250500),
    (0.37, 626350, 9999999, 751600, 9999999, 626350),
]
for rate, sl, su, ml, mu, hl in rates_num:
    style_cell(ws_tt, r, 2, rate, BLACK, None, CENTER, THIN_BORDER, '0%')
    style_cell(ws_tt, r, 3, sl, BLACK, None, RIGHT, THIN_BORDER, CURR_FMT)
    style_cell(ws_tt, r, 4, su, BLACK, None, RIGHT, THIN_BORDER, CURR_FMT)
    style_cell(ws_tt, r, 5, ml, BLACK, None, RIGHT, THIN_BORDER, CURR_FMT)
    style_cell(ws_tt, r, 6, mu, BLACK, None, RIGHT, THIN_BORDER, CURR_FMT)
    style_cell(ws_tt, r, 7, hl, BLACK, None, RIGHT, THIN_BORDER, CURR_FMT)
    r += 1
TT_RATES_END = r - 1
r += 1

# ── Standard Deductions ──────────────────────────────────────────────────────
section_header(ws_tt, r, 2, 4, 'STANDARD DEDUCTIONS'); r += 1
SD_ROW_START = r
sd_items = [('Single', 15750), ('Married Filing Jointly', 31500),
            ('Married Filing Separately', 15750), ('Head of Household', 23625),
            ('Qualifying Surviving Spouse', 31500),
            ('Additional — Age 65+ or Blind (Single/HOH)', 2000),
            ('Additional — Age 65+ or Blind (MFJ/MFS, per person)', 1600)]
for label, val in sd_items:
    style_cell(ws_tt, r, 2, label, BLACK, LIGHT_GRAY, LEFT, THIN_BORDER)
    style_cell(ws_tt, r, 3, val, BLACK, None, RIGHT, THIN_BORDER, CURR_FMT)
    r += 1
r += 1

# ── Capital Gains Rates ──────────────────────────────────────────────────────
section_header(ws_tt, r, 2, 5, 'LONG-TERM CAPITAL GAINS RATES'); r += 1
for i, h in enumerate(['Rate', 'Single', 'MFJ', 'HOH'], 2):
    style_cell(ws_tt, r, i, h, HEADER_FONT, MEDIUM_BLUE, CENTER, THIN_BORDER)
r += 1
CG_RATES_START = r
cg_data = [
    ('0%', '$0 – $48,350', '$0 – $96,700', '$0 – $64,750'),
    ('15%', '$48,350 – $533,400', '$96,700 – $600,050', '$64,750 – $566,700'),
    ('20%', 'Over $533,400', 'Over $600,050', 'Over $566,700'),
]
for rate, s, m, h in cg_data:
    style_cell(ws_tt, r, 2, rate, BLACK_BOLD, LIGHT_GRAY, CENTER, THIN_BORDER)
    style_cell(ws_tt, r, 3, s, BLACK, None, CENTER, THIN_BORDER)
    style_cell(ws_tt, r, 4, m, BLACK, None, CENTER, THIN_BORDER)
    style_cell(ws_tt, r, 5, h, BLACK, None, CENTER, THIN_BORDER)
    r += 1
r += 1

# ── Key Thresholds ────────────────────────────────────────────────────────────
section_header(ws_tt, r, 2, 4, 'KEY THRESHOLDS & LIMITS'); r += 1
KEY_START = r
thresholds = [
    ('Social Security Wage Base', 176100),
    ('SE Tax Rate — Social Security (12.4%)', 0.124),
    ('SE Tax Rate — Medicare (2.9%)', 0.029),
    ('Additional Medicare Tax Threshold (Single)', 200000),
    ('Additional Medicare Tax Threshold (MFJ)', 250000),
    ('Additional Medicare Tax Rate', 0.009),
    ('Net Investment Income Tax Threshold (Single)', 200000),
    ('Net Investment Income Tax Threshold (MFJ)', 250000),
    ('NIIT Rate', 0.038),
    ('AMT Exemption — Single', 88100),
    ('AMT Exemption — MFJ', 137000),
    ('AMT 28% Threshold', 239100),
    ('AMT Phaseout — Single', 626350),
    ('AMT Phaseout — MFJ', 1252700),
    ('Child Tax Credit (per qualifying child)', 2200),
    ('Child Tax Credit — Refundable Portion', 1700),
    ('SALT Deduction Cap', 10000),
    ('QBI Deduction Threshold — Single', 197300),
    ('QBI Deduction Threshold — MFJ', 394600),
    ('Annual Gift Tax Exclusion', 19000),
    ('FATCA Threshold — Domestic Single (Year-End)', 50000),
    ('FATCA Threshold — Domestic Single (Any Time)', 75000),
    ('FATCA Threshold — Domestic MFJ (Year-End)', 100000),
    ('FATCA Threshold — Domestic MFJ (Any Time)', 150000),
    ('FATCA Threshold — Foreign Single (Year-End)', 200000),
    ('FATCA Threshold — Foreign Single (Any Time)', 300000),
    ('FATCA Threshold — Foreign MFJ (Year-End)', 400000),
    ('FATCA Threshold — Foreign MFJ (Any Time)', 600000),
    ('FBAR Threshold (aggregate)', 10000),
    ('Form 3520 Gift Threshold — Foreign Individuals', 100000),
    ('Form 3520 Gift Threshold — Foreign Corporations', 19570),
    ('Educator Expense Deduction Max', 300),
    ('Student Loan Interest Deduction Max', 2500),
]
for label, val in thresholds:
    style_cell(ws_tt, r, 2, label, BLACK, LIGHT_GRAY, LEFT, THIN_BORDER)
    fmt = PCT_FMT if isinstance(val, float) and val < 1 else CURR_FMT
    style_cell(ws_tt, r, 3, val, BLACK, None, RIGHT, THIN_BORDER, fmt)
    r += 1

print(f"Tax Tables complete. Last row: {r}")


# ═══════════════════════════════════════════════════════════════════════════════
# TAB 3: FORM 1040
# ═══════════════════════════════════════════════════════════════════════════════
ws_1040 = wb.create_sheet('Form 1040')
ws_1040.sheet_properties.tabColor = '00B050'
ws_1040.column_dimensions['A'].width = 3
ws_1040.column_dimensions['B'].width = 8
ws_1040.column_dimensions['C'].width = 50
ws_1040.column_dimensions['D'].width = 20
ws_1040.column_dimensions['E'].width = 20
ws_1040.column_dimensions['F'].width = 25

r = 1
style_cell(ws_1040, r, 2, 'Form 1040 — U.S. Individual Income Tax Return (2025)', TITLE_FONT, None, LEFT, None, None, 6); r += 1
style_cell(ws_1040, r, 2, 'All amounts pull from Intake and Schedule tabs via formulas', SMALL_FONT); r += 2

section_header(ws_1040, r, 2, 5, 'INCOME'); r += 1
lines_income = [
    ('1a', 'Wages, salaries, tips (W-2, Box 1)', f"=Intake!D{TOTAL_WAGES_ROW}"),
    ('1b', 'Household employee income (not on W-2)', "=Intake!C{0}".format(OTHER_INC_START + 14)),
    ('1c', 'Tip income not on W-2', "=Intake!C{0}".format(OTHER_INC_START + 14)),
    ('1z', 'Total from lines 1a through 1c', '=D{r1}+D{r2}'),
    ('2a', 'Tax-exempt interest', f"=Intake!E{TOTAL_INT_ROW}"),
    ('2b', 'Taxable interest', f"=Intake!D{TOTAL_INT_ROW}"),
    ('3a', 'Qualified dividends', f"=Intake!D{TOTAL_QUAL_DIV_ROW}"),
    ('3b', 'Ordinary dividends', f"=Intake!C{TOTAL_ORD_DIV_ROW}"),
    ('4a', 'IRA distributions — Gross', f"=Intake!C{OTHER_INC_START + 2}"),
    ('4b', 'IRA distributions — Taxable', f"=Intake!C{OTHER_INC_START + 3}"),
    ('5a', 'Pensions and annuities — Gross', f"=Intake!C{OTHER_INC_START + 4}"),
    ('5b', 'Pensions and annuities — Taxable', f"=Intake!C{OTHER_INC_START + 5}"),
    ('6a', 'Social Security benefits — Gross', f"=Intake!C{OTHER_INC_START + 6}"),
    ('6b', 'Social Security benefits — Taxable', f"=Intake!C{OTHER_INC_START + 7}"),
    ('7', 'Capital gain or (loss) — Schedule D', "='Schedule D'!D50"),
    ('8', 'Other income from Schedule 1, line 10', "='Schedule 1'!D30"),
    ('9', 'Total income (add 1z, 2b, 3b, 4b, 5b, 6b, 7, 8)', 'SUM_INCOME'),
]

LINE_ROWS = {}
for line_num, desc, formula in lines_income:
    style_cell(ws_1040, r, 2, line_num, BLACK_BOLD, LIGHT_GRAY, CENTER, THIN_BORDER)
    style_cell(ws_1040, r, 3, desc, BLACK, LIGHT_GRAY, LEFT, THIN_BORDER)
    LINE_ROWS[line_num] = r
    if formula == 'SUM_INCOME':
        f_str = f"=D{LINE_ROWS['1z']}+D{LINE_ROWS['2b']}+D{LINE_ROWS['3b']}+D{LINE_ROWS['4b']}+D{LINE_ROWS['5b']}+D{LINE_ROWS['6b']}+D{LINE_ROWS['7']}+D{LINE_ROWS['8']}"
        style_cell(ws_1040, r, 4, f_str, BLACK_BOLD, LIGHT_GREEN, RIGHT, THICK_BOTTOM, CURR_FMT)
    elif line_num == '1z':
        f_str = f"=D{LINE_ROWS['1a']}+D{LINE_ROWS.get('1b', LINE_ROWS['1a'])}+D{LINE_ROWS.get('1c', LINE_ROWS['1a'])}"
        style_cell(ws_1040, r, 4, f_str, BLACK, LIGHT_BLUE, RIGHT, THIN_BORDER, CURR_FMT)
    else:
        font = GREEN_LINK if '!' in str(formula) else BLACK
        style_cell(ws_1040, r, 4, formula, font, LIGHT_BLUE, RIGHT, THIN_BORDER, CURR_FMT)
    r += 1
r += 1

# Adjusted Gross Income
section_header(ws_1040, r, 2, 5, 'ADJUSTED GROSS INCOME'); r += 1
agi_lines = [
    ('10', 'Adjustments from Schedule 1, line 26', "='Schedule 1'!D60"),
    ('11', 'Adjusted gross income (line 9 minus line 10)', 'CALC_AGI'),
]
for line_num, desc, formula in agi_lines:
    style_cell(ws_1040, r, 2, line_num, BLACK_BOLD, LIGHT_GRAY, CENTER, THIN_BORDER)
    style_cell(ws_1040, r, 3, desc, BLACK, LIGHT_GRAY, LEFT, THIN_BORDER)
    LINE_ROWS[line_num] = r
    if formula == 'CALC_AGI':
        style_cell(ws_1040, r, 4, f"=D{LINE_ROWS['9']}-D{LINE_ROWS['10']}", BLACK_BOLD, LIGHT_GREEN, RIGHT, THICK_BOTTOM, CURR_FMT)
    else:
        style_cell(ws_1040, r, 4, formula, GREEN_LINK, LIGHT_BLUE, RIGHT, THIN_BORDER, CURR_FMT)
    r += 1
r += 1

# Deductions & Taxable Income
section_header(ws_1040, r, 2, 5, 'DEDUCTIONS & TAXABLE INCOME'); r += 1
ded_lines = [
    ('12', 'Standard deduction or itemized deductions', "='Deduction Calc'!D10"),
    ('13', 'Qualified business income deduction (Sec 199A)', "=Intake!C{0}".format(ADJ_START)),
    ('14', 'Total deductions (line 12 + line 13)', 'SUM_DED'),
    ('15', 'Taxable income (line 11 minus line 14, not less than 0)', 'TAXABLE'),
]
for line_num, desc, formula in ded_lines:
    style_cell(ws_1040, r, 2, line_num, BLACK_BOLD, LIGHT_GRAY, CENTER, THIN_BORDER)
    style_cell(ws_1040, r, 3, desc, BLACK, LIGHT_GRAY, LEFT, THIN_BORDER)
    LINE_ROWS[line_num] = r
    if formula == 'SUM_DED':
        style_cell(ws_1040, r, 4, f"=D{LINE_ROWS['12']}+D{LINE_ROWS['13']}", BLACK, LIGHT_BLUE, RIGHT, THIN_BORDER, CURR_FMT)
    elif formula == 'TAXABLE':
        style_cell(ws_1040, r, 4, f"=MAX(0,D{LINE_ROWS['11']}-D{LINE_ROWS['14']})", BLACK_BOLD, LIGHT_GREEN, RIGHT, THICK_BOTTOM, CURR_FMT)
    else:
        style_cell(ws_1040, r, 4, formula, GREEN_LINK, LIGHT_BLUE, RIGHT, THIN_BORDER, CURR_FMT)
    r += 1
r += 1

# Tax Computation
section_header(ws_1040, r, 2, 5, 'TAX COMPUTATION'); r += 1
tax_lines = [
    ('16', 'Tax (from Tax Computation worksheet)', "='Tax Computation'!D30"),
    ('17', 'Amount from Schedule 2, line 21 (additional taxes)', "='Schedule 2'!D25"),
    ('18', 'Total tax before credits (line 16 + 17)', 'SUM_TAX'),
    ('19', 'Nonrefundable credits from Schedule 3, line 8', "='Schedule 3'!D15"),
    ('20', 'Subtract line 19 from 18 (not less than 0)', 'NET_TAX1'),
    ('21', 'Other taxes from Schedule 2, line 21', "='Schedule 2'!D25"),
    ('22', 'Total tax (line 20 + 21)', 'TOTAL_TAX'),
]
for line_num, desc, formula in tax_lines:
    style_cell(ws_1040, r, 2, line_num, BLACK_BOLD, LIGHT_GRAY, CENTER, THIN_BORDER)
    style_cell(ws_1040, r, 3, desc, BLACK, LIGHT_GRAY, LEFT, THIN_BORDER)
    LINE_ROWS[line_num] = r
    if formula == 'SUM_TAX':
        style_cell(ws_1040, r, 4, f"=D{LINE_ROWS['16']}+D{LINE_ROWS['17']}", BLACK, LIGHT_BLUE, RIGHT, THIN_BORDER, CURR_FMT)
    elif formula == 'NET_TAX1':
        style_cell(ws_1040, r, 4, f"=MAX(0,D{LINE_ROWS['18']}-D{LINE_ROWS['19']})", BLACK, LIGHT_BLUE, RIGHT, THIN_BORDER, CURR_FMT)
    elif formula == 'TOTAL_TAX':
        style_cell(ws_1040, r, 4, f"=D{LINE_ROWS['20']}+D{LINE_ROWS['21']}", BLACK_BOLD, LIGHT_GREEN, RIGHT, THICK_BOTTOM, CURR_FMT)
    else:
        style_cell(ws_1040, r, 4, formula, GREEN_LINK, LIGHT_BLUE, RIGHT, THIN_BORDER, CURR_FMT)
    r += 1
r += 1

# Payments
section_header(ws_1040, r, 2, 5, 'PAYMENTS'); r += 1
pay_lines = [
    ('23', 'Federal income tax withheld', f"=Intake!C{PAY_START}"),
    ('24', 'Estimated tax payments + amount applied from prior year',
     f"=Intake!C{PAY_START+1}+Intake!C{PAY_START+2}+Intake!C{PAY_START+3}+Intake!C{PAY_START+4}+Intake!C{PAY_START+5}"),
    ('25', 'Earned income credit', f"=Intake!C{CREDIT_START+12}"),
    ('26', 'Additional child tax credit (Form 8812)', f"=Intake!C{CREDIT_START+9}"),
    ('27', 'American opportunity credit (Form 8863)', f"=Intake!C{CREDIT_START+10}"),
    ('28', 'Other refundable credits (Schedule 3)', "='Schedule 3'!D25"),
    ('29', 'Total other payments and refundable credits', f"=D{0}"),
    ('30', 'Extension payment (Form 4868)', f"=Intake!C{PAY_START+6}"),
    ('31', 'Excess SS tax withheld', f"=Intake!C{PAY_START+7}"),
    ('32', 'Other payments', f"=Intake!C{PAY_START+8}"),
    ('33', 'Total payments', 'TOTAL_PAY'),
]
for line_num, desc, formula in pay_lines:
    style_cell(ws_1040, r, 2, line_num, BLACK_BOLD, LIGHT_GRAY, CENTER, THIN_BORDER)
    style_cell(ws_1040, r, 3, desc, BLACK, LIGHT_GRAY, LEFT, THIN_BORDER)
    LINE_ROWS[line_num] = r
    if formula == 'TOTAL_PAY':
        style_cell(ws_1040, r, 4, f"=SUM(D{LINE_ROWS['23']}:D{LINE_ROWS['32']})", BLACK_BOLD, LIGHT_GREEN, RIGHT, THICK_BOTTOM, CURR_FMT)
    elif line_num == '29':
        style_cell(ws_1040, r, 4, f"=D{LINE_ROWS['25']}+D{LINE_ROWS['26']}+D{LINE_ROWS['27']}+D{LINE_ROWS['28']}", BLACK, LIGHT_BLUE, RIGHT, THIN_BORDER, CURR_FMT)
    else:
        font = GREEN_LINK if '!' in str(formula) else BLACK
        style_cell(ws_1040, r, 4, formula, font, LIGHT_BLUE, RIGHT, THIN_BORDER, CURR_FMT)
    r += 1
r += 1

# Refund or Amount Owed
section_header(ws_1040, r, 2, 5, 'REFUND OR AMOUNT OWED'); r += 1
style_cell(ws_1040, r, 2, '34', BLACK_BOLD, LIGHT_GRAY, CENTER, THIN_BORDER)
style_cell(ws_1040, r, 3, 'Overpayment (if line 33 > line 22)', BLACK, LIGHT_GRAY, LEFT, THIN_BORDER)
LINE_ROWS['34'] = r
style_cell(ws_1040, r, 4, f"=IF(D{LINE_ROWS['33']}>D{LINE_ROWS['22']},D{LINE_ROWS['33']}-D{LINE_ROWS['22']},0)", BLACK_BOLD, LIGHT_GREEN, RIGHT, THIN_BORDER, CURR_FMT)
r += 1
style_cell(ws_1040, r, 2, '35a', BLACK_BOLD, LIGHT_GRAY, CENTER, THIN_BORDER)
style_cell(ws_1040, r, 3, 'Amount of line 34 you want REFUNDED to you', BLACK, LIGHT_GRAY, LEFT, THIN_BORDER)
LINE_ROWS['35a'] = r
style_cell(ws_1040, r, 4, f"=D{LINE_ROWS['34']}", BLACK, LIGHT_BLUE, RIGHT, THIN_BORDER, CURR_FMT)
r += 1
style_cell(ws_1040, r, 2, '37', BLACK_BOLD, LIGHT_GRAY, CENTER, THIN_BORDER)
style_cell(ws_1040, r, 3, 'Amount you OWE (if line 22 > line 33)', BLACK, LIGHT_GRAY, LEFT, THIN_BORDER)
LINE_ROWS['37'] = r
style_cell(ws_1040, r, 4, f"=IF(D{LINE_ROWS['22']}>D{LINE_ROWS['33']},D{LINE_ROWS['22']}-D{LINE_ROWS['33']},0)",
           Font(name='Arial', size=10, color='FF0000', bold=True), LIGHT_RED, RIGHT, THICK_BOTTOM, CURR_FMT)
r += 1
style_cell(ws_1040, r, 2, '38', BLACK_BOLD, LIGHT_GRAY, CENTER, THIN_BORDER)
style_cell(ws_1040, r, 3, 'Estimated tax penalty', BLACK, LIGHT_GRAY, LEFT, THIN_BORDER)
LINE_ROWS['38'] = r
style_cell(ws_1040, r, 4, None, BLUE_INPUT, LIGHT_YELLOW, RIGHT, THIN_BORDER, CURR_FMT)

print(f"Form 1040 complete. Last row: {r}")


# ═══════════════════════════════════════════════════════════════════════════════
# TAB: DEDUCTION CALC (helper for Standard vs Itemized)
# ═══════════════════════════════════════════════════════════════════════════════
ws_ded = wb.create_sheet('Deduction Calc')
ws_ded.sheet_properties.tabColor = 'FFC000'
ws_ded.column_dimensions['A'].width = 3
ws_ded.column_dimensions['B'].width = 5
ws_ded.column_dimensions['C'].width = 50
ws_ded.column_dimensions['D'].width = 20

r = 1
style_cell(ws_ded, r, 2, 'Deduction Determination Worksheet', TITLE_FONT); r += 2
section_header(ws_ded, r, 2, 4, 'STANDARD DEDUCTION CALCULATION'); r += 1

style_cell(ws_ded, r, 2, '1', BLACK_BOLD, LIGHT_GRAY, CENTER, THIN_BORDER)
style_cell(ws_ded, r, 3, 'Filing Status', BLACK, LIGHT_GRAY, LEFT, THIN_BORDER)
style_cell(ws_ded, r, 4, "=Intake!C6", GREEN_LINK, LIGHT_BLUE, CENTER, THIN_BORDER)
FS_ROW = r; r += 1

style_cell(ws_ded, r, 2, '2', BLACK_BOLD, LIGHT_GRAY, CENTER, THIN_BORDER)
style_cell(ws_ded, r, 3, 'Base standard deduction for filing status', BLACK, LIGHT_GRAY, LEFT, THIN_BORDER)
# VLOOKUP: 1=Single(15750), 2=MFJ(31500), 3=MFS(15750), 4=HOH(23625), 5=QSS(31500)
style_cell(ws_ded, r, 4,
    f'=IF(D{FS_ROW}=1,15750,IF(D{FS_ROW}=2,31500,IF(D{FS_ROW}=3,15750,IF(D{FS_ROW}=4,23625,IF(D{FS_ROW}=5,31500,0)))))',
    BLACK, LIGHT_BLUE, RIGHT, THIN_BORDER, CURR_FMT)
BASE_SD_ROW = r; r += 1

style_cell(ws_ded, r, 2, '3', BLACK_BOLD, LIGHT_GRAY, CENTER, THIN_BORDER)
style_cell(ws_ded, r, 3, 'Additional for age 65+ (primary)', BLACK, LIGHT_GRAY, LEFT, THIN_BORDER)
style_cell(ws_ded, r, 4,
    f'=IF(Intake!C17="Y",IF(OR(D{FS_ROW}=1,D{FS_ROW}=4),2000,1600),0)',
    BLACK, LIGHT_BLUE, RIGHT, THIN_BORDER, CURR_FMT)
ADD_SD1 = r; r += 1

style_cell(ws_ded, r, 2, '4', BLACK_BOLD, LIGHT_GRAY, CENTER, THIN_BORDER)
style_cell(ws_ded, r, 3, 'Additional for age 65+ (spouse, if MFJ)', BLACK, LIGHT_GRAY, LEFT, THIN_BORDER)
style_cell(ws_ded, r, 4,
    f'=IF(AND(D{FS_ROW}=2,Intake!E17="Y"),1600,0)',
    BLACK, LIGHT_BLUE, RIGHT, THIN_BORDER, CURR_FMT)
ADD_SD2 = r; r += 1

style_cell(ws_ded, r, 2, '5', BLACK_BOLD, LIGHT_GRAY, CENTER, THIN_BORDER)
style_cell(ws_ded, r, 3, 'Additional for blindness (primary)', BLACK, LIGHT_GRAY, LEFT, THIN_BORDER)
style_cell(ws_ded, r, 4,
    f'=IF(Intake!C16="Y",IF(OR(D{FS_ROW}=1,D{FS_ROW}=4),2000,1600),0)',
    BLACK, LIGHT_BLUE, RIGHT, THIN_BORDER, CURR_FMT)
ADD_SD3 = r; r += 1

style_cell(ws_ded, r, 2, '6', BLACK_BOLD, LIGHT_GRAY, CENTER, THIN_BORDER)
style_cell(ws_ded, r, 3, 'Total standard deduction', BLACK_BOLD, LIGHT_GRAY, LEFT, THIN_BORDER)
style_cell(ws_ded, r, 4, f'=SUM(D{BASE_SD_ROW}:D{ADD_SD3})', BLACK_BOLD, LIGHT_GREEN, RIGHT, THICK_BOTTOM, CURR_FMT)
TOT_SD_ROW = r; r += 2

section_header(ws_ded, r, 2, 4, 'ITEMIZED DEDUCTIONS TOTAL'); r += 1
style_cell(ws_ded, r, 2, '7', BLACK_BOLD, LIGHT_GRAY, CENTER, THIN_BORDER)
style_cell(ws_ded, r, 3, 'Total itemized deductions (from Schedule A)', BLACK, LIGHT_GRAY, LEFT, THIN_BORDER)
style_cell(ws_ded, r, 4, "='Schedule A'!D30", GREEN_LINK, LIGHT_BLUE, RIGHT, THIN_BORDER, CURR_FMT)
ITEM_ROW = r; r += 2

section_header(ws_ded, r, 2, 4, 'DEDUCTION USED ON RETURN'); r += 1
# D10 is the key cell referenced by Form 1040
style_cell(ws_ded, r, 2, '8', BLACK_BOLD, LIGHT_GRAY, CENTER, THIN_BORDER)
style_cell(ws_ded, r, 3, 'Greater of standard or itemized (or per taxpayer choice)', BLACK_BOLD, LIGHT_GRAY, LEFT, THIN_BORDER)
style_cell(ws_ded, r, 4,
    f'=IF(Intake!C8="I",D{ITEM_ROW},IF(Intake!C8="S",D{TOT_SD_ROW},MAX(D{TOT_SD_ROW},D{ITEM_ROW})))',
    BLACK_BOLD, LIGHT_GREEN, RIGHT, THICK_BOTTOM, CURR_FMT)

print("Deduction Calc complete.")


# ═══════════════════════════════════════════════════════════════════════════════
# TAB: TAX COMPUTATION
# ═══════════════════════════════════════════════════════════════════════════════
ws_tc = wb.create_sheet('Tax Computation')
ws_tc.sheet_properties.tabColor = 'FF0000'
ws_tc.column_dimensions['A'].width = 3
ws_tc.column_dimensions['B'].width = 5
ws_tc.column_dimensions['C'].width = 55
ws_tc.column_dimensions['D'].width = 20

r = 1
style_cell(ws_tc, r, 2, 'Tax Computation Worksheet — 2025', TITLE_FONT); r += 2
style_cell(ws_tc, r, 2, 'Calculates ordinary income tax using progressive bracket rates', SMALL_FONT); r += 2

section_header(ws_tc, r, 2, 4, 'INPUTS'); r += 1
style_cell(ws_tc, r, 3, 'Filing Status', BLACK, LIGHT_GRAY, LEFT, THIN_BORDER)
style_cell(ws_tc, r, 4, "=Intake!C6", GREEN_LINK, LIGHT_BLUE, CENTER, THIN_BORDER)
TC_FS = r; r += 1
style_cell(ws_tc, r, 3, 'Taxable Income (from Form 1040, Line 15)', BLACK, LIGHT_GRAY, LEFT, THIN_BORDER)
style_cell(ws_tc, r, 4, f"='Form 1040'!D{LINE_ROWS['15']}", GREEN_LINK, LIGHT_BLUE, RIGHT, THIN_BORDER, CURR_FMT)
TC_TI = r; r += 2

section_header(ws_tc, r, 2, 4, 'PROGRESSIVE TAX CALCULATION'); r += 1

# Use the numeric bracket thresholds from Tax Tables
# For each bracket, compute tax = rate * (min(income, upper) - lower), if income > lower
bracket_info = [
    (0.10, 'B', 'C', 'D', 'E', 'F'),  # cols in Tax Tables numeric section
]

# Build bracket calculation rows
# We'll reference Tax Tables for thresholds
# Filing status determines which column to use: Single=C/D, MFJ=E/F, HOH=G
style_cell(ws_tc, r, 2, '#', HEADER_FONT, MEDIUM_BLUE, CENTER, THIN_BORDER)
style_cell(ws_tc, r, 3, 'Bracket Description', HEADER_FONT, MEDIUM_BLUE, CENTER, THIN_BORDER)
style_cell(ws_tc, r, 4, 'Tax in Bracket', HEADER_FONT, MEDIUM_BLUE, CENTER, THIN_BORDER)
r += 1

# Hardcode 7 bracket calculation rows
rates = [0.10, 0.12, 0.22, 0.24, 0.32, 0.35, 0.37]
single_brackets = [(0, 11925), (11925, 48475), (48475, 103350), (103350, 197300),
                   (197300, 250525), (250525, 626350), (626350, 99999999)]
mfj_brackets = [(0, 23850), (23850, 96950), (96950, 206700), (206700, 394600),
                (394600, 501050), (501050, 751600), (751600, 99999999)]
hoh_brackets = [(0, 17000), (17000, 64850), (64850, 103350), (103350, 197300),
                (197300, 250500), (250500, 626350), (626350, 99999999)]

TC_BRACKET_START = r
for i, rate in enumerate(rates):
    pct = int(rate * 100)
    sl, su = single_brackets[i]
    ml, mu = mfj_brackets[i]
    hl, hu = hoh_brackets[i]
    style_cell(ws_tc, r, 2, str(i+1), BLACK_BOLD, LIGHT_GRAY, CENTER, THIN_BORDER)
    style_cell(ws_tc, r, 3, f'{pct}% bracket', BLACK, LIGHT_GRAY, LEFT, THIN_BORDER)
    # Formula: IF filing_status=1 use single, =2 use MFJ, =3 use MFS (same as single), =4 use HOH
    formula = (
        f'=IF(D${TC_TI}>0,'
        f'IF(OR(D${TC_FS}=1,D${TC_FS}=3),'
        f'{rate}*MAX(0,MIN(D${TC_TI},{su})-{sl}),'
        f'IF(D${TC_FS}=2,'
        f'{rate}*MAX(0,MIN(D${TC_TI},{mu})-{ml}),'
        f'IF(D${TC_FS}=4,'
        f'{rate}*MAX(0,MIN(D${TC_TI},{hu})-{hl}),'
        f'{rate}*MAX(0,MIN(D${TC_TI},{su})-{sl})))),'
        f'0)'
    )
    style_cell(ws_tc, r, 4, formula, BLACK, LIGHT_BLUE, RIGHT, THIN_BORDER, CURR_FMT)
    r += 1
TC_BRACKET_END = r - 1
r += 1

style_cell(ws_tc, r, 2, '', BLACK_BOLD, LIGHT_GRAY, CENTER, THICK_BOTTOM)
style_cell(ws_tc, r, 3, 'Total Ordinary Income Tax', BLACK_BOLD, LIGHT_GRAY, LEFT, THICK_BOTTOM)
style_cell(ws_tc, r, 4, f'=SUM(D{TC_BRACKET_START}:D{TC_BRACKET_END})', BLACK_BOLD, LIGHT_GREEN, RIGHT, THICK_BOTTOM, CURR_FMT)
TC_TOTAL_ORD = r; r += 2

# Qualified dividends / LTCG preferential rate (simplified)
section_header(ws_tc, r, 2, 4, 'QUALIFIED DIVIDENDS & LTCG TAX (SIMPLIFIED)'); r += 1
style_cell(ws_tc, r, 3, 'Qualified dividends', BLACK, LIGHT_GRAY, LEFT, THIN_BORDER)
style_cell(ws_tc, r, 4, f"='Form 1040'!D{LINE_ROWS['3a']}", GREEN_LINK, LIGHT_BLUE, RIGHT, THIN_BORDER, CURR_FMT)
QD_ROW = r; r += 1
style_cell(ws_tc, r, 3, 'Net LTCG from Schedule D', BLACK, LIGHT_GRAY, LEFT, THIN_BORDER)
style_cell(ws_tc, r, 4, "='Schedule D'!D48", GREEN_LINK, LIGHT_BLUE, RIGHT, THIN_BORDER, CURR_FMT)
LTCG_ROW = r; r += 1
style_cell(ws_tc, r, 3, 'Total preferential rate income', BLACK, LIGHT_GRAY, LEFT, THIN_BORDER)
style_cell(ws_tc, r, 4, f'=D{QD_ROW}+MAX(0,D{LTCG_ROW})', BLACK, LIGHT_BLUE, RIGHT, THIN_BORDER, CURR_FMT)
PREF_INC = r; r += 1
style_cell(ws_tc, r, 3, 'Note: Full QDCG worksheet computation recommended for complex returns', SMALL_FONT)
r += 2

# Line 30 is the key output row referenced by Form 1040
# For simplicity, use ordinary rates on full taxable income (agent can enhance)
while r < 30:
    r += 1
style_cell(ws_tc, 30, 2, '', BLACK_BOLD, LIGHT_GRAY, CENTER, THICK_BOTTOM)
style_cell(ws_tc, 30, 3, 'TAX (Line 16 of Form 1040)', BLACK_BOLD, LIGHT_GRAY, LEFT, THICK_BOTTOM)
style_cell(ws_tc, 30, 4, f'=D{TC_TOTAL_ORD}', BLACK_BOLD, LIGHT_GREEN, RIGHT, THICK_BOTTOM, CURR_FMT)

print("Tax Computation complete.")


# ═══════════════════════════════════════════════════════════════════════════════
# TAB: SCHEDULE 1
# ═══════════════════════════════════════════════════════════════════════════════
ws_s1 = wb.create_sheet('Schedule 1')
ws_s1.sheet_properties.tabColor = '4472C4'
ws_s1.column_dimensions['A'].width = 3
ws_s1.column_dimensions['B'].width = 5
ws_s1.column_dimensions['C'].width = 55
ws_s1.column_dimensions['D'].width = 20

r = 1
style_cell(ws_s1, r, 2, 'Schedule 1 — Additional Income and Adjustments to Income', TITLE_FONT); r += 2

section_header(ws_s1, r, 2, 4, 'PART I — ADDITIONAL INCOME'); r += 1
s1_income = [
    ('1', 'Taxable refunds of state/local taxes', f"=Intake!C{OTHER_INC_START}"),
    ('2a', 'Alimony received (pre-2019 divorce)', f"=Intake!C{OTHER_INC_START+1}"),
    ('3', 'Business income or (loss) — Schedule C', "='Schedule C'!D45"),
    ('4', 'Other gains or (losses) — Form 4797', None),
    ('5', 'Rental real estate, royalties, partnerships, S corps — Schedule E', "='Schedule E'!D50"),
    ('6', 'Farm income or (loss) — Schedule F', f"=Intake!C{OTHER_INC_START+12}"),
    ('7', 'Unemployment compensation', f"=Intake!C{OTHER_INC_START+8}"),
    ('8a', 'Net operating loss', None),
    ('8b', 'Gambling income', f"=Intake!C{OTHER_INC_START+9}"),
    ('8c', 'Cancellation of debt', None),
    ('8d', 'Foreign earned income exclusion (Form 2555)', None),
    ('8z', 'Other income (list type and amount)', f"=Intake!C{OTHER_INC_START+11}"),
    ('9', 'Total other income (sum 8a through 8z)', 'SUM8'),
    ('10', 'Total additional income (add lines 1 through 7 plus 9)', 'SUM_ALL'),
]

S1_ROWS = {}
for line_num, desc, formula in s1_income:
    style_cell(ws_s1, r, 2, line_num, BLACK_BOLD, LIGHT_GRAY, CENTER, THIN_BORDER)
    style_cell(ws_s1, r, 3, desc, BLACK, LIGHT_GRAY, LEFT, THIN_BORDER)
    S1_ROWS[line_num] = r
    if formula == 'SUM8':
        style_cell(ws_s1, r, 4, f"=SUM(D{S1_ROWS.get('8a',r)}:D{S1_ROWS.get('8z',r)})", BLACK, LIGHT_BLUE, RIGHT, THIN_BORDER, CURR_FMT)
    elif formula == 'SUM_ALL':
        style_cell(ws_s1, r, 4,
            f"=D{S1_ROWS['1']}+D{S1_ROWS['2a']}+D{S1_ROWS['3']}+D{S1_ROWS.get('4',r)}+D{S1_ROWS['5']}+D{S1_ROWS['6']}+D{S1_ROWS['7']}+D{S1_ROWS['9']}",
            BLACK_BOLD, LIGHT_GREEN, RIGHT, THICK_BOTTOM, CURR_FMT)
    elif formula:
        font = GREEN_LINK if '!' in str(formula) else BLACK
        style_cell(ws_s1, r, 4, formula, font, LIGHT_BLUE, RIGHT, THIN_BORDER, CURR_FMT)
    else:
        style_cell(ws_s1, r, 4, None, BLUE_INPUT, LIGHT_YELLOW, RIGHT, THIN_BORDER, CURR_FMT)
    r += 1

# Ensure line 30 = row 30 for the reference from Form 1040
while r < 28:
    r += 1
r = 28
style_cell(ws_s1, r, 3, 'Carried to Form 1040, Line 8', SMALL_FONT)
r += 1
style_cell(ws_s1, 30, 2, '10', BLACK_BOLD, LIGHT_GRAY, CENTER, THICK_BOTTOM)
style_cell(ws_s1, 30, 3, 'Total additional income → Form 1040 Line 8', BLACK_BOLD, LIGHT_GRAY, LEFT, THICK_BOTTOM)
style_cell(ws_s1, 30, 4, f"=D{S1_ROWS['10']}", BLACK_BOLD, LIGHT_GREEN, RIGHT, THICK_BOTTOM, CURR_FMT)
r = 32

section_header(ws_s1, r, 2, 4, 'PART II — ADJUSTMENTS TO INCOME'); r += 1
s1_adj = [
    ('11', 'Educator expenses', f"=Intake!C{ADJ_START}"),
    ('12', 'Business expenses (reservists, etc.)', f"=Intake!C{ADJ_START+1}"),
    ('13', 'HSA deduction', f"=Intake!C{ADJ_START+2}"),
    ('14', 'Moving expenses (military)', f"=Intake!C{ADJ_START+3}"),
    ('15', 'Deductible part of self-employment tax', "='Schedule SE'!D20"),
    ('16', 'Self-employed SEP/SIMPLE/qualified plans', f"=Intake!C{ADJ_START+5}"),
    ('17', 'Self-employed health insurance', f"=Intake!C{ADJ_START+6}"),
    ('18', 'Penalty on early withdrawal of savings', f"=Intake!C{ADJ_START+7}"),
    ('19', 'IRA deduction', f"=Intake!C{ADJ_START+8}"),
    ('20', 'Student loan interest deduction', f"=Intake!C{ADJ_START+9}"),
    ('22', 'Other adjustments', f"=Intake!C{ADJ_START+13}"),
    ('25', 'Total Part II adjustments', 'SUM_ADJ'),
    ('26', 'Total adjustments → Form 1040 Line 10', 'TOTAL_ADJ'),
]

S1_ADJ_ROWS = {}
for line_num, desc, formula in s1_adj:
    style_cell(ws_s1, r, 2, line_num, BLACK_BOLD, LIGHT_GRAY, CENTER, THIN_BORDER)
    style_cell(ws_s1, r, 3, desc, BLACK, LIGHT_GRAY, LEFT, THIN_BORDER)
    S1_ADJ_ROWS[line_num] = r
    if formula == 'SUM_ADJ':
        style_cell(ws_s1, r, 4,
            f"=SUM(D{S1_ADJ_ROWS['11']}:D{S1_ADJ_ROWS['22']})",
            BLACK, LIGHT_BLUE, RIGHT, THIN_BORDER, CURR_FMT)
    elif formula == 'TOTAL_ADJ':
        style_cell(ws_s1, r, 4, f"=D{S1_ADJ_ROWS['25']}", BLACK_BOLD, LIGHT_GREEN, RIGHT, THICK_BOTTOM, CURR_FMT)
    else:
        font = GREEN_LINK if '!' in str(formula) else BLACK
        style_cell(ws_s1, r, 4, formula, font, LIGHT_BLUE, RIGHT, THIN_BORDER, CURR_FMT)
    r += 1

# Put total at row 60 for Form 1040 reference
while r < 60:
    r += 1
style_cell(ws_s1, 60, 2, '26', BLACK_BOLD, LIGHT_GRAY, CENTER, THICK_BOTTOM)
style_cell(ws_s1, 60, 3, 'Total adjustments to income → Form 1040 Line 10', BLACK_BOLD, LIGHT_GRAY, LEFT, THICK_BOTTOM)
style_cell(ws_s1, 60, 4, f"=D{S1_ADJ_ROWS['26']}", BLACK_BOLD, LIGHT_GREEN, RIGHT, THICK_BOTTOM, CURR_FMT)

print("Schedule 1 complete.")


# ═══════════════════════════════════════════════════════════════════════════════
# TAB: SCHEDULE 2 — Additional Taxes
# ═══════════════════════════════════════════════════════════════════════════════
ws_s2 = wb.create_sheet('Schedule 2')
ws_s2.sheet_properties.tabColor = 'ED7D31'
ws_s2.column_dimensions['A'].width = 3
ws_s2.column_dimensions['B'].width = 5
ws_s2.column_dimensions['C'].width = 55
ws_s2.column_dimensions['D'].width = 20

r = 1
style_cell(ws_s2, r, 2, 'Schedule 2 — Additional Taxes', TITLE_FONT); r += 2

section_header(ws_s2, r, 2, 4, 'PART I — TAX'); r += 1
style_cell(ws_s2, r, 2, '1', BLACK_BOLD, LIGHT_GRAY, CENTER, THIN_BORDER)
style_cell(ws_s2, r, 3, 'AMT (Form 6251)', BLACK, LIGHT_GRAY, LEFT, THIN_BORDER)
style_cell(ws_s2, r, 4, None, BLUE_INPUT, LIGHT_YELLOW, RIGHT, THIN_BORDER, CURR_FMT)
S2_AMT = r; r += 1
style_cell(ws_s2, r, 2, '2', BLACK_BOLD, LIGHT_GRAY, CENTER, THIN_BORDER)
style_cell(ws_s2, r, 3, 'Excess premium tax credit repayment', BLACK, LIGHT_GRAY, LEFT, THIN_BORDER)
style_cell(ws_s2, r, 4, None, BLUE_INPUT, LIGHT_YELLOW, RIGHT, THIN_BORDER, CURR_FMT)
S2_PTC = r; r += 1
style_cell(ws_s2, r, 2, '3', BLACK_BOLD, LIGHT_GRAY, CENTER, THIN_BORDER)
style_cell(ws_s2, r, 3, 'Total (add lines 1 and 2) → Form 1040 Line 17', BLACK_BOLD, LIGHT_GRAY, LEFT, THIN_BORDER)
style_cell(ws_s2, r, 4, f'=D{S2_AMT}+D{S2_PTC}', BLACK_BOLD, LIGHT_GREEN, RIGHT, THICK_BOTTOM, CURR_FMT)
r += 2

section_header(ws_s2, r, 2, 4, 'PART II — OTHER TAXES'); r += 1
s2_other = [
    ('6', 'Self-employment tax (Schedule SE)', "='Schedule SE'!D18"),
    ('8', 'Additional Medicare Tax (Form 8959)', None),
    ('9', 'Net investment income tax (Form 8960)', None),
    ('10', 'Uncollected SS/Medicare on tips', None),
    ('11', 'Additional tax on IRAs/retirement (10%)', None),
    ('12', 'Additional tax on HSA distributions', None),
    ('17', 'Total additional taxes (sum)', 'SUM_S2'),
    ('21', 'Total Other Taxes → Form 1040 Line 21', 'TOTAL_S2'),
]
S2_ROWS = {}
for line_num, desc, formula in s2_other:
    style_cell(ws_s2, r, 2, line_num, BLACK_BOLD, LIGHT_GRAY, CENTER, THIN_BORDER)
    style_cell(ws_s2, r, 3, desc, BLACK, LIGHT_GRAY, LEFT, THIN_BORDER)
    S2_ROWS[line_num] = r
    if formula == 'SUM_S2':
        style_cell(ws_s2, r, 4, f"=SUM(D{S2_ROWS['6']}:D{S2_ROWS['12']})", BLACK, LIGHT_BLUE, RIGHT, THIN_BORDER, CURR_FMT)
    elif formula == 'TOTAL_S2':
        style_cell(ws_s2, r, 4, f"=D{S2_ROWS['17']}", BLACK_BOLD, LIGHT_GREEN, RIGHT, THICK_BOTTOM, CURR_FMT)
    elif formula:
        style_cell(ws_s2, r, 4, formula, GREEN_LINK, LIGHT_BLUE, RIGHT, THIN_BORDER, CURR_FMT)
    else:
        style_cell(ws_s2, r, 4, None, BLUE_INPUT, LIGHT_YELLOW, RIGHT, THIN_BORDER, CURR_FMT)
    r += 1

# Set row 25 as the output row
while r <= 25:
    r += 1
style_cell(ws_s2, 25, 2, '21', BLACK_BOLD, LIGHT_GRAY, CENTER, THICK_BOTTOM)
style_cell(ws_s2, 25, 3, 'Schedule 2 Total → Form 1040', BLACK_BOLD, LIGHT_GRAY, LEFT, THICK_BOTTOM)
style_cell(ws_s2, 25, 4, f"=D{S2_ROWS['21']}", BLACK_BOLD, LIGHT_GREEN, RIGHT, THICK_BOTTOM, CURR_FMT)

print("Schedule 2 complete.")


# ═══════════════════════════════════════════════════════════════════════════════
# TAB: SCHEDULE 3 — Additional Credits and Payments
# ═══════════════════════════════════════════════════════════════════════════════
ws_s3 = wb.create_sheet('Schedule 3')
ws_s3.sheet_properties.tabColor = '70AD47'
ws_s3.column_dimensions['A'].width = 3
ws_s3.column_dimensions['B'].width = 5
ws_s3.column_dimensions['C'].width = 55
ws_s3.column_dimensions['D'].width = 20

r = 1
style_cell(ws_s3, r, 2, 'Schedule 3 — Additional Credits and Payments', TITLE_FONT); r += 2

section_header(ws_s3, r, 2, 4, 'PART I — NONREFUNDABLE CREDITS'); r += 1
s3_credits = [
    ('1', 'Foreign tax credit (Form 1116)', f"=Intake!C{CREDIT_START}"),
    ('2', 'Child and dependent care credit (Form 2441)', f"=Intake!C{CREDIT_START+1}"),
    ('3', 'Education credits (Form 8863)', f"=Intake!C{CREDIT_START+2}+Intake!C{CREDIT_START+3}"),
    ('4', 'Retirement savings credit (Form 8880)', f"=Intake!C{CREDIT_START+4}"),
    ('5a', 'Residential energy credit', f"=Intake!C{CREDIT_START+6}"),
    ('5b', 'EV credit', f"=Intake!C{CREDIT_START+7}"),
    ('7', 'Other nonrefundable credits', f"=Intake!C{CREDIT_START+8}"),
    ('8', 'Total nonrefundable credits → Form 1040 Line 19', 'SUM_NR'),
]
S3_ROWS = {}
for line_num, desc, formula in s3_credits:
    style_cell(ws_s3, r, 2, line_num, BLACK_BOLD, LIGHT_GRAY, CENTER, THIN_BORDER)
    style_cell(ws_s3, r, 3, desc, BLACK, LIGHT_GRAY, LEFT, THIN_BORDER)
    S3_ROWS[line_num] = r
    if formula == 'SUM_NR':
        style_cell(ws_s3, r, 4, f"=SUM(D{S3_ROWS['1']}:D{S3_ROWS['7']})", BLACK_BOLD, LIGHT_GREEN, RIGHT, THICK_BOTTOM, CURR_FMT)
    else:
        font = GREEN_LINK if '!' in str(formula) else BLUE_INPUT
        fill = LIGHT_BLUE if '!' in str(formula) else LIGHT_YELLOW
        style_cell(ws_s3, r, 4, formula, font, fill, RIGHT, THIN_BORDER, CURR_FMT)
    r += 1

# Row 15 = output for Form 1040
while r <= 15:
    r += 1
style_cell(ws_s3, 15, 2, '8', BLACK_BOLD, LIGHT_GRAY, CENTER, THICK_BOTTOM)
style_cell(ws_s3, 15, 3, 'Total nonrefundable credits → Form 1040 Line 19', BLACK_BOLD, LIGHT_GRAY, LEFT, THICK_BOTTOM)
style_cell(ws_s3, 15, 4, f"=D{S3_ROWS['8']}", BLACK_BOLD, LIGHT_GREEN, RIGHT, THICK_BOTTOM, CURR_FMT)

r = 17
section_header(ws_s3, r, 2, 4, 'PART II — OTHER PAYMENTS AND REFUNDABLE CREDITS'); r += 1
s3_refund = [
    ('9', 'Net premium tax credit (Form 8962)', f"=Intake!C{CREDIT_START+11}"),
    ('10', 'Amount paid with extension', f"=Intake!C{PAY_START+6}"),
    ('11', 'Excess SS tax withheld', f"=Intake!C{PAY_START+7}"),
    ('13', 'Other refundable credits', f"=Intake!C{CREDIT_START+13}"),
    ('15', 'Total other payments → Form 1040', 'SUM_REF'),
]
S3_REF_ROWS = {}
for line_num, desc, formula in s3_refund:
    style_cell(ws_s3, r, 2, line_num, BLACK_BOLD, LIGHT_GRAY, CENTER, THIN_BORDER)
    style_cell(ws_s3, r, 3, desc, BLACK, LIGHT_GRAY, LEFT, THIN_BORDER)
    S3_REF_ROWS[line_num] = r
    if formula == 'SUM_REF':
        style_cell(ws_s3, r, 4, f"=SUM(D{S3_REF_ROWS['9']}:D{S3_REF_ROWS['13']})", BLACK_BOLD, LIGHT_GREEN, RIGHT, THICK_BOTTOM, CURR_FMT)
    else:
        style_cell(ws_s3, r, 4, formula, GREEN_LINK, LIGHT_BLUE, RIGHT, THIN_BORDER, CURR_FMT)
    r += 1

# Row 25 = output
while r <= 25:
    r += 1
style_cell(ws_s3, 25, 2, '15', BLACK_BOLD, LIGHT_GRAY, CENTER, THICK_BOTTOM)
style_cell(ws_s3, 25, 3, 'Total other payments/refundable credits → Form 1040', BLACK_BOLD, LIGHT_GRAY, LEFT, THICK_BOTTOM)
style_cell(ws_s3, 25, 4, f"=D{S3_REF_ROWS['15']}", BLACK_BOLD, LIGHT_GREEN, RIGHT, THICK_BOTTOM, CURR_FMT)

print("Schedule 3 complete.")


# ═══════════════════════════════════════════════════════════════════════════════
# TAB: SCHEDULE A — Itemized Deductions
# ═══════════════════════════════════════════════════════════════════════════════
ws_sa = wb.create_sheet('Schedule A')
ws_sa.sheet_properties.tabColor = 'A5A5A5'
ws_sa.column_dimensions['A'].width = 3
ws_sa.column_dimensions['B'].width = 5
ws_sa.column_dimensions['C'].width = 55
ws_sa.column_dimensions['D'].width = 20

r = 1
style_cell(ws_sa, r, 2, 'Schedule A — Itemized Deductions', TITLE_FONT); r += 2

section_header(ws_sa, r, 2, 4, 'MEDICAL AND DENTAL EXPENSES'); r += 1
style_cell(ws_sa, r, 2, '1', BLACK_BOLD, LIGHT_GRAY, CENTER, THIN_BORDER)
style_cell(ws_sa, r, 3, 'Medical and dental expenses', BLACK, LIGHT_GRAY, LEFT, THIN_BORDER)
style_cell(ws_sa, r, 4, f"=Intake!C{SCHA_START}", GREEN_LINK, LIGHT_BLUE, RIGHT, THIN_BORDER, CURR_FMT)
SA_MED = r; r += 1
style_cell(ws_sa, r, 2, '2', BLACK_BOLD, LIGHT_GRAY, CENTER, THIN_BORDER)
style_cell(ws_sa, r, 3, 'AGI', BLACK, LIGHT_GRAY, LEFT, THIN_BORDER)
style_cell(ws_sa, r, 4, f"='Form 1040'!D{LINE_ROWS['11']}", GREEN_LINK, LIGHT_BLUE, RIGHT, THIN_BORDER, CURR_FMT)
SA_AGI = r; r += 1
style_cell(ws_sa, r, 2, '3', BLACK_BOLD, LIGHT_GRAY, CENTER, THIN_BORDER)
style_cell(ws_sa, r, 3, '7.5% of AGI threshold', BLACK, LIGHT_GRAY, LEFT, THIN_BORDER)
style_cell(ws_sa, r, 4, f'=D{SA_AGI}*0.075', BLACK, LIGHT_BLUE, RIGHT, THIN_BORDER, CURR_FMT)
SA_THRESH = r; r += 1
style_cell(ws_sa, r, 2, '4', BLACK_BOLD, LIGHT_GRAY, CENTER, THIN_BORDER)
style_cell(ws_sa, r, 3, 'Deductible medical expenses (excess over 7.5% AGI)', BLACK_BOLD, LIGHT_GRAY, LEFT, THIN_BORDER)
style_cell(ws_sa, r, 4, f'=MAX(0,D{SA_MED}-D{SA_THRESH})', BLACK_BOLD, LIGHT_GREEN, RIGHT, THICK_BOTTOM, CURR_FMT)
SA_MED_DED = r; r += 2

section_header(ws_sa, r, 2, 4, 'TAXES YOU PAID'); r += 1
style_cell(ws_sa, r, 2, '5a', BLACK_BOLD, LIGHT_GRAY, CENTER, THIN_BORDER)
style_cell(ws_sa, r, 3, 'State/local income taxes OR sales taxes', BLACK, LIGHT_GRAY, LEFT, THIN_BORDER)
style_cell(ws_sa, r, 4, f"=MAX(Intake!C{SCHA_START+1},Intake!C{SCHA_START+2})", GREEN_LINK, LIGHT_BLUE, RIGHT, THIN_BORDER, CURR_FMT)
SA_SALT1 = r; r += 1
style_cell(ws_sa, r, 2, '5b', BLACK_BOLD, LIGHT_GRAY, CENTER, THIN_BORDER)
style_cell(ws_sa, r, 3, 'Real estate taxes', BLACK, LIGHT_GRAY, LEFT, THIN_BORDER)
style_cell(ws_sa, r, 4, f"=Intake!C{SCHA_START+3}", GREEN_LINK, LIGHT_BLUE, RIGHT, THIN_BORDER, CURR_FMT)
SA_RE_TAX = r; r += 1
style_cell(ws_sa, r, 2, '5c', BLACK_BOLD, LIGHT_GRAY, CENTER, THIN_BORDER)
style_cell(ws_sa, r, 3, 'Personal property taxes', BLACK, LIGHT_GRAY, LEFT, THIN_BORDER)
style_cell(ws_sa, r, 4, f"=Intake!C{SCHA_START+4}", GREEN_LINK, LIGHT_BLUE, RIGHT, THIN_BORDER, CURR_FMT)
SA_PP_TAX = r; r += 1
style_cell(ws_sa, r, 2, '5d', BLACK_BOLD, LIGHT_GRAY, CENTER, THIN_BORDER)
style_cell(ws_sa, r, 3, 'Total SALT (capped at $10,000)', BLACK_BOLD, LIGHT_GRAY, LEFT, THIN_BORDER)
style_cell(ws_sa, r, 4, f'=MIN(10000,D{SA_SALT1}+D{SA_RE_TAX}+D{SA_PP_TAX})', BLACK_BOLD, LIGHT_GREEN, RIGHT, THICK_BOTTOM, CURR_FMT)
SA_SALT_DED = r; r += 2

section_header(ws_sa, r, 2, 4, 'INTEREST YOU PAID'); r += 1
style_cell(ws_sa, r, 2, '8a', BLACK_BOLD, LIGHT_GRAY, CENTER, THIN_BORDER)
style_cell(ws_sa, r, 3, 'Home mortgage interest (Form 1098)', BLACK, LIGHT_GRAY, LEFT, THIN_BORDER)
style_cell(ws_sa, r, 4, f"=Intake!C{SCHA_START+6}", GREEN_LINK, LIGHT_BLUE, RIGHT, THIN_BORDER, CURR_FMT)
SA_MORT = r; r += 1
style_cell(ws_sa, r, 2, '8d', BLACK_BOLD, LIGHT_GRAY, CENTER, THIN_BORDER)
style_cell(ws_sa, r, 3, 'Mortgage insurance premiums', BLACK, LIGHT_GRAY, LEFT, THIN_BORDER)
style_cell(ws_sa, r, 4, f"=Intake!C{SCHA_START+7}", GREEN_LINK, LIGHT_BLUE, RIGHT, THIN_BORDER, CURR_FMT)
SA_MI = r; r += 1
style_cell(ws_sa, r, 2, '9', BLACK_BOLD, LIGHT_GRAY, CENTER, THIN_BORDER)
style_cell(ws_sa, r, 3, 'Investment interest expense', BLACK, LIGHT_GRAY, LEFT, THIN_BORDER)
style_cell(ws_sa, r, 4, f"=Intake!C{SCHA_START+8}", GREEN_LINK, LIGHT_BLUE, RIGHT, THIN_BORDER, CURR_FMT)
SA_INV_INT = r; r += 1
style_cell(ws_sa, r, 2, '10', BLACK_BOLD, LIGHT_GRAY, CENTER, THIN_BORDER)
style_cell(ws_sa, r, 3, 'Total interest deduction', BLACK_BOLD, LIGHT_GRAY, LEFT, THIN_BORDER)
style_cell(ws_sa, r, 4, f'=D{SA_MORT}+D{SA_MI}+D{SA_INV_INT}', BLACK_BOLD, LIGHT_GREEN, RIGHT, THICK_BOTTOM, CURR_FMT)
SA_INT_DED = r; r += 2

section_header(ws_sa, r, 2, 4, 'GIFTS TO CHARITY'); r += 1
style_cell(ws_sa, r, 2, '11', BLACK_BOLD, LIGHT_GRAY, CENTER, THIN_BORDER)
style_cell(ws_sa, r, 3, 'Cash contributions', BLACK, LIGHT_GRAY, LEFT, THIN_BORDER)
style_cell(ws_sa, r, 4, f"=Intake!C{SCHA_START+9}", GREEN_LINK, LIGHT_BLUE, RIGHT, THIN_BORDER, CURR_FMT)
SA_CASH_CHAR = r; r += 1
style_cell(ws_sa, r, 2, '12', BLACK_BOLD, LIGHT_GRAY, CENTER, THIN_BORDER)
style_cell(ws_sa, r, 3, 'Noncash contributions', BLACK, LIGHT_GRAY, LEFT, THIN_BORDER)
style_cell(ws_sa, r, 4, f"=Intake!C{SCHA_START+10}", GREEN_LINK, LIGHT_BLUE, RIGHT, THIN_BORDER, CURR_FMT)
SA_NC_CHAR = r; r += 1
style_cell(ws_sa, r, 2, '13', BLACK_BOLD, LIGHT_GRAY, CENTER, THIN_BORDER)
style_cell(ws_sa, r, 3, 'Carryover from prior year', BLACK, LIGHT_GRAY, LEFT, THIN_BORDER)
style_cell(ws_sa, r, 4, f"=Intake!C{SCHA_START+11}", GREEN_LINK, LIGHT_BLUE, RIGHT, THIN_BORDER, CURR_FMT)
SA_CARRY = r; r += 1
style_cell(ws_sa, r, 2, '14', BLACK_BOLD, LIGHT_GRAY, CENTER, THIN_BORDER)
style_cell(ws_sa, r, 3, 'Total charitable deduction', BLACK_BOLD, LIGHT_GRAY, LEFT, THIN_BORDER)
style_cell(ws_sa, r, 4, f'=D{SA_CASH_CHAR}+D{SA_NC_CHAR}+D{SA_CARRY}', BLACK_BOLD, LIGHT_GREEN, RIGHT, THICK_BOTTOM, CURR_FMT)
SA_CHAR_DED = r; r += 2

section_header(ws_sa, r, 2, 4, 'OTHER DEDUCTIONS'); r += 1
style_cell(ws_sa, r, 2, '15', BLACK_BOLD, LIGHT_GRAY, CENTER, THIN_BORDER)
style_cell(ws_sa, r, 3, 'Casualty/theft losses (federally declared disaster)', BLACK, LIGHT_GRAY, LEFT, THIN_BORDER)
style_cell(ws_sa, r, 4, f"=Intake!C{SCHA_START+12}", GREEN_LINK, LIGHT_BLUE, RIGHT, THIN_BORDER, CURR_FMT)
SA_CAS = r; r += 1
style_cell(ws_sa, r, 2, '16', BLACK_BOLD, LIGHT_GRAY, CENTER, THIN_BORDER)
style_cell(ws_sa, r, 3, 'Gambling losses (limited to winnings)', BLACK, LIGHT_GRAY, LEFT, THIN_BORDER)
style_cell(ws_sa, r, 4, f"=Intake!C{SCHA_START+13}", GREEN_LINK, LIGHT_BLUE, RIGHT, THIN_BORDER, CURR_FMT)
SA_GAMB = r; r += 1
style_cell(ws_sa, r, 2, '17', BLACK_BOLD, LIGHT_GRAY, CENTER, THIN_BORDER)
style_cell(ws_sa, r, 3, 'Other itemized deductions', BLACK, LIGHT_GRAY, LEFT, THIN_BORDER)
style_cell(ws_sa, r, 4, f"=Intake!C{SCHA_START+14}", GREEN_LINK, LIGHT_BLUE, RIGHT, THIN_BORDER, CURR_FMT)
SA_OTHER = r; r += 2

# Row 30 = total itemized deductions (referenced by Deduction Calc)
SA_TOTAL_ROW = 30
style_cell(ws_sa, SA_TOTAL_ROW, 2, '18', BLACK_BOLD, LIGHT_GRAY, CENTER, THICK_BOTTOM)
style_cell(ws_sa, SA_TOTAL_ROW, 3, 'TOTAL ITEMIZED DEDUCTIONS', BLACK_BOLD, LIGHT_GRAY, LEFT, THICK_BOTTOM)
style_cell(ws_sa, SA_TOTAL_ROW, 4,
    f'=D{SA_MED_DED}+D{SA_SALT_DED}+D{SA_INT_DED}+D{SA_CHAR_DED}+D{SA_CAS}+D{SA_GAMB}+D{SA_OTHER}',
    BLACK_BOLD, LIGHT_GREEN, RIGHT, THICK_BOTTOM, CURR_FMT)

print("Schedule A complete.")


# ═══════════════════════════════════════════════════════════════════════════════
# TAB: SCHEDULE B — Interest and Ordinary Dividends
# ═══════════════════════════════════════════════════════════════════════════════
ws_sb = wb.create_sheet('Schedule B')
ws_sb.sheet_properties.tabColor = '5B9BD5'
ws_sb.column_dimensions['A'].width = 3
ws_sb.column_dimensions['B'].width = 5
ws_sb.column_dimensions['C'].width = 40
ws_sb.column_dimensions['D'].width = 20

r = 1
style_cell(ws_sb, r, 2, 'Schedule B — Interest and Ordinary Dividends', TITLE_FONT); r += 2

section_header(ws_sb, r, 2, 4, 'PART I — INTEREST (pulls from Intake Section 6)'); r += 1
style_cell(ws_sb, r, 3, 'See Intake tab for detailed 1099-INT entries', SMALL_FONT); r += 1
style_cell(ws_sb, r, 2, '1', BLACK_BOLD, LIGHT_GRAY, CENTER, THIN_BORDER)
style_cell(ws_sb, r, 3, 'Total interest income', BLACK, LIGHT_GRAY, LEFT, THIN_BORDER)
style_cell(ws_sb, r, 4, f"=Intake!D{TOTAL_INT_ROW}", GREEN_LINK, LIGHT_BLUE, RIGHT, THIN_BORDER, CURR_FMT)
r += 1
style_cell(ws_sb, r, 2, '4', BLACK_BOLD, LIGHT_GRAY, CENTER, THIN_BORDER)
style_cell(ws_sb, r, 3, 'Taxable interest → Form 1040 Line 2b', BLACK_BOLD, LIGHT_GRAY, LEFT, THIN_BORDER)
style_cell(ws_sb, r, 4, f"=Intake!D{TOTAL_INT_ROW}", GREEN_LINK, LIGHT_GREEN, RIGHT, THICK_BOTTOM, CURR_FMT)
r += 2

section_header(ws_sb, r, 2, 4, 'PART II — ORDINARY DIVIDENDS (pulls from Intake Section 7)'); r += 1
style_cell(ws_sb, r, 2, '5', BLACK_BOLD, LIGHT_GRAY, CENTER, THIN_BORDER)
style_cell(ws_sb, r, 3, 'Total ordinary dividends', BLACK, LIGHT_GRAY, LEFT, THIN_BORDER)
style_cell(ws_sb, r, 4, f"=Intake!C{TOTAL_ORD_DIV_ROW}", GREEN_LINK, LIGHT_BLUE, RIGHT, THIN_BORDER, CURR_FMT)
r += 1
style_cell(ws_sb, r, 2, '6', BLACK_BOLD, LIGHT_GRAY, CENTER, THIN_BORDER)
style_cell(ws_sb, r, 3, 'Ordinary dividends → Form 1040 Line 3b', BLACK_BOLD, LIGHT_GRAY, LEFT, THIN_BORDER)
style_cell(ws_sb, r, 4, f"=Intake!C{TOTAL_ORD_DIV_ROW}", GREEN_LINK, LIGHT_GREEN, RIGHT, THICK_BOTTOM, CURR_FMT)
r += 2

section_header(ws_sb, r, 2, 4, 'PART III — FOREIGN ACCOUNTS AND TRUSTS'); r += 1
style_cell(ws_sb, r, 2, '7a', BLACK_BOLD, LIGHT_GRAY, CENTER, THIN_BORDER)
style_cell(ws_sb, r, 3, 'Do you have foreign accounts or foreign trusts? (Y/N)', BLACK, LIGHT_GRAY, LEFT, THIN_BORDER)
style_cell(ws_sb, r, 4, f"=Intake!C{FA_START-5}", GREEN_LINK, LIGHT_BLUE, CENTER, THIN_BORDER)
r += 1
style_cell(ws_sb, r, 2, '7b', BLACK_BOLD, LIGHT_GRAY, CENTER, THIN_BORDER)
style_cell(ws_sb, r, 3, 'Country where accounts are located', BLACK, LIGHT_GRAY, LEFT, THIN_BORDER)
style_cell(ws_sb, r, 4, None, BLUE_INPUT, LIGHT_YELLOW, LEFT, THIN_BORDER)
r += 1
style_cell(ws_sb, r, 2, '8', BLACK_BOLD, LIGHT_GRAY, CENTER, THIN_BORDER)
style_cell(ws_sb, r, 3, 'Are you required to file FinCEN Form 114 (FBAR)? (Y/N)', BLACK, LIGHT_GRAY, LEFT, THIN_BORDER)
style_cell(ws_sb, r, 4, f'=IF(Intake!F{FA_START-3}>10000,"YES","NO")', BLACK, LIGHT_BLUE, CENTER, THIN_BORDER)

print("Schedule B complete.")


# ═══════════════════════════════════════════════════════════════════════════════
# TAB: SCHEDULE C — Profit or Loss from Business
# ═══════════════════════════════════════════════════════════════════════════════
ws_sc = wb.create_sheet('Schedule C')
ws_sc.sheet_properties.tabColor = 'BF8F00'
ws_sc.column_dimensions['A'].width = 3
ws_sc.column_dimensions['B'].width = 5
ws_sc.column_dimensions['C'].width = 55
ws_sc.column_dimensions['D'].width = 20

r = 1
style_cell(ws_sc, r, 2, 'Schedule C — Profit or Loss from Business', TITLE_FONT); r += 2

section_header(ws_sc, r, 2, 4, 'BUSINESS INFORMATION'); r += 1
sc_info = [
    ('A', 'Principal business or profession', f"=Intake!C{SCHC_START+2}"),
    ('B', 'Business name', f"=Intake!C{SCHC_START}"),
    ('C', 'Business EIN', f"=Intake!C{SCHC_START+1}"),
    ('D', 'Business address', f"=Intake!C{SCHC_START+3}"),
    ('F', 'Accounting method', f"=Intake!C{SCHC_START+4}"),
]
for ln, desc, formula in sc_info:
    style_cell(ws_sc, r, 2, ln, BLACK_BOLD, LIGHT_GRAY, CENTER, THIN_BORDER)
    style_cell(ws_sc, r, 3, desc, BLACK, LIGHT_GRAY, LEFT, THIN_BORDER)
    style_cell(ws_sc, r, 4, formula, GREEN_LINK, LIGHT_BLUE, LEFT, THIN_BORDER)
    r += 1
r += 1

section_header(ws_sc, r, 2, 4, 'INCOME'); r += 1
style_cell(ws_sc, r, 2, '1', BLACK_BOLD, LIGHT_GRAY, CENTER, THIN_BORDER)
style_cell(ws_sc, r, 3, 'Gross receipts or sales', BLACK, LIGHT_GRAY, LEFT, THIN_BORDER)
style_cell(ws_sc, r, 4, f"=Intake!C{SCHC_START+5}", GREEN_LINK, LIGHT_BLUE, RIGHT, THIN_BORDER, CURR_FMT)
SC_GROSS = r; r += 1
style_cell(ws_sc, r, 2, '2', BLACK_BOLD, LIGHT_GRAY, CENTER, THIN_BORDER)
style_cell(ws_sc, r, 3, 'Returns and allowances', BLACK, LIGHT_GRAY, LEFT, THIN_BORDER)
style_cell(ws_sc, r, 4, f"=Intake!C{SCHC_START+6}", GREEN_LINK, LIGHT_BLUE, RIGHT, THIN_BORDER, CURR_FMT)
SC_RET = r; r += 1
style_cell(ws_sc, r, 2, '4', BLACK_BOLD, LIGHT_GRAY, CENTER, THIN_BORDER)
style_cell(ws_sc, r, 3, 'Cost of goods sold', BLACK, LIGHT_GRAY, LEFT, THIN_BORDER)
style_cell(ws_sc, r, 4, f"=Intake!C{SCHC_START+7}", GREEN_LINK, LIGHT_BLUE, RIGHT, THIN_BORDER, CURR_FMT)
SC_COGS = r; r += 1
style_cell(ws_sc, r, 2, '5', BLACK_BOLD, LIGHT_GRAY, CENTER, THIN_BORDER)
style_cell(ws_sc, r, 3, 'Gross profit (Line 1 - 2 - 4)', BLACK_BOLD, LIGHT_GRAY, LEFT, THIN_BORDER)
style_cell(ws_sc, r, 4, f'=D{SC_GROSS}-D{SC_RET}-D{SC_COGS}', BLACK_BOLD, LIGHT_GREEN, RIGHT, THICK_BOTTOM, CURR_FMT)
SC_GP = r; r += 2

section_header(ws_sc, r, 2, 4, 'EXPENSES'); r += 1
expense_map = [
    ('8', 'Advertising', SCHC_START + 8),
    ('9', 'Car and truck expenses', SCHC_START + 9),
    ('10', 'Commissions and fees', SCHC_START + 10),
    ('11', 'Contract labor', SCHC_START + 11),
    ('13', 'Depreciation (Form 4562)', SCHC_START + 12),
    ('14', 'Employee benefit programs', SCHC_START + 13),
    ('15', 'Insurance (other than health)', SCHC_START + 14),
    ('16a', 'Mortgage interest', SCHC_START + 15),
    ('16b', 'Other interest', SCHC_START + 16),
    ('17', 'Legal and professional services', SCHC_START + 17),
    ('18', 'Office expense', SCHC_START + 18),
    ('20a', 'Rent — vehicles/machinery', SCHC_START + 19),
    ('20b', 'Rent — other', SCHC_START + 20),
    ('21', 'Repairs and maintenance', SCHC_START + 21),
    ('22', 'Supplies', SCHC_START + 22),
    ('23', 'Taxes and licenses', SCHC_START + 23),
    ('24a', 'Travel', SCHC_START + 24),
    ('24b', 'Meals (50% deductible)', SCHC_START + 25),
    ('25', 'Utilities', SCHC_START + 26),
    ('26', 'Wages', SCHC_START + 27),
    ('27a', 'Other expenses', SCHC_START + 29),
]
SC_EXP_ROWS = {}
for ln, desc, intake_row in expense_map:
    style_cell(ws_sc, r, 2, ln, BLACK_BOLD, LIGHT_GRAY, CENTER, THIN_BORDER)
    style_cell(ws_sc, r, 3, desc, BLACK, LIGHT_GRAY, LEFT, THIN_BORDER)
    style_cell(ws_sc, r, 4, f"=Intake!C{intake_row}", GREEN_LINK, LIGHT_BLUE, RIGHT, THIN_BORDER, CURR_FMT)
    SC_EXP_ROWS[ln] = r
    r += 1

style_cell(ws_sc, r, 2, '28', BLACK_BOLD, LIGHT_GRAY, CENTER, THIN_BORDER)
style_cell(ws_sc, r, 3, 'Total expenses', BLACK_BOLD, LIGHT_GRAY, LEFT, THIN_BORDER)
first_exp = list(SC_EXP_ROWS.values())[0]
last_exp = list(SC_EXP_ROWS.values())[-1]
style_cell(ws_sc, r, 4, f'=SUM(D{first_exp}:D{last_exp})', BLACK_BOLD, LIGHT_GREEN, RIGHT, THICK_BOTTOM, CURR_FMT)
SC_TOTAL_EXP = r; r += 1

style_cell(ws_sc, r, 2, '29', BLACK_BOLD, LIGHT_GRAY, CENTER, THIN_BORDER)
style_cell(ws_sc, r, 3, 'Tentative profit (loss)', BLACK, LIGHT_GRAY, LEFT, THIN_BORDER)
style_cell(ws_sc, r, 4, f'=D{SC_GP}-D{SC_TOTAL_EXP}', BLACK, LIGHT_BLUE, RIGHT, THIN_BORDER, CURR_FMT)
SC_TENT = r; r += 1

style_cell(ws_sc, r, 2, '30', BLACK_BOLD, LIGHT_GRAY, CENTER, THIN_BORDER)
style_cell(ws_sc, r, 3, 'Business use of home (Form 8829)', BLACK, LIGHT_GRAY, LEFT, THIN_BORDER)
style_cell(ws_sc, r, 4, f"=Intake!C{SCHC_START+30}", GREEN_LINK, LIGHT_BLUE, RIGHT, THIN_BORDER, CURR_FMT)
SC_HOME = r; r += 1

# Row 45 = Net profit/loss (referenced by Schedule 1)
while r < 45:
    r += 1
style_cell(ws_sc, 45, 2, '31', BLACK_BOLD, LIGHT_GRAY, CENTER, THICK_BOTTOM)
style_cell(ws_sc, 45, 3, 'NET PROFIT (LOSS) → Schedule 1 Line 3', BLACK_BOLD, LIGHT_GRAY, LEFT, THICK_BOTTOM)
style_cell(ws_sc, 45, 4, f'=D{SC_TENT}-D{SC_HOME}', BLACK_BOLD, LIGHT_GREEN, RIGHT, THICK_BOTTOM, CURR_FMT)

print("Schedule C complete.")


# ═══════════════════════════════════════════════════════════════════════════════
# TAB: SCHEDULE D — Capital Gains and Losses
# ═══════════════════════════════════════════════════════════════════════════════
ws_sd = wb.create_sheet('Schedule D')
ws_sd.sheet_properties.tabColor = 'C00000'
ws_sd.column_dimensions['A'].width = 3
ws_sd.column_dimensions['B'].width = 5
ws_sd.column_dimensions['C'].width = 55
ws_sd.column_dimensions['D'].width = 20

r = 1
style_cell(ws_sd, r, 2, 'Schedule D — Capital Gains and Losses', TITLE_FONT); r += 2

section_header(ws_sd, r, 2, 4, 'PART I — SHORT-TERM CAPITAL GAINS AND LOSSES'); r += 1
style_cell(ws_sd, r, 2, '7', BLACK_BOLD, LIGHT_GRAY, CENTER, THIN_BORDER)
style_cell(ws_sd, r, 3, 'Net short-term gain (loss) from Form 8949', BLACK, LIGHT_GRAY, LEFT, THIN_BORDER)
style_cell(ws_sd, r, 4, "='Form 8949'!D30", GREEN_LINK, LIGHT_BLUE, RIGHT, THIN_BORDER, CURR_FMT)
SD_ST = r; r += 1
style_cell(ws_sd, r, 2, '8', BLACK_BOLD, LIGHT_GRAY, CENTER, THIN_BORDER)
style_cell(ws_sd, r, 3, 'Short-term gain from K-1s and other forms', BLACK, LIGHT_GRAY, LEFT, THIN_BORDER)
style_cell(ws_sd, r, 4, None, BLUE_INPUT, LIGHT_YELLOW, RIGHT, THIN_BORDER, CURR_FMT)
SD_ST_K1 = r; r += 1
style_cell(ws_sd, r, 2, '9', BLACK_BOLD, LIGHT_GRAY, CENTER, THIN_BORDER)
style_cell(ws_sd, r, 3, 'Short-term capital loss carryover', BLACK, LIGHT_GRAY, LEFT, THIN_BORDER)
style_cell(ws_sd, r, 4, None, BLUE_INPUT, LIGHT_YELLOW, RIGHT, THIN_BORDER, CURR_FMT)
SD_ST_CARRY = r; r += 1
style_cell(ws_sd, r, 2, '10', BLACK_BOLD, LIGHT_GRAY, CENTER, THIN_BORDER)
style_cell(ws_sd, r, 3, 'Net short-term capital gain (loss)', BLACK_BOLD, LIGHT_GRAY, LEFT, THIN_BORDER)
style_cell(ws_sd, r, 4, f'=D{SD_ST}+D{SD_ST_K1}+D{SD_ST_CARRY}', BLACK_BOLD, LIGHT_GREEN, RIGHT, THICK_BOTTOM, CURR_FMT)
SD_NET_ST = r; r += 2

section_header(ws_sd, r, 2, 4, 'PART II — LONG-TERM CAPITAL GAINS AND LOSSES'); r += 1
style_cell(ws_sd, r, 2, '15', BLACK_BOLD, LIGHT_GRAY, CENTER, THIN_BORDER)
style_cell(ws_sd, r, 3, 'Net long-term gain (loss) from Form 8949', BLACK, LIGHT_GRAY, LEFT, THIN_BORDER)
style_cell(ws_sd, r, 4, "='Form 8949'!D60", GREEN_LINK, LIGHT_BLUE, RIGHT, THIN_BORDER, CURR_FMT)
SD_LT = r; r += 1
style_cell(ws_sd, r, 2, '16', BLACK_BOLD, LIGHT_GRAY, CENTER, THIN_BORDER)
style_cell(ws_sd, r, 3, 'Long-term gain from K-1s and Form 1099-DIV cap gain dist', BLACK, LIGHT_GRAY, LEFT, THIN_BORDER)
style_cell(ws_sd, r, 4, None, BLUE_INPUT, LIGHT_YELLOW, RIGHT, THIN_BORDER, CURR_FMT)
SD_LT_K1 = r; r += 1
style_cell(ws_sd, r, 2, '17', BLACK_BOLD, LIGHT_GRAY, CENTER, THIN_BORDER)
style_cell(ws_sd, r, 3, 'Long-term capital loss carryover', BLACK, LIGHT_GRAY, LEFT, THIN_BORDER)
style_cell(ws_sd, r, 4, None, BLUE_INPUT, LIGHT_YELLOW, RIGHT, THIN_BORDER, CURR_FMT)
SD_LT_CARRY = r; r += 1
style_cell(ws_sd, r, 2, '18', BLACK_BOLD, LIGHT_GRAY, CENTER, THIN_BORDER)
style_cell(ws_sd, r, 3, 'Net long-term capital gain (loss)', BLACK_BOLD, LIGHT_GRAY, LEFT, THIN_BORDER)
style_cell(ws_sd, r, 4, f'=D{SD_LT}+D{SD_LT_K1}+D{SD_LT_CARRY}', BLACK_BOLD, LIGHT_GREEN, RIGHT, THICK_BOTTOM, CURR_FMT)
SD_NET_LT = r; r += 2

section_header(ws_sd, r, 2, 4, 'PART III — SUMMARY'); r += 1

# Row 48 = Net LTCG for Tax Computation
style_cell(ws_sd, r, 2, '19', BLACK_BOLD, LIGHT_GRAY, CENTER, THIN_BORDER)
style_cell(ws_sd, r, 3, 'Combine lines 10 and 18', BLACK, LIGHT_GRAY, LEFT, THIN_BORDER)
style_cell(ws_sd, r, 4, f'=D{SD_NET_ST}+D{SD_NET_LT}', BLACK, LIGHT_BLUE, RIGHT, THIN_BORDER, CURR_FMT)
SD_COMBINED = r; r += 1

style_cell(ws_sd, r, 2, '20', BLACK_BOLD, LIGHT_GRAY, CENTER, THIN_BORDER)
style_cell(ws_sd, r, 3, 'Capital loss limitation ($3,000 max or $1,500 MFS)', BLACK, LIGHT_GRAY, LEFT, THIN_BORDER)
style_cell(ws_sd, r, 4, f'=IF(D{SD_COMBINED}<0,MAX(D{SD_COMBINED},-3000),D{SD_COMBINED})', BLACK, LIGHT_BLUE, RIGHT, THIN_BORDER, CURR_FMT)
SD_LIMITED = r

# Set specific output rows
while r < 48:
    r += 1
style_cell(ws_sd, 48, 2, '', BLACK_BOLD, LIGHT_GRAY, CENTER, THIN_BORDER)
style_cell(ws_sd, 48, 3, 'Net LTCG (for Tax Computation)', BLACK, LIGHT_GRAY, LEFT, THIN_BORDER)
style_cell(ws_sd, 48, 4, f'=MAX(0,D{SD_NET_LT})', BLACK, LIGHT_BLUE, RIGHT, THIN_BORDER, CURR_FMT)

style_cell(ws_sd, 50, 2, '21', BLACK_BOLD, LIGHT_GRAY, CENTER, THICK_BOTTOM)
style_cell(ws_sd, 50, 3, 'NET CAPITAL GAIN (LOSS) → Form 1040 Line 7', BLACK_BOLD, LIGHT_GRAY, LEFT, THICK_BOTTOM)
style_cell(ws_sd, 50, 4, f'=D{SD_LIMITED}', BLACK_BOLD, LIGHT_GREEN, RIGHT, THICK_BOTTOM, CURR_FMT)

print("Schedule D complete.")


# ═══════════════════════════════════════════════════════════════════════════════
# TAB: FORM 8949 — Sales and Dispositions of Capital Assets
# ═══════════════════════════════════════════════════════════════════════════════
ws_8949 = wb.create_sheet('Form 8949')
ws_8949.sheet_properties.tabColor = 'FF6600'
ws_8949.column_dimensions['A'].width = 3
ws_8949.column_dimensions['B'].width = 30
ws_8949.column_dimensions['C'].width = 15
ws_8949.column_dimensions['D'].width = 18
ws_8949.column_dimensions['E'].width = 18
ws_8949.column_dimensions['F'].width = 18
ws_8949.column_dimensions['G'].width = 15

r = 1
style_cell(ws_8949, r, 2, 'Form 8949 — Sales and Other Dispositions of Capital Assets', TITLE_FONT); r += 2

# Part I — Short-Term
section_header(ws_8949, r, 2, 7, 'PART I — SHORT-TERM (held one year or less)'); r += 1
for i, h in enumerate(['Description', 'Date Acquired', 'Date Sold', 'Proceeds', 'Cost Basis', 'Gain/(Loss)'], 2):
    style_cell(ws_8949, r, i, h, HEADER_FONT, MEDIUM_BLUE, CENTER, THIN_BORDER)
r += 1
F8949_ST_START = r
for row_i in range(15):
    style_cell(ws_8949, r, 2, None, BLUE_INPUT, LIGHT_YELLOW, LEFT, THIN_BORDER)
    style_cell(ws_8949, r, 3, None, BLUE_INPUT, LIGHT_YELLOW, CENTER, THIN_BORDER, DATE_FMT)
    style_cell(ws_8949, r, 4, None, BLUE_INPUT, LIGHT_YELLOW, CENTER, THIN_BORDER, DATE_FMT)
    style_cell(ws_8949, r, 5, None, BLUE_INPUT, LIGHT_YELLOW, RIGHT, THIN_BORDER, CURR_FMT)
    style_cell(ws_8949, r, 6, None, BLUE_INPUT, LIGHT_YELLOW, RIGHT, THIN_BORDER, CURR_FMT)
    style_cell(ws_8949, r, 7, f'=IF(OR(E{r}="",F{r}=""),"",E{r}-F{r})', BLACK, LIGHT_BLUE, RIGHT, THIN_BORDER, CURR_FMT)
    r += 1
F8949_ST_END = r - 1

# Row 30 = Short-term total
while r < 29:
    r += 1
r = 29
style_cell(ws_8949, r, 2, 'Total Short-Term', BLACK_BOLD, LIGHT_GREEN, LEFT, THICK_BOTTOM)
style_cell(ws_8949, r, 5, f'=SUM(E{F8949_ST_START}:E{F8949_ST_END})', BLACK_BOLD, LIGHT_GREEN, RIGHT, THICK_BOTTOM, CURR_FMT)
style_cell(ws_8949, r, 6, f'=SUM(F{F8949_ST_START}:F{F8949_ST_END})', BLACK_BOLD, LIGHT_GREEN, RIGHT, THICK_BOTTOM, CURR_FMT)
style_cell(ws_8949, 30, 2, 'Net Short-Term Gain/(Loss) → Schedule D', BLACK_BOLD, LIGHT_GREEN, LEFT, THICK_BOTTOM)
style_cell(ws_8949, 30, 4, f'=SUM(G{F8949_ST_START}:G{F8949_ST_END})', BLACK_BOLD, LIGHT_GREEN, RIGHT, THICK_BOTTOM, CURR_FMT)
r = 32

# Part II — Long-Term
section_header(ws_8949, r, 2, 7, 'PART II — LONG-TERM (held more than one year)'); r += 1
for i, h in enumerate(['Description', 'Date Acquired', 'Date Sold', 'Proceeds', 'Cost Basis', 'Gain/(Loss)'], 2):
    style_cell(ws_8949, r, i, h, HEADER_FONT, MEDIUM_BLUE, CENTER, THIN_BORDER)
r += 1
F8949_LT_START = r
for row_i in range(15):
    style_cell(ws_8949, r, 2, None, BLUE_INPUT, LIGHT_YELLOW, LEFT, THIN_BORDER)
    style_cell(ws_8949, r, 3, None, BLUE_INPUT, LIGHT_YELLOW, CENTER, THIN_BORDER, DATE_FMT)
    style_cell(ws_8949, r, 4, None, BLUE_INPUT, LIGHT_YELLOW, CENTER, THIN_BORDER, DATE_FMT)
    style_cell(ws_8949, r, 5, None, BLUE_INPUT, LIGHT_YELLOW, RIGHT, THIN_BORDER, CURR_FMT)
    style_cell(ws_8949, r, 6, None, BLUE_INPUT, LIGHT_YELLOW, RIGHT, THIN_BORDER, CURR_FMT)
    style_cell(ws_8949, r, 7, f'=IF(OR(E{r}="",F{r}=""),"",E{r}-F{r})', BLACK, LIGHT_BLUE, RIGHT, THIN_BORDER, CURR_FMT)
    r += 1
F8949_LT_END = r - 1

# Row 60 = Long-term total
while r < 59:
    r += 1
r = 59
style_cell(ws_8949, r, 2, 'Total Long-Term', BLACK_BOLD, LIGHT_GREEN, LEFT, THICK_BOTTOM)
style_cell(ws_8949, r, 5, f'=SUM(E{F8949_LT_START}:E{F8949_LT_END})', BLACK_BOLD, LIGHT_GREEN, RIGHT, THICK_BOTTOM, CURR_FMT)
style_cell(ws_8949, r, 6, f'=SUM(F{F8949_LT_START}:F{F8949_LT_END})', BLACK_BOLD, LIGHT_GREEN, RIGHT, THICK_BOTTOM, CURR_FMT)
style_cell(ws_8949, 60, 2, 'Net Long-Term Gain/(Loss) → Schedule D', BLACK_BOLD, LIGHT_GREEN, LEFT, THICK_BOTTOM)
style_cell(ws_8949, 60, 4, f'=SUM(G{F8949_LT_START}:G{F8949_LT_END})', BLACK_BOLD, LIGHT_GREEN, RIGHT, THICK_BOTTOM, CURR_FMT)

print("Form 8949 complete.")


# ═══════════════════════════════════════════════════════════════════════════════
# TAB: SCHEDULE E — Supplemental Income and Loss
# ═══════════════════════════════════════════════════════════════════════════════
ws_se_rental = wb.create_sheet('Schedule E')
ws_se_rental.sheet_properties.tabColor = '00B0F0'
ws_se_rental.column_dimensions['A'].width = 3
ws_se_rental.column_dimensions['B'].width = 5
ws_se_rental.column_dimensions['C'].width = 35
ws_se_rental.column_dimensions['D'].width = 18
ws_se_rental.column_dimensions['E'].width = 18
ws_se_rental.column_dimensions['F'].width = 18

r = 1
style_cell(ws_se_rental, r, 2, 'Schedule E — Supplemental Income and Loss', TITLE_FONT); r += 2

section_header(ws_se_rental, r, 2, 6, 'PART I — RENTAL REAL ESTATE AND ROYALTIES'); r += 1
style_cell(ws_se_rental, r, 3, 'Item', HEADER_FONT, MEDIUM_BLUE, CENTER, THIN_BORDER)
style_cell(ws_se_rental, r, 4, 'Property 1', HEADER_FONT, MEDIUM_BLUE, CENTER, THIN_BORDER)
style_cell(ws_se_rental, r, 5, 'Property 2', HEADER_FONT, MEDIUM_BLUE, CENTER, THIN_BORDER)
style_cell(ws_se_rental, r, 6, 'Property 3', HEADER_FONT, MEDIUM_BLUE, CENTER, THIN_BORDER)
r += 1

rental_items = [
    ('3', 'Rents received', SCHE_START + 4),
    ('5', 'Advertising', SCHE_START + 5),
    ('6', 'Auto and travel', SCHE_START + 6),
    ('7', 'Cleaning and maintenance', SCHE_START + 7),
    ('8', 'Commissions', SCHE_START + 8),
    ('9', 'Insurance', SCHE_START + 9),
    ('10', 'Legal and professional', SCHE_START + 10),
    ('11', 'Management fees', SCHE_START + 11),
    ('12', 'Mortgage interest', SCHE_START + 12),
    ('13', 'Other interest', SCHE_START + 13),
    ('14', 'Repairs', SCHE_START + 14),
    ('15', 'Supplies', SCHE_START + 15),
    ('16', 'Taxes', SCHE_START + 16),
    ('17', 'Utilities', SCHE_START + 17),
    ('18', 'Depreciation', SCHE_START + 18),
    ('19', 'Other', SCHE_START + 19),
]

SE_ROWS = {}
SE_RENT_ROW = None
SE_EXP_START = None
SE_EXP_END = None
for ln, desc, intake_row in rental_items:
    style_cell(ws_se_rental, r, 2, ln, BLACK_BOLD, LIGHT_GRAY, CENTER, THIN_BORDER)
    style_cell(ws_se_rental, r, 3, desc, BLACK, LIGHT_GRAY, LEFT, THIN_BORDER)
    for c in range(4, 7):
        col_letter = get_column_letter(c - 1)  # Intake columns C, D, E
        style_cell(ws_se_rental, r, c, f"=Intake!{col_letter}{intake_row}", GREEN_LINK, LIGHT_BLUE, RIGHT, THIN_BORDER, CURR_FMT)
    SE_ROWS[ln] = r
    if ln == '3':
        SE_RENT_ROW = r
    if ln == '5':
        SE_EXP_START = r
    if ln == '19':
        SE_EXP_END = r
    r += 1

r += 1
style_cell(ws_se_rental, r, 2, '20', BLACK_BOLD, LIGHT_GRAY, CENTER, THIN_BORDER)
style_cell(ws_se_rental, r, 3, 'Total expenses', BLACK_BOLD, LIGHT_GRAY, LEFT, THIN_BORDER)
for c in range(4, 7):
    cl = get_column_letter(c)
    style_cell(ws_se_rental, r, c, f'=SUM({cl}{SE_EXP_START}:{cl}{SE_EXP_END})', BLACK_BOLD, LIGHT_GREEN, RIGHT, THICK_BOTTOM, CURR_FMT)
SE_TOTAL_EXP = r; r += 1

style_cell(ws_se_rental, r, 2, '21', BLACK_BOLD, LIGHT_GRAY, CENTER, THIN_BORDER)
style_cell(ws_se_rental, r, 3, 'Net rental income (loss) per property', BLACK, LIGHT_GRAY, LEFT, THIN_BORDER)
for c in range(4, 7):
    cl = get_column_letter(c)
    style_cell(ws_se_rental, r, c, f'={cl}{SE_RENT_ROW}-{cl}{SE_TOTAL_EXP}', BLACK, LIGHT_BLUE, RIGHT, THIN_BORDER, CURR_FMT)
SE_NET_PER = r; r += 2

section_header(ws_se_rental, r, 2, 6, 'PART II — INCOME FROM PARTNERSHIPS AND S CORPORATIONS'); r += 1
style_cell(ws_se_rental, r, 3, 'See Intake Section 19 for K-1 details', SMALL_FONT); r += 1
style_cell(ws_se_rental, r, 2, '32', BLACK_BOLD, LIGHT_GRAY, CENTER, THIN_BORDER)
style_cell(ws_se_rental, r, 3, 'Total K-1 ordinary income', BLACK, LIGHT_GRAY, LEFT, THIN_BORDER)
style_cell(ws_se_rental, r, 4, f"=SUM(Intake!E{K1_START}:E{K1_END})", GREEN_LINK, LIGHT_BLUE, RIGHT, THIN_BORDER, CURR_FMT)
SE_K1_ORD = r; r += 1
style_cell(ws_se_rental, r, 2, '33', BLACK_BOLD, LIGHT_GRAY, CENTER, THIN_BORDER)
style_cell(ws_se_rental, r, 3, 'Total K-1 rental income', BLACK, LIGHT_GRAY, LEFT, THIN_BORDER)
style_cell(ws_se_rental, r, 4, f"=SUM(Intake!F{K1_START}:F{K1_END})", GREEN_LINK, LIGHT_BLUE, RIGHT, THIN_BORDER, CURR_FMT)
SE_K1_RENT = r; r += 2

# Row 50 = Total Schedule E
while r < 50:
    r += 1
style_cell(ws_se_rental, 50, 2, '41', BLACK_BOLD, LIGHT_GRAY, CENTER, THICK_BOTTOM)
style_cell(ws_se_rental, 50, 3, 'TOTAL SCH E INCOME (LOSS) → Schedule 1', BLACK_BOLD, LIGHT_GRAY, LEFT, THICK_BOTTOM)
style_cell(ws_se_rental, 50, 4,
    f'=D{SE_NET_PER}+E{SE_NET_PER}+F{SE_NET_PER}+D{SE_K1_ORD}+D{SE_K1_RENT}',
    BLACK_BOLD, LIGHT_GREEN, RIGHT, THICK_BOTTOM, CURR_FMT)

print("Schedule E complete.")


# ═══════════════════════════════════════════════════════════════════════════════
# TAB: SCHEDULE SE — Self-Employment Tax
# ═══════════════════════════════════════════════════════════════════════════════
ws_se = wb.create_sheet('Schedule SE')
ws_se.sheet_properties.tabColor = '9DC3E6'
ws_se.column_dimensions['A'].width = 3
ws_se.column_dimensions['B'].width = 5
ws_se.column_dimensions['C'].width = 55
ws_se.column_dimensions['D'].width = 20

r = 1
style_cell(ws_se, r, 2, 'Schedule SE — Self-Employment Tax', TITLE_FONT); r += 2

section_header(ws_se, r, 2, 4, 'SELF-EMPLOYMENT TAX CALCULATION'); r += 1
style_cell(ws_se, r, 2, '1', BLACK_BOLD, LIGHT_GRAY, CENTER, THIN_BORDER)
style_cell(ws_se, r, 3, 'Net farm profit (loss) — Schedule F', BLACK, LIGHT_GRAY, LEFT, THIN_BORDER)
style_cell(ws_se, r, 4, None, BLUE_INPUT, LIGHT_YELLOW, RIGHT, THIN_BORDER, CURR_FMT)
SE_FARM = r; r += 1

style_cell(ws_se, r, 2, '2', BLACK_BOLD, LIGHT_GRAY, CENTER, THIN_BORDER)
style_cell(ws_se, r, 3, 'Net profit (loss) from Schedule C', BLACK, LIGHT_GRAY, LEFT, THIN_BORDER)
style_cell(ws_se, r, 4, "='Schedule C'!D45", GREEN_LINK, LIGHT_BLUE, RIGHT, THIN_BORDER, CURR_FMT)
SE_BIZ = r; r += 1

style_cell(ws_se, r, 2, '3', BLACK_BOLD, LIGHT_GRAY, CENTER, THIN_BORDER)
style_cell(ws_se, r, 3, 'Combined (lines 1 + 2)', BLACK, LIGHT_GRAY, LEFT, THIN_BORDER)
style_cell(ws_se, r, 4, f'=D{SE_FARM}+D{SE_BIZ}', BLACK, LIGHT_BLUE, RIGHT, THIN_BORDER, CURR_FMT)
SE_COMB = r; r += 1

style_cell(ws_se, r, 2, '4', BLACK_BOLD, LIGHT_GRAY, CENTER, THIN_BORDER)
style_cell(ws_se, r, 3, '92.35% of line 3 (net earnings subject to SE tax)', BLACK, LIGHT_GRAY, LEFT, THIN_BORDER)
style_cell(ws_se, r, 4, f'=D{SE_COMB}*0.9235', BLACK, LIGHT_BLUE, RIGHT, THIN_BORDER, CURR_FMT)
SE_NET_EARN = r; r += 1

style_cell(ws_se, r, 2, '5', BLACK_BOLD, LIGHT_GRAY, CENTER, THIN_BORDER)
style_cell(ws_se, r, 3, 'Social Security wage base limit ($176,100)', BLACK, LIGHT_GRAY, LEFT, THIN_BORDER)
style_cell(ws_se, r, 4, 176100, BLACK, None, RIGHT, THIN_BORDER, CURR_FMT)
SE_LIMIT = r; r += 1

style_cell(ws_se, r, 2, '6', BLACK_BOLD, LIGHT_GRAY, CENTER, THIN_BORDER)
style_cell(ws_se, r, 3, 'W-2 Social Security wages (reduces SE base)', BLACK, LIGHT_GRAY, LEFT, THIN_BORDER)
style_cell(ws_se, r, 4, f"=Intake!D{TOTAL_WAGES_ROW}", GREEN_LINK, LIGHT_BLUE, RIGHT, THIN_BORDER, CURR_FMT)
SE_W2_SS = r; r += 1

style_cell(ws_se, r, 2, '7', BLACK_BOLD, LIGHT_GRAY, CENTER, THIN_BORDER)
style_cell(ws_se, r, 3, 'Remaining SS wage room', BLACK, LIGHT_GRAY, LEFT, THIN_BORDER)
style_cell(ws_se, r, 4, f'=MAX(0,D{SE_LIMIT}-D{SE_W2_SS})', BLACK, LIGHT_BLUE, RIGHT, THIN_BORDER, CURR_FMT)
SE_ROOM = r; r += 1

style_cell(ws_se, r, 2, '8', BLACK_BOLD, LIGHT_GRAY, CENTER, THIN_BORDER)
style_cell(ws_se, r, 3, 'Social Security portion (12.4% on lesser of net earnings or room)', BLACK, LIGHT_GRAY, LEFT, THIN_BORDER)
style_cell(ws_se, r, 4, f'=MIN(D{SE_NET_EARN},D{SE_ROOM})*0.124', BLACK, LIGHT_BLUE, RIGHT, THIN_BORDER, CURR_FMT)
SE_SS = r; r += 1

style_cell(ws_se, r, 2, '9', BLACK_BOLD, LIGHT_GRAY, CENTER, THIN_BORDER)
style_cell(ws_se, r, 3, 'Medicare portion (2.9% on all net earnings)', BLACK, LIGHT_GRAY, LEFT, THIN_BORDER)
style_cell(ws_se, r, 4, f'=D{SE_NET_EARN}*0.029', BLACK, LIGHT_BLUE, RIGHT, THIN_BORDER, CURR_FMT)
SE_MED = r; r += 1

# Row 18 = Total SE Tax
style_cell(ws_se, r, 2, '10', BLACK_BOLD, LIGHT_GRAY, CENTER, THICK_BOTTOM)
style_cell(ws_se, r, 3, 'TOTAL SELF-EMPLOYMENT TAX → Schedule 2', BLACK_BOLD, LIGHT_GRAY, LEFT, THICK_BOTTOM)
style_cell(ws_se, r, 4, f'=D{SE_SS}+D{SE_MED}', BLACK_BOLD, LIGHT_GREEN, RIGHT, THICK_BOTTOM, CURR_FMT)
SE_TOTAL = r
# Ensure row 18
while r < 18:
    r += 1
style_cell(ws_se, 18, 2, '', BLACK_BOLD, LIGHT_GRAY, CENTER, THICK_BOTTOM)
style_cell(ws_se, 18, 3, 'Self-Employment Tax (output row)', BLACK_BOLD, LIGHT_GRAY, LEFT, THICK_BOTTOM)
style_cell(ws_se, 18, 4, f'=D{SE_TOTAL}', BLACK_BOLD, LIGHT_GREEN, RIGHT, THICK_BOTTOM, CURR_FMT)
r = 20
style_cell(ws_se, 20, 2, '11', BLACK_BOLD, LIGHT_GRAY, CENTER, THICK_BOTTOM)
style_cell(ws_se, 20, 3, 'Deductible part of SE tax (50%) → Schedule 1', BLACK_BOLD, LIGHT_GRAY, LEFT, THICK_BOTTOM)
style_cell(ws_se, 20, 4, f'=D{SE_TOTAL}*0.5', BLACK_BOLD, LIGHT_GREEN, RIGHT, THICK_BOTTOM, CURR_FMT)

print("Schedule SE complete.")


# ═══════════════════════════════════════════════════════════════════════════════
# TAB: FORM 8938 — FATCA
# ═══════════════════════════════════════════════════════════════════════════════
ws_fatca = wb.create_sheet('Form 8938')
ws_fatca.sheet_properties.tabColor = '7030A0'
ws_fatca.column_dimensions['A'].width = 3
ws_fatca.column_dimensions['B'].width = 5
ws_fatca.column_dimensions['C'].width = 45
ws_fatca.column_dimensions['D'].width = 20
ws_fatca.column_dimensions['E'].width = 20
ws_fatca.column_dimensions['F'].width = 20

r = 1
style_cell(ws_fatca, r, 2, 'Form 8938 — Statement of Specified Foreign Financial Assets', TITLE_FONT); r += 2

section_header(ws_fatca, r, 2, 6, 'FILING THRESHOLD DETERMINATION'); r += 1
style_cell(ws_fatca, r, 3, 'Filing Status', BLACK, LIGHT_GRAY, LEFT, THIN_BORDER)
style_cell(ws_fatca, r, 4, "=Intake!C6", GREEN_LINK, LIGHT_BLUE, CENTER, THIN_BORDER)
F8938_FS = r; r += 1
style_cell(ws_fatca, r, 3, 'Taxpayer lives abroad? (Y/N)', BLACK, LIGHT_GRAY, LEFT, THIN_BORDER)
style_cell(ws_fatca, r, 4, None, BLUE_INPUT, LIGHT_YELLOW, CENTER, THIN_BORDER)
F8938_ABROAD = r; r += 1
style_cell(ws_fatca, r, 3, 'Maximum aggregate value during year', BLACK, LIGHT_GRAY, LEFT, THIN_BORDER)
style_cell(ws_fatca, r, 4, f"=Intake!F{FA_START-3}", GREEN_LINK, LIGHT_BLUE, RIGHT, THIN_BORDER, CURR_FMT)
F8938_MAX = r; r += 1
style_cell(ws_fatca, r, 3, 'Year-end aggregate value', BLACK, LIGHT_GRAY, LEFT, THIN_BORDER)
style_cell(ws_fatca, r, 4, f"=SUM(Intake!G{FA_START}:G{FA_END})", GREEN_LINK, LIGHT_BLUE, RIGHT, THIN_BORDER, CURR_FMT)
F8938_YE = r; r += 1

style_cell(ws_fatca, r, 3, 'Filing threshold (year-end)', BLACK, LIGHT_GRAY, LEFT, THIN_BORDER)
style_cell(ws_fatca, r, 4,
    f'=IF(D{F8938_ABROAD}="Y",'
    f'IF(OR(D{F8938_FS}=1,D{F8938_FS}=3),200000,IF(D{F8938_FS}=2,400000,200000)),'
    f'IF(OR(D{F8938_FS}=1,D{F8938_FS}=3),50000,IF(D{F8938_FS}=2,100000,50000)))',
    BLACK, LIGHT_BLUE, RIGHT, THIN_BORDER, CURR_FMT)
F8938_THRESH_YE = r; r += 1

style_cell(ws_fatca, r, 3, 'Filing threshold (any time during year)', BLACK, LIGHT_GRAY, LEFT, THIN_BORDER)
style_cell(ws_fatca, r, 4,
    f'=IF(D{F8938_ABROAD}="Y",'
    f'IF(OR(D{F8938_FS}=1,D{F8938_FS}=3),300000,IF(D{F8938_FS}=2,600000,300000)),'
    f'IF(OR(D{F8938_FS}=1,D{F8938_FS}=3),75000,IF(D{F8938_FS}=2,150000,75000)))',
    BLACK, LIGHT_BLUE, RIGHT, THIN_BORDER, CURR_FMT)
F8938_THRESH_MAX = r; r += 1

style_cell(ws_fatca, r, 3, 'FORM 8938 REQUIRED?', BLACK_BOLD, LIGHT_GRAY, LEFT, THICK_BOTTOM)
style_cell(ws_fatca, r, 4,
    f'=IF(OR(D{F8938_YE}>D{F8938_THRESH_YE},D{F8938_MAX}>D{F8938_THRESH_MAX}),"YES - FILING REQUIRED","NO")',
    Font(name='Arial', size=10, color='FF0000', bold=True), LIGHT_RED, CENTER, THICK_BOTTOM)
r += 2

section_header(ws_fatca, r, 2, 6, 'PART I — FOREIGN DEPOSIT AND CUSTODIAL ACCOUNTS'); r += 1
style_cell(ws_fatca, r, 3, '(Pulls from Intake Section 16)', SMALL_FONT); r += 1
for i, h in enumerate(['Account Type', 'Institution', 'Country', 'Max Value', 'Year-End Value'], 2):
    style_cell(ws_fatca, r, i, h, HEADER_FONT, MEDIUM_BLUE, CENTER, THIN_BORDER)
r += 1
for row_i in range(10):
    intake_r = FA_START + row_i
    style_cell(ws_fatca, r, 2, f"=Intake!B{intake_r}", GREEN_LINK, LIGHT_BLUE, LEFT, THIN_BORDER)
    style_cell(ws_fatca, r, 3, f"=Intake!C{intake_r}", GREEN_LINK, LIGHT_BLUE, LEFT, THIN_BORDER)
    style_cell(ws_fatca, r, 4, f"=Intake!D{intake_r}", GREEN_LINK, LIGHT_BLUE, LEFT, THIN_BORDER)
    style_cell(ws_fatca, r, 5, f"=Intake!F{intake_r}", GREEN_LINK, LIGHT_BLUE, RIGHT, THIN_BORDER, CURR_FMT)
    style_cell(ws_fatca, r, 6, f"=Intake!G{intake_r}", GREEN_LINK, LIGHT_BLUE, RIGHT, THIN_BORDER, CURR_FMT)
    r += 1

r += 1
section_header(ws_fatca, r, 2, 6, 'PART IV — SUMMARY OF TAX ITEMS'); r += 1
style_cell(ws_fatca, r, 3, 'Total interest from foreign accounts', BLACK, LIGHT_GRAY, LEFT, THIN_BORDER)
style_cell(ws_fatca, r, 4, f"=SUM(Intake!H{FA_START}:H{FA_END})", GREEN_LINK, LIGHT_BLUE, RIGHT, THIN_BORDER, CURR_FMT)
r += 1
style_cell(ws_fatca, r, 3, 'Total dividends from foreign accounts', BLACK, LIGHT_GRAY, LEFT, THIN_BORDER)
style_cell(ws_fatca, r, 4, None, BLUE_INPUT, LIGHT_YELLOW, RIGHT, THIN_BORDER, CURR_FMT)
r += 1
style_cell(ws_fatca, r, 3, 'Total gains from foreign accounts', BLACK, LIGHT_GRAY, LEFT, THIN_BORDER)
style_cell(ws_fatca, r, 4, None, BLUE_INPUT, LIGHT_YELLOW, RIGHT, THIN_BORDER, CURR_FMT)

print("Form 8938 complete.")


# ═══════════════════════════════════════════════════════════════════════════════
# TAB: FORM 8621 — PFIC
# ═══════════════════════════════════════════════════════════════════════════════
ws_pfic = wb.create_sheet('Form 8621')
ws_pfic.sheet_properties.tabColor = '548235'
ws_pfic.column_dimensions['A'].width = 3
ws_pfic.column_dimensions['B'].width = 5
ws_pfic.column_dimensions['C'].width = 40
ws_pfic.column_dimensions['D'].width = 20
ws_pfic.column_dimensions['E'].width = 20

r = 1
style_cell(ws_pfic, r, 2, 'Form 8621 — Information Return by a Shareholder of a PFIC', TITLE_FONT); r += 2

section_header(ws_pfic, r, 2, 5, 'PFIC INFORMATION (from Intake Section 18)'); r += 1
for i, h in enumerate(['PFIC Name', 'Country', 'EIN/Ref ID', 'FMV (Year-End)'], 2):
    style_cell(ws_pfic, r, i, h, HEADER_FONT, MEDIUM_BLUE, CENTER, THIN_BORDER)
r += 1
for row_i in range(5):
    intake_r = PFIC_START + row_i
    style_cell(ws_pfic, r, 2, f"=Intake!B{intake_r}", GREEN_LINK, LIGHT_BLUE, LEFT, THIN_BORDER)
    style_cell(ws_pfic, r, 3, f"=Intake!C{intake_r}", GREEN_LINK, LIGHT_BLUE, LEFT, THIN_BORDER)
    style_cell(ws_pfic, r, 4, f"=Intake!D{intake_r}", GREEN_LINK, LIGHT_BLUE, LEFT, THIN_BORDER)
    style_cell(ws_pfic, r, 5, f"=Intake!G{intake_r}", GREEN_LINK, LIGHT_BLUE, RIGHT, THIN_BORDER, CURR_FMT)
    r += 1
r += 1

section_header(ws_pfic, r, 2, 5, 'ELECTION AND INCOME DETAILS'); r += 1
style_cell(ws_pfic, r, 3, 'Note: Complete one Form 8621 per PFIC. This is a summary worksheet.', SMALL_FONT); r += 1
for i, h in enumerate(['PFIC Name', 'Election Type', 'Ordinary Income', 'Net Capital Gain'], 2):
    style_cell(ws_pfic, r, i, h, HEADER_FONT, MEDIUM_BLUE, CENTER, THIN_BORDER)
r += 1
for row_i in range(5):
    intake_r = PFIC_START + row_i
    style_cell(ws_pfic, r, 2, f"=Intake!B{intake_r}", GREEN_LINK, LIGHT_BLUE, LEFT, THIN_BORDER)
    style_cell(ws_pfic, r, 3, f"=Intake!H{intake_r}", GREEN_LINK, LIGHT_BLUE, LEFT, THIN_BORDER)
    style_cell(ws_pfic, r, 4, None, BLUE_INPUT, LIGHT_YELLOW, RIGHT, THIN_BORDER, CURR_FMT)
    style_cell(ws_pfic, r, 5, None, BLUE_INPUT, LIGHT_YELLOW, RIGHT, THIN_BORDER, CURR_FMT)
    r += 1

print("Form 8621 complete.")


# ═══════════════════════════════════════════════════════════════════════════════
# TAB: FORM 3520 — Foreign Trusts and Gifts
# ═══════════════════════════════════════════════════════════════════════════════
ws_3520 = wb.create_sheet('Form 3520')
ws_3520.sheet_properties.tabColor = 'FF0066'
ws_3520.column_dimensions['A'].width = 3
ws_3520.column_dimensions['B'].width = 5
ws_3520.column_dimensions['C'].width = 50
ws_3520.column_dimensions['D'].width = 20

r = 1
style_cell(ws_3520, r, 2, 'Form 3520 — Annual Return to Report Transactions with Foreign Trusts', TITLE_FONT); r += 1
style_cell(ws_3520, r, 2, 'and Receipt of Certain Foreign Gifts', SUBTITLE_FONT); r += 2

section_header(ws_3520, r, 2, 4, 'PART IV — US PERSONS RECEIVING FOREIGN GIFTS'); r += 1

# Threshold determination
f3520_start = r
style_cell(ws_3520, r, 3, 'Total gifts from foreign individuals', BLACK, LIGHT_GRAY, LEFT, THIN_BORDER)
style_cell(ws_3520, r, 4, f"=Intake!C{FA_END+4}", GREEN_LINK, LIGHT_BLUE, RIGHT, THIN_BORDER, CURR_FMT)
F3520_IND = r; r += 1
style_cell(ws_3520, r, 3, 'Total gifts from foreign corporations', BLACK, LIGHT_GRAY, LEFT, THIN_BORDER)
style_cell(ws_3520, r, 4, f"=Intake!C{FA_END+5}", GREEN_LINK, LIGHT_BLUE, RIGHT, THIN_BORDER, CURR_FMT)
F3520_CORP = r; r += 1
style_cell(ws_3520, r, 3, 'FORM 3520 REQUIRED FOR GIFTS?', BLACK_BOLD, LIGHT_GRAY, LEFT, THICK_BOTTOM)
style_cell(ws_3520, r, 4,
    f'=IF(OR(D{F3520_IND}>100000,D{F3520_CORP}>19570),"YES - FILING REQUIRED","NO")',
    Font(name='Arial', size=10, color='FF0000', bold=True), LIGHT_RED, CENTER, THICK_BOTTOM)
r += 2

section_header(ws_3520, r, 2, 4, 'PART III — US PERSONS WITH FOREIGN TRUSTS'); r += 1
trust_fields = [
    ('Name of foreign trust', f"=Intake!C{FA_END+6}"),
    ('Country of trust', f"=Intake!C{FA_END+7}"),
    ('EIN or Reference ID', f"=Intake!C{FA_END+8}"),
    ('Distributions received from trust', f"=Intake!C{FA_END+9}"),
    ('Transfers to trust during year', f"=Intake!C{FA_END+10}"),
]
for desc, formula in trust_fields:
    style_cell(ws_3520, r, 3, desc, BLACK, LIGHT_GRAY, LEFT, THIN_BORDER)
    fmt = CURR_FMT if 'Distributions' in desc or 'Transfers' in desc else None
    style_cell(ws_3520, r, 4, formula, GREEN_LINK, LIGHT_BLUE, RIGHT if fmt else LEFT, THIN_BORDER, fmt)
    r += 1
r += 1
style_cell(ws_3520, r, 3, 'FORM 3520 REQUIRED FOR TRUST?', BLACK_BOLD, LIGHT_GRAY, LEFT, THICK_BOTTOM)
style_cell(ws_3520, r, 4,
    f'=IF(OR(D{r-2}<>0,D{r-1}<>0),"YES - FILING REQUIRED","NO")',
    Font(name='Arial', size=10, color='FF0000', bold=True), LIGHT_RED, CENTER, THICK_BOTTOM)

print("Form 3520 complete.")


# ═══════════════════════════════════════════════════════════════════════════════
# TAB: SUMMARY DASHBOARD
# ═══════════════════════════════════════════════════════════════════════════════
ws_dash = wb.create_sheet('Dashboard')
ws_dash.sheet_properties.tabColor = '002060'
wb.move_sheet('Dashboard', offset=-len(wb.sheetnames)+2)  # Move after Intake
ws_dash.column_dimensions['A'].width = 3
ws_dash.column_dimensions['B'].width = 35
ws_dash.column_dimensions['C'].width = 20
ws_dash.column_dimensions['D'].width = 25
ws_dash.column_dimensions['E'].width = 20

r = 1
style_cell(ws_dash, r, 2, 'TAX RETURN SUMMARY DASHBOARD', TITLE_FONT); r += 1
style_cell(ws_dash, r, 2, 'Tax Year 2025 — All amounts auto-calculated', SMALL_FONT); r += 2

section_header(ws_dash, r, 2, 3, 'TAXPAYER INFORMATION'); r += 1
style_cell(ws_dash, r, 2, 'Name', BLACK, LIGHT_GRAY, LEFT, THIN_BORDER)
style_cell(ws_dash, r, 3, '=Intake!C13&" "&Intake!C15', GREEN_LINK, LIGHT_BLUE, LEFT, THIN_BORDER)
r += 1
style_cell(ws_dash, r, 2, 'Filing Status', BLACK, LIGHT_GRAY, LEFT, THIN_BORDER)
style_cell(ws_dash, r, 3,
    '=IF(Intake!C6=1,"Single",IF(Intake!C6=2,"Married Filing Jointly",IF(Intake!C6=3,"Married Filing Separately",IF(Intake!C6=4,"Head of Household","Qualifying Surviving Spouse"))))',
    GREEN_LINK, LIGHT_BLUE, LEFT, THIN_BORDER)
r += 2

section_header(ws_dash, r, 2, 3, 'INCOME SUMMARY'); r += 1
dash_inc = [
    ('Wages & Salaries', f"='Form 1040'!D{LINE_ROWS['1z']}"),
    ('Taxable Interest', f"='Form 1040'!D{LINE_ROWS['2b']}"),
    ('Ordinary Dividends', f"='Form 1040'!D{LINE_ROWS['3b']}"),
    ('Capital Gains/(Losses)', f"='Form 1040'!D{LINE_ROWS['7']}"),
    ('Business Income (Sch C)', "='Schedule C'!D45"),
    ('Rental Income (Sch E)', "='Schedule E'!D50"),
    ('Other Income (Sch 1)', f"='Form 1040'!D{LINE_ROWS['8']}"),
    ('TOTAL INCOME', f"='Form 1040'!D{LINE_ROWS['9']}"),
]
for label, formula in dash_inc:
    bold = 'TOTAL' in label
    style_cell(ws_dash, r, 2, label, BLACK_BOLD if bold else BLACK, LIGHT_GRAY, LEFT, THICK_BOTTOM if bold else THIN_BORDER)
    style_cell(ws_dash, r, 3, formula, GREEN_LINK if not bold else BLACK_BOLD,
               LIGHT_GREEN if bold else LIGHT_BLUE, RIGHT, THICK_BOTTOM if bold else THIN_BORDER, CURR_FMT)
    r += 1
r += 1

section_header(ws_dash, r, 2, 3, 'TAX COMPUTATION SUMMARY'); r += 1
dash_tax = [
    ('Adjusted Gross Income', f"='Form 1040'!D{LINE_ROWS['11']}"),
    ('Deductions', f"='Form 1040'!D{LINE_ROWS['14']}"),
    ('Taxable Income', f"='Form 1040'!D{LINE_ROWS['15']}"),
    ('Income Tax', f"='Form 1040'!D{LINE_ROWS['16']}"),
    ('Self-Employment Tax', "='Schedule SE'!D18"),
    ('Total Tax', f"='Form 1040'!D{LINE_ROWS['22']}"),
    ('Total Payments', f"='Form 1040'!D{LINE_ROWS['33']}"),
    ('REFUND', f"='Form 1040'!D{LINE_ROWS['34']}"),
    ('AMOUNT OWED', f"='Form 1040'!D{LINE_ROWS['37']}"),
]
for label, formula in dash_tax:
    is_key = label in ['REFUND', 'AMOUNT OWED']
    bold = is_key or 'Total' in label
    fill = LIGHT_GREEN if label == 'REFUND' else (LIGHT_RED if label == 'AMOUNT OWED' else (LIGHT_BLUE if not bold else LIGHT_GREEN))
    style_cell(ws_dash, r, 2, label, BLACK_BOLD if bold else BLACK, LIGHT_GRAY, LEFT, THICK_BOTTOM if is_key else THIN_BORDER)
    style_cell(ws_dash, r, 3, formula,
               Font(name='Arial', size=12, color='008000' if label == 'REFUND' else ('FF0000' if label == 'AMOUNT OWED' else '000000'), bold=bold),
               fill, RIGHT, THICK_BOTTOM if is_key else THIN_BORDER, CURR_FMT)
    r += 1
r += 1

section_header(ws_dash, r, 2, 3, 'FOREIGN REPORTING REQUIREMENTS'); r += 1
style_cell(ws_dash, r, 2, 'FBAR Required?', BLACK, LIGHT_GRAY, LEFT, THIN_BORDER)
style_cell(ws_dash, r, 3, f"=IF(Intake!F{FA_START-3}>10000,\"YES\",\"NO\")", GREEN_LINK, LIGHT_BLUE, CENTER, THIN_BORDER)
r += 1
style_cell(ws_dash, r, 2, 'Form 8938 (FATCA) Required?', BLACK, LIGHT_GRAY, LEFT, THIN_BORDER)
style_cell(ws_dash, r, 3, f"='Form 8938'!D{F8938_THRESH_YE+2}", GREEN_LINK, LIGHT_BLUE, CENTER, THIN_BORDER)
r += 1
style_cell(ws_dash, r, 2, 'Form 3520 Required?', BLACK, LIGHT_GRAY, LEFT, THIN_BORDER)
style_cell(ws_dash, r, 3, f"=IF(OR('Form 3520'!D{F3520_IND}>100000,'Form 3520'!D{F3520_CORP}>19570),\"YES\",\"NO\")", GREEN_LINK, LIGHT_BLUE, CENTER, THIN_BORDER)
r += 1
style_cell(ws_dash, r, 2, 'Form 8621 (PFIC) Required?', BLACK, LIGHT_GRAY, LEFT, THIN_BORDER)
style_cell(ws_dash, r, 3, f"=IF(COUNTA(Intake!B{PFIC_START}:B{PFIC_END})>0,\"YES\",\"NO\")", GREEN_LINK, LIGHT_BLUE, CENTER, THIN_BORDER)

print("Dashboard complete.")

# ═══════════════════════════════════════════════════════════════════════════════
# SAVE
# ═══════════════════════════════════════════════════════════════════════════════
OUTPUT = '/sessions/serene-upbeat-einstein/mnt/Tax templates/1040_Tax_Return_2025.xlsx'
wb.save(OUTPUT)
print(f"\nWorkbook saved to: {OUTPUT}")
print(f"Tabs: {wb.sheetnames}")
